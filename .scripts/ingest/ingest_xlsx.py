#!/usr/bin/env python3
"""Excel workbook structure extraction for ingest.py."""

from __future__ import annotations

from pathlib import Path
import html


def _load_workbook(path: Path, data_only: bool):
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise ImportError("openpyxl is required for XLSX ingest") from exc
    return load_workbook(filename=str(path), data_only=data_only, read_only=False)


def _safe(value) -> str:
    if value is None:
        return ""
    return html.escape(str(value).replace("\n", " ").strip())


def workbook_to_markdown(path: Path, max_rows_per_sheet: int = 80) -> str:
    formula_book = _load_workbook(path, data_only=False)
    value_book = _load_workbook(path, data_only=True)

    parts = [f"# {path.name}", "", "## Workbook Map", ""]
    parts.append("| Sheet | Max row | Max column | Hidden | Formula cells |")
    parts.append("|---|---:|---:|---|---:|")

    formula_counts = {}
    for sheet in formula_book.worksheets:
        formula_count = 0
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
        formula_counts[sheet.title] = formula_count
        parts.append(f"| {_safe(sheet.title)} | {sheet.max_row} | {sheet.max_column} | {sheet.sheet_state} | {formula_count} |")

    for sheet in formula_book.worksheets:
        value_sheet = value_book[sheet.title]
        parts.extend(["", f"## Sheet: {sheet.title}", "", "### Preview", ""])
        max_row = min(sheet.max_row, max_rows_per_sheet)
        max_col = min(sheet.max_column, 20)
        if max_row == 0 or max_col == 0:
            parts.append("[empty sheet]")
            continue

        headers = [f"C{col}" for col in range(1, max_col + 1)]
        parts.append("| " + " | ".join(headers) + " |")
        parts.append("| " + " | ".join("---" for _ in headers) + " |")
        for row_idx in range(1, max_row + 1):
            row_values = []
            for col_idx in range(1, max_col + 1):
                formula_value = sheet.cell(row=row_idx, column=col_idx).value
                data_value = value_sheet.cell(row=row_idx, column=col_idx).value
                if isinstance(formula_value, str) and formula_value.startswith("="):
                    row_values.append(_safe(f"{data_value} [{formula_value}]"))
                else:
                    row_values.append(_safe(data_value if data_value is not None else formula_value))
            parts.append("| " + " | ".join(row_values) + " |")

        formula_cells = []
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_cells.append(f"- {cell.coordinate}: `{cell.value}`")
                if len(formula_cells) >= 80:
                    break
            if len(formula_cells) >= 80:
                break
        if formula_cells:
            parts.extend(["", "### Formula Topology", ""])
            parts.extend(formula_cells)
            if formula_counts[sheet.title] > len(formula_cells):
                parts.append(f"- [... {formula_counts[sheet.title] - len(formula_cells)} more formula cells omitted]")

    return "\n".join(parts) + "\n"
