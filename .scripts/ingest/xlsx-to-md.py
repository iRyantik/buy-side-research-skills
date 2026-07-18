#!/usr/bin/env python3
"""xlsx-to-md.py — Convert Excel workbook to markdown tables.

Usage:
    python xlsx-to-md.py <input.xlsx> [--output <path>] [--all-sheets]

One markdown table per sheet.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path


def convert(input_path: Path, output_path: Path) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(input_path, data_only=True)
    parts = [f"# {input_path.name}", ""]
    sheet_count = 0

    for name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row == 0:
            continue
        sheet_count += 1
        parts.append(f"## {name}")
        parts.append("")

        # Build markdown table from rows
        table = []
        for row in ws.iter_rows(max_row=min(ws.max_row, 500), values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            table.append(cells)

        if not table:
            continue

        # Write table
        max_cols = max(len(r) for r in table)
        for i, row in enumerate(table):
            while len(row) < max_cols:
                row.append("")
            parts.append("| " + " | ".join(row) + " |")
            if i == 0:
                parts.append("| " + " | ".join("---" for _ in range(max_cols)) + " |")
        parts.append("")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    text = "\n".join(parts)
    output_path.write_text(text, encoding="utf-8")

    return {
        "status": "converted",
        "engine": "openpyxl",
        "sheets": sheet_count,
        "rows": sum(ws.max_row or 0 for ws in wb.worksheets),
        "chars": len(text),
    }


def main():
    parser = argparse.ArgumentParser(description="Convert Excel to markdown tables")
    parser.add_argument("input", help="Excel file path (.xlsx/.xlsm)")
    parser.add_argument("--output", "-o", help="Output markdown path")
    args = parser.parse_args()

    src = Path(args.input)
    if not src.exists():
        print(f"ERROR: file not found: {src}", file=sys.stderr)
        sys.exit(1)

    dst = Path(args.output) if args.output else src.with_suffix(".md")
    result = convert(src, dst)
    print(f"OK {result['sheets']} sheets, {result['rows']} rows, {result['chars']:,} chars → {dst}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
