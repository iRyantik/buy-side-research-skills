#!/usr/bin/env python3
"""xlsx-to-json.py — Convert Excel workbook to structured JSON.

Usage:
    python xlsx-to-json.py <input.xlsx> [--output <path>] [--sheet <name>]

Outputs [{"col": val, ...}, ...] per sheet.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path


def convert(input_path: Path, output_path: Path, sheet_name: str | None = None) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(input_path, data_only=True)
    result = {}
    sheet_count = 0
    total_rows = 0

    sheets = [sheet_name] if sheet_name else wb.sheetnames
    for name in sheets:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        rows = []
        headers = None
        for row in ws.iter_rows(max_row=min(ws.max_row, 1000), values_only=True):
            cells = [cell if cell is not None else None for cell in row]
            if headers is None:
                headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(cells)]
                continue
            if not any(c is not None for c in cells):
                continue
            rows.append({headers[i]: cells[i] for i in range(min(len(headers), len(cells)))})

        result[name] = rows
        sheet_count += 1
        total_rows += len(rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    text = json.dumps(result if not sheet_name else result.get(sheet_name, []), indent=2, ensure_ascii=False)
    output_path.write_text(text, encoding="utf-8")

    return {
        "status": "converted",
        "sheets": sheet_count,
        "rows": total_rows,
    }


def main():
    parser = argparse.ArgumentParser(description="Convert Excel to structured JSON")
    parser.add_argument("input", help="Excel file path")
    parser.add_argument("--output", "-o", help="Output JSON path")
    parser.add_argument("--sheet", help="Specific sheet name (default: all sheets)")
    args = parser.parse_args()

    src = Path(args.input)
    if not src.exists():
        print(f"ERROR: file not found: {src}", file=sys.stderr)
        sys.exit(1)

    dst = Path(args.output) if args.output else src.with_suffix(".json")
    result = convert(src, dst, args.sheet)
    print(f"OK {result['sheets']} sheets, {result['rows']} rows → {dst}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
