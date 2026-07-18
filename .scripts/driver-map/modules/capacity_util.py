"""capacity_util module — Capacity × Utilization × ASP revenue model.

Contract:
  render(ws, R, ll, anchor_info, ctx) -> dict

  Revenue = Capacity × Utilization% × ASP / unit_scale
  No explicit Volume — Volume is derived from Capacity × Utilization.
"""
from openpyxl.utils import get_column_letter


def render(ws, R, ll, anchor_info, ctx):
    C = ctx['C']; I = ctx['I']; CF = ctx.get('CF', C); HL = ctx.get('HL', C)
    nf = ctx['nf']; bf = ctx['bf']; itf = ctx['itf']
    NUM = ctx['NUM']; DEC = ctx['DEC']; PCT = ctx['PCT']; INT = ctx['INT']
    DS = ctx['DS']; FY0 = ctx['FY0']; LC = ctx['LC']
    proj_n = ctx['proj_n']; bfyr = ctx['bfyr']

    ln = ll['name']
    _cap_li = ctx.get('_cap_li', lambda fy: 0)
    _util_li = ctx.get('_util_li', lambda fy: 0.85)
    _asp_li = ctx.get('_asp_li', lambda fy, ti=0, sc=None: 0)
    _share_li = ctx.get('_share_li', lambda fy, ti=0: 0)
    _PROJ_FYS = ctx.get('_PROJ_FYS', [])
    scale = ctx.get('_us_li', ll.get('unit_scale', 100))
    tiers = ll.get('tiers', [])
    cap_unit = ll.get('capacity', {}).get('unit', 'units')
    asp_unit = ll.get('asp_unit', '')
    FY0_KEY = f'FY{bfyr}'; FY1_KEY = f'FY{bfyr-1}'; FY2_KEY = f'FY{bfyr-2}'

    # ── Nameplate Capacity ──
    for fy_key, col in [(FY2_KEY, DS), (FY1_KEY, DS + 1)]:
        if _cap_li(fy_key): I(ws, R, col, _cap_li(fy_key), fmt=INT)
        else: C(ws, R, col, '', fmt=INT)
    I(ws, R, FY0, _cap_li(FY0_KEY), fmt=INT)
    for i in range(proj_n):
        if i < len(_PROJ_FYS):
            I(ws, R, FY0 + 1 + i, _cap_li(_PROJ_FYS[i]), fmt=INT)
    C(ws, R, 2, ln, font=bf)
    HL(ws, R, 3, f'Nameplate Capacity ({cap_unit})')
    cap_r = R; R += 1

    # ── Utilization % ──
    for fy_key, col in [(FY2_KEY, DS), (FY1_KEY, DS + 1)]:
        if _util_li(fy_key): I(ws, R, col, _util_li(fy_key), fmt=PCT)
        else: C(ws, R, col, '', fmt=PCT)
    I(ws, R, FY0, _util_li(FY0_KEY), fmt=PCT)
    for i in range(proj_n):
        if i < len(_PROJ_FYS):
            I(ws, R, FY0 + 1 + i, _util_li(_PROJ_FYS[i]), fmt=PCT)
    C(ws, R, 3, '  Utilization %', font=itf)
    util_r = R; R += 1

    # ── Implied Volume = Capacity × Utilization ──
    for ci in range(DS, LC + 1):
        cl = get_column_letter(ci)
        C(ws, R, ci, f'=IFERROR({cl}{cap_r}*{cl}{util_r},"")', fmt=INT)
    C(ws, R, 3, f'  Implied Volume', font=itf)
    vol_r = R; R += 1

    # ── ASP rows (same pattern as vol_asp) ──
    asp_rows = []; share_rows = []
    for t_idx, t in enumerate(tiers):
        tn = t.get('name', f'tier{t_idx}')
        is_last = (t_idx == len(tiers) - 1)

        # Share % (all tiers except last)
        if not is_last:
            for ci in range(DS, DS + 2):
                C(ws, R, ci, '', fmt=PCT)
            I(ws, R, FY0, _share_li(FY0_KEY, t_idx), fmt=PCT)
            for i in range(proj_n):
                if i < len(_PROJ_FYS):
                    I(ws, R, FY0 + 1 + i, _share_li(_PROJ_FYS[i], t_idx), fmt=PCT)
            C(ws, R, 3, f'  {tn} Share %', font=itf)
            share_rows.append(R); R += 1
        else:
            share_rows.append(0)

        # ASP
        for fy_key, col in [(FY2_KEY, DS), (FY1_KEY, DS + 1)]:
            av = _asp_li(fy_key, t_idx)
            if av: I(ws, R, col, av, fmt=DEC)
            else: C(ws, R, col, '', fmt=NUM)
        I(ws, R, FY0, _asp_li(FY0_KEY, t_idx), fmt=DEC)
        for i in range(proj_n):
            if i < len(_PROJ_FYS):
                I(ws, R, FY0 + 1 + i, _asp_li(_PROJ_FYS[i], t_idx), fmt=DEC)
        HL(ws, R, 3, f'  {tn} ASP ({asp_unit})')
        asp_rows.append(R)
        R += 1

    # ── Revenue = Implied Vol × Σ(Share × ASP) / scale ──
    rev_r = R
    for ci in range(DS, FY0):
        C(ws, R, ci, '', fmt=NUM)
    for col_idx in [FY0] + [FY0 + 1 + i for i in range(proj_n)]:
        cl = get_column_letter(col_idx)
        parts = []
        ssum = '+'.join([f'{cl}{sr}' for sr in share_rows if sr > 0])
        for ti in range(len(tiers)):
            ac = f'{cl}{asp_rows[ti]}'
            if ti == len(tiers) - 1:
                parts.append(
                    f'({cl}{vol_r}*(1-({ssum}))*{ac})' if ssum
                    else f'({cl}{vol_r}*{ac})')
            else:
                parts.append(f'({cl}{vol_r}*{cl}{share_rows[ti]}*{ac})')
        C(ws, R, col_idx, '=(' + '+'.join(parts) + ')/' + str(scale), fmt=NUM)
    C(ws, R, 3, 'Revenue')
    R += 1

    # ── Rev YoY ──
    for ci in range(DS, DS + 2):
        C(ws, R, ci, '', fmt=PCT)
    C(ws, R, FY0, f'=IFERROR({get_column_letter(FY0)}{rev_r}/{get_column_letter(FY0-1)}{rev_r}-1,"")', fmt=PCT)
    for ci in range(FY0 + 1, LC + 1):
        cl = get_column_letter(ci); pl = get_column_letter(ci - 1)
        C(ws, R, ci, f'=IFERROR({cl}{rev_r}/{pl}{rev_r}-1,"")', fmt=PCT)
    C(ws, R, 3, 'Rev YoY')
    R += 1

    return {
        'next_R': R,
        'rev_r': rev_r,
        'gm_r': None, 'gp_r': None,
        'vol_r': vol_r, 'cap_r': cap_r, 'util_r': util_r,
        'share_rows': share_rows, 'asp_rows': asp_rows,
        'asp_h_r': asp_rows[0] if asp_rows else 0,
        'module': 'capacity_util',
    }
