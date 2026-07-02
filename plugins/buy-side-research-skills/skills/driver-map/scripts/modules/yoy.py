"""yoy module — Year-over-Year revenue model.

Contract:
  render(ws, R, ll, anchor_info, ctx) -> dict

  Input:  logic_line JSON with module="yoy", yoy {bull, base, bear}
  Output: {next_R, rev_r, gm_r:None, gp_r:None, yb, ybs, ybe, ya, module}
  Caller writes GM/GP after render() returns, fills gm_r/gp_r into dict.
"""

from openpyxl.utils import get_column_letter


def render(ws, R, ll, anchor_info, ctx):
    """Base yoy: Revenue = Section 1 ref, YoY Active (BBE group)."""
    C = ctx['C']; I = ctx['I']; CF = ctx.get('CF', C)
    nf = ctx['nf']; bf = ctx['bf']; itf = ctx['itf']
    NUM = ctx['NUM']; PCT = ctx['PCT']
    DS = ctx['DS']; FY0 = ctx['FY0']; LC = ctx['LC']
    proj_n = ctx['proj_n']

    ln = ll['name']
    yoy = ll['yoy']
    bull = yoy['bull']; base = yoy['base']; bear = yoy['bear']
    s1r, _ = anchor_info.get(ln, (0, 0))

    # ── Revenue (history + FY0 = Section 1 ref) ──
    for col in (DS, DS + 1):
        cl = get_column_letter(col)
        C(ws, R, col, f'={cl}{s1r}', fmt=NUM)
    C(ws, R, FY0, f'={get_column_letter(FY0)}{s1r}', fmt=NUM)
    rev_r = R
    C(ws, R, 2, ln, font=bf)
    C(ws, R, 3, 'Revenue')
    R += 1

    # ── YoY Active (history blank, FY26E+ formula) ──
    ya = R
    for ci in range(DS, FY0):
        C(ws, R, ci, '', fmt=PCT)
    cl_e = get_column_letter(DS + 1); cl_d = get_column_letter(DS)
    C(ws, R, DS + 1, f'=IFERROR({cl_e}{rev_r}/{cl_d}{rev_r}-1,"")', fmt=PCT)
    f0 = get_column_letter(FY0); f_1 = get_column_letter(FY0 - 1)
    C(ws, R, FY0, f'=IFERROR({f0}{rev_r}/{f_1}{rev_r}-1,"")', fmt=PCT)
    for i in range(proj_n):
        C(ws, R, FY0 + 1 + i, 0, fmt=PCT)
    C(ws, R, 3, 'Rev YoY')
    R += 1

    # ── Rev QoQ (right below Rev YoY) ──
    Q_START = ctx.get('Q_START', 0)
    Q_END = ctx.get('Q_END', 0)
    has_q = ctx.get('q_actual_n', 0) + ctx.get('q_proj_n', 0) > 0
    if has_q:
        for ci in range(DS, ctx.get('LC', DS + 5)):
            C(ws, R, ci, '', fmt=PCT)
        for qi in range(Q_START, Q_END + 1):
            cl = get_column_letter(qi); pl = get_column_letter(qi - 1)
            if qi == Q_START: C(ws, R, qi, '', fmt=PCT)
            else: C(ws, R, qi, f'=IFERROR({cl}{rev_r}/{pl}{rev_r}-1,"")', fmt=PCT)
        C(ws, R, 3, '  Rev QoQ', font=itf)
        R += 1

    # ── BBE YoY hidden rows ──
    yb = ys = ye = 0
    for arr, label in [(bull, 'Bull'), (base, 'Base'), (bear, 'Bear')]:
        for ci in range(DS, DS + 2):
            C(ws, R, ci, '', fmt=PCT)
        C(ws, R, FY0, '', fmt=PCT)
        for i, v in enumerate(arr):
            I(ws, R, FY0 + 1 + i, v, fmt=PCT)
        C(ws, R, 3, f'  {label}', font=itf)
        ws.row_dimensions[R].hidden = True
        if label == 'Bull':   yb = R
        elif label == 'Base': ys = R
        elif label == 'Bear': ye = R
        R += 1

    ws.row_dimensions.group(yb, ye, outline_level=1, hidden=True)

    # ── YoY Active formulas + Revenue FY26+ formulas ──
    for i in range(proj_n):
        ci = FY0 + 1 + i
        cl = get_column_letter(ci)
        CF(ws, ya, ci, f'=IF(B1="Bull",{cl}{yb},IF(B1="Bear",{cl}{ye},{cl}{ys}))', fmt=PCT)
        CF(ws, rev_r, FY0 + 1 + i, f'={get_column_letter(FY0 + i)}{rev_r}*(1+{cl}{ya})', fmt=NUM)

    return {
        'next_R': R,
        'rev_r': rev_r,
        'gm_r': None,
        'gp_r': None,
        'yb': yb, 'ybs': ys, 'ybe': ye, 'ya': ya,
        'module': 'yoy',
    }
