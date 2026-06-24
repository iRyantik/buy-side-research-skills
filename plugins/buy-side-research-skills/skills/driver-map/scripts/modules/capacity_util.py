"""capacity_util module — Capacity × Utilization% × ASP revenue model.

Contract:
  render(ws, R, ll, anchor_info, ctx) -> dict

  Input:  logic_line JSON with module="capacity_util"
          capacity: {fy0, proj, unit}
          util_rate: {fy0, proj}  (0-1)
          asp: {fy0, proj, unit}
  Output: {next_R, rev_r, gm_r:None, gp_r:None}
"""

from openpyxl.utils import get_column_letter


def render(ws, R, ll, anchor_info, ctx):
    C = ctx['C']; I = ctx['I']
    nf = ctx['nf']; bf = ctx['bf']; itf = ctx['itf']
    NUM = ctx['NUM']; DEC = ctx['DEC']; PCT = ctx['PCT']
    DS = ctx['DS']; FY0 = ctx['FY0']; LC = ctx['LC']
    proj_n = ctx['proj_n']

    ln = ll['name']
    cap = ll['capacity']; util = ll['util_rate']; asp = ll['asp']

    # ── Capacity ──
    for ci in range(DS, DS + 2):
        C(ws, R, ci, '', fmt=NUM)
    I(ws, R, FY0, cap['fy0'], fmt=DEC)
    for i, v in enumerate(cap['proj']):
        I(ws, R, FY0 + 1 + i, v, fmt=DEC)
    C(ws, R, 2, ln, font=bf)
    C(ws, R, 3, f'Capacity ({cap["unit"]})')
    cap_r = R; R += 1

    # ── Utilization % ──
    for ci in range(DS, DS + 2):
        C(ws, R, ci, '', fmt=PCT)
    I(ws, R, FY0, util['fy0'], fmt=PCT)
    for i, v in enumerate(util['proj']):
        I(ws, R, FY0 + 1 + i, v, fmt=PCT)
    C(ws, R, 3, 'Utilization %')
    util_r = R; R += 1

    # ── ASP ──
    for ci in range(DS, DS + 2):
        C(ws, R, ci, '', fmt=NUM)
    I(ws, R, FY0, asp['fy0'], fmt=DEC)
    for i, v in enumerate(asp['proj']):
        I(ws, R, FY0 + 1 + i, v, fmt=DEC)
    C(ws, R, 3, f'ASP ({asp["unit"]})')
    asp_r = R; R += 1

    # ── Revenue = Capacity × Util% × ASP ──
    rev_r = R
    for col_idx in [FY0] + [FY0 + 1 + i for i in range(proj_n)]:
        cl = get_column_letter(col_idx)
        C(ws, R, col_idx,
          f'={cl}{cap_r}*{cl}{util_r}*{cl}{asp_r}', fmt=NUM)
    for ci in range(DS, DS + 2):
        C(ws, R, ci, '', fmt=NUM)
    C(ws, R, 3, 'Revenue')
    R += 1

    # ── Check row (collapsible) ──
    s1r, s1v = anchor_info.get(ln, (0, 0))
    for ci in range(DS, DS + 2):
        C(ws, R, ci, '', fmt=NUM)
    if s1r:
        C(ws, R, FY0, f'={get_column_letter(FY0)}{s1r}', fmt=NUM)
    for ci in range(FY0 + 1, LC + 1):
        C(ws, R, ci, '', fmt=NUM)
    C(ws, R, 3, f'  Check (anchor {s1v}M)', font=itf)
    ws.row_dimensions.group(R, R, outline_level=1, hidden=True)
    R += 1

    # ── Implied YoY ──
    for ci in range(DS, DS + 2):
        C(ws, R, ci, '', fmt=PCT)
    f0 = get_column_letter(FY0); f_1 = get_column_letter(FY0 - 1)
    C(ws, R, FY0, f'=IFERROR({f0}{rev_r}/{f_1}{rev_r}-1,"")', fmt=PCT)
    for ci in range(FY0 + 1, LC + 1):
        cl = get_column_letter(ci); pl = get_column_letter(ci - 1)
        C(ws, R, ci, f'=IFERROR({cl}{rev_r}/{pl}{rev_r}-1,"")', fmt=PCT)
    C(ws, R, 3, 'Implied YoY')
    R += 1

    return {
        'next_R': R,
        'rev_r': rev_r,
        'gm_r': None,
        'gp_r': None,
        'cap_r': cap_r,
        'util_r': util_r,
        'asp_r': asp_r,
        'module': 'capacity_util',
    }
