"""vol_asp module — Volume × Share% × ASP revenue model.

Contract:
  render(ws, R, ll, anchor_info, ctx) -> dict

  Input:  logic_line JSON with module="vol_asp", volume, tiers
  Output: {next_R, rev_r, vol_r, share_rows, asp_rows}
  Caller writes GM/GP after render() returns, fills gm_r/gp_r into dict.
"""

from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment


def render(ws, R, ll, anchor_info, ctx):
    """Render one vol_asp logic line. Returns row references dict."""
    C = ctx['C']; I = ctx['I']; A = ctx.get('A', C); CF = ctx.get('CF', C); HL = ctx.get('HL', C)
    nf = ctx['nf']; bf = ctx['bf']; itf = ctx['itf']
    NUM = ctx['NUM']; DEC = ctx['DEC']; PCT = ctx['PCT']; INT = ctx['INT']
    DS = ctx['DS']; FY0 = ctx['FY0']; LC = ctx['LC']; SC = ctx['SC']
    proj_n = ctx['proj_n']; bfyr = ctx['bfyr']

    ln = ll['name']
    vol = ll['volume']; tiers = ll['tiers']
    v0 = vol['fy0']; vp = vol['proj']
    cap = ll.get('capacity')
    scale = ll.get('unit_scale', 100)
    asp_unit = ll.get('asp_unit', '万/t')
    # History layer: fy-2 (col DS) and fy-1 (col DS+1)
    hist_all = ll.get('history', {})
    H2 = hist_all.get('fy-2', {}); H1 = hist_all.get('fy-1', {})

    # ── Volume row ──
    for yr_h, col in [(H2, DS), (H1, DS + 1)]:
        if yr_h.get('volume'): I(ws, R, col, yr_h['volume'], fmt=INT)
        else: C(ws, R, col, '', fmt=INT)
    I(ws, R, FY0, v0, fmt=INT)
    for i, v in enumerate(vp):
        I(ws, R, FY0 + 1 + i, v, fmt=INT)
    C(ws, R, 2, ln, font=bf)
    HL(ws, R, 3, f'Volume ({vol["unit"]})')
    vol_r = R; R += 1

    # ── Nameplate Capacity + Utilization (if capacity field exists) ──
    cap_r = 0
    if cap:
        for ci in range(DS, DS + 2):
            C(ws, R, ci, '', fmt=INT)
        I(ws, R, FY0, cap['fy0'], fmt=INT)
        for i, v in enumerate(cap['proj']):
            I(ws, R, FY0 + 1 + i, v, fmt=INT)
        C(ws, R, 3, f'  Nameplate Capacity ({cap["unit"]})', font=itf)
        cap_r = R; R += 1

        # Ramp-up notes as cell comments
        for key, note in cap.get('ramp_notes', {}).items():
            try:
                yr = int(key.replace('fy', ''))
                offset = yr - (bfyr % 100)
                col = FY0 + offset
                if DS <= col <= LC:
                    ws.cell(row=cap_r, column=col).comment = Comment(note, 'Analyst')
            except (ValueError, KeyError):
                pass

        for ci in range(DS, LC + 1):
            cl = get_column_letter(ci)
            C(ws, R, ci, f'=IFERROR({cl}{vol_r}/{cl}{cap_r},"")', fmt=PCT)
        C(ws, R, 3, '  Utilization %', font=itf)
        util_r = R; R += 1

    share_rows = []; asp_rows = []
    tier_bbe = []  # per-tier: (t_idx, active_row, bull_row, base_row, bear_row) or None for non-BBE

    # ── Tier rows ──
    for t_idx, t in enumerate(tiers):
        tn = t['name']
        is_last = (t_idx == len(tiers) - 1)

        # Share % row (every tier except last)
        if not is_last:
            for ci in range(DS, DS + 2):
                C(ws, R, ci, '', fmt=PCT)
            I(ws, R, FY0, t.get('share_fy0', 0), fmt=PCT)
            for i, v in enumerate(t.get('share_proj', [])):
                I(ws, R, FY0 + 1 + i, v, fmt=PCT)
            C(ws, R, 3, f'  {tn} Share %', font=itf)
            share_rows.append(R)
            R += 1
        else:
            share_rows.append(0)

        # ASP row — two paths: BBE (3-scenario) or simple array
        if any(k in t for k in ('asp_bull', 'asp_base', 'asp_bear')):
            # ── BBE ASP ──
            bull = t.get('asp_bull', [])
            base = t.get('asp_base', [])
            bear = t.get('asp_bear', [])

            # ASP history + Active
            asp_h_2 = H2.get(f'{tn}_asp'); asp_h_1 = H1.get(f'{tn}_asp')
            for col, h_val in [(DS, asp_h_2), (DS + 1, asp_h_1)]:
                if h_val: I(ws, R, col, h_val, fmt=DEC)
                else: C(ws, R, col, '', fmt=NUM)
            for ci in range(FY0, LC + 1):
                C(ws, R, ci, 0, fmt=NUM)
            HL(ws, R, 3, f'  {tn} ASP Active ({asp_unit})')
            asp_a_r = R; R += 1

            asp_b_r = asp_bs_r = asp_be_r = 0
            for arr, label in [(bull, 'Bull'), (base, 'Base'), (bear, 'Bear')]:
                for col, h_val in [(DS, asp_h_2), (DS + 1, asp_h_1)]:
                    if h_val: I(ws, R, col, h_val, fmt=DEC)
                    else: C(ws, R, col, '', fmt=NUM)
                I(ws, R, FY0, arr[0] if arr else 0, fmt=DEC)
                for i, v in enumerate(arr[1:] if len(arr) > 1 else []):
                    if i < proj_n:
                        I(ws, R, FY0 + 1 + i, v, fmt=DEC)
                C(ws, R, 3, f'    {label} ({asp_unit})', font=itf)
                ws.row_dimensions[R].hidden = True
                if label == 'Bull':    asp_b_r = R
                elif label == 'Base':  asp_bs_r = R
                elif label == 'Bear':  asp_be_r = R
                R += 1

            ws.row_dimensions.group(asp_b_r, asp_be_r, outline_level=1, hidden=True)

            # ASP Active formula
            f0_val = base[0] if base else 0
            cll = FY0
            C(ws, asp_a_r, cll, f0_val, fmt=DEC)
            for i in range(proj_n):
                ci = FY0 + 1 + i
                cl = get_column_letter(ci)
                CF(ws, asp_a_r, ci,
                   f'=IF(B1="Bull",{cl}{asp_b_r},IF(B1="Bear",{cl}{asp_be_r},{cl}{asp_bs_r}))', fmt=DEC)

            asp_rows.append(asp_a_r)
            tier_bbe.append((t_idx, asp_a_r, asp_b_r, asp_bs_r, asp_be_r))

        else:
            # ── Simple ASP array ──
            ap = t.get('asp', [])
            afy0 = t.get('asp_fy0')
            if afy0 is None and ap:
                afy0 = ap[0]; ap = ap[1:]
            elif afy0 is None:
                afy0 = 0

            asp_h_2 = H2.get(f'{tn}_asp'); asp_h_1 = H1.get(f'{tn}_asp')
            for col, h_val in [(DS, asp_h_2), (DS + 1, asp_h_1)]:
                if h_val: I(ws, R, col, h_val, fmt=DEC)
                else: C(ws, R, col, '', fmt=NUM)
            I(ws, R, FY0, afy0, fmt=DEC)
            for i, v in enumerate(ap):
                if i < proj_n:
                    I(ws, R, FY0 + 1 + i, v, fmt=DEC)
            HL(ws, R, 3, f'  {tn} ASP ({asp_unit})')
            asp_rows.append(R)
            R += 1
            tier_bbe.append(None)

    # ── Revenue formula row ──
    rev_r = R
    # History columns: =Vol × ASP / scale if data exists
    for col, h_key in [(DS, H2), (DS + 1, H1)]:
        cl = get_column_letter(col)
        if h_key.get('volume') and asp_rows:
            C(ws, R, col, f'=({cl}{vol_r}*{cl}{asp_rows[0]})/{scale}', fmt=NUM)
        else:
            C(ws, R, col, '', fmt=NUM)
    for col_idx in [FY0] + [FY0 + 1 + i for i in range(proj_n)]:
        cl = get_column_letter(col_idx)
        parts = []
        ssum = '+'.join([f'{cl}{sr}' for sr in share_rows if sr > 0])
        for ti in range(len(tiers)):
            ac = f'{cl}{asp_rows[ti]}'
            if ti == len(tiers) - 1:
                parts.append(
                    f'({cl}{vol_r}*(1-({ssum}))*{ac})' if ssum
                    else f'({cl}{vol_r}*{ac})')
            else:
                parts.append(f'({cl}{vol_r}*{cl}{share_rows[ti]}*{ac})')
        C(ws, R, col_idx, '=(' + '+'.join(parts) + ')/' + str(scale), fmt=NUM)
    C(ws, R, 3, 'Revenue')
    R += 1

    # ── Scenario Revenue (cached, for Scenario Summary) ──
    sc_cl = get_column_letter(SC)
    yb = ys = ye = 0
    if any(tb is not None for tb in tier_bbe):
        for arr_suffix, arr_idx, label in [('b', 2, 'Bull'), ('bs', 3, 'Base'), ('be', 4, 'Bear')]:
            parts = []
            ssum = '+'.join([f'{sc_cl}{sr}' for sr in share_rows if sr > 0])
            for ti in range(len(tiers)):
                tb = tier_bbe[ti]
                if tb is not None:
                    ac = f'{sc_cl}{tb[arr_idx]}'  # BBE scenario row
                else:
                    ac = f'{sc_cl}{asp_rows[ti]}'  # simple ASP row
                if ti == len(tiers) - 1:
                    parts.append(f'({sc_cl}{vol_r}*(1-({ssum}))*{ac})' if ssum
                                 else f'({sc_cl}{vol_r}*{ac})')
                else:
                    parts.append(f'({sc_cl}{vol_r}*{sc_cl}{share_rows[ti]}*{ac})')
            for ci in range(DS, DS + 2):
                C(ws, R, ci, '', fmt=NUM)
            C(ws, R, 3, f'    {label} Rev @ SOTP', font=itf)
            C(ws, R, SC, '=(' + '+'.join(parts) + ')/' + str(scale), fmt=NUM)
            ws.row_dimensions[R].hidden = True
            if label == 'Bull':    yb = R
            elif label == 'Base':  ys = R
            elif label == 'Bear':  ye = R
            R += 1
        ws.row_dimensions.group(yb, ye, outline_level=1, hidden=True)

    # ── Implied YoY ──
    for ci in range(DS, DS + 2):
        C(ws, R, ci, '', fmt=PCT)
    f0 = get_column_letter(FY0)
    f_1 = get_column_letter(FY0 - 1)
    C(ws, R, FY0, f'=IFERROR({f0}{rev_r}/{f_1}{rev_r}-1,"")', fmt=PCT)
    cl_e = get_column_letter(DS + 1); cl_d = get_column_letter(DS)
    C(ws, R, DS + 1, f'=IFERROR({cl_e}{rev_r}/{cl_d}{rev_r}-1,"")', fmt=PCT)
    for ci in range(FY0 + 1, LC + 1):
        cl = get_column_letter(ci); pl = get_column_letter(ci - 1)
        C(ws, R, ci, f'=IFERROR({cl}{rev_r}/{pl}{rev_r}-1,"")', fmt=PCT)
    C(ws, R, 3, 'Rev YoY')
    R += 1

    # ── Rev QoQ (right below Rev YoY) ──
    Q_START = ctx.get('Q_START', 0)
    Q_END = ctx.get('Q_END', 0)
    has_q = ctx.get('q_actual_n', 0) + ctx.get('q_proj_n', 0) > 0
    if has_q:
        for ci in range(DS, ctx.get('LC', DS + 5)):
            C(ws, R, ci, '', fmt=PCT)
        for qi in range(Q_START, Q_END + 1):
            cl = get_column_letter(qi); pl = get_column_letter(qi - 1)
            if qi == Q_START: C(ws, R, qi, '', fmt=PCT)
            else: C(ws, R, qi, f'=IFERROR({cl}{rev_r}/{pl}{rev_r}-1,"")', fmt=PCT)
        C(ws, R, 3, '  Rev QoQ', font=itf)
        R += 1

    result = {
        'next_R': R,
        'rev_r': rev_r,
        'gm_r': None,
        'gp_r': None,
        'vol_r': vol_r,
        'cap_r': cap_r,
        'util_r': util_r if cap else 0,
        'share_rows': share_rows,
        'asp_rows': asp_rows,
        'module': 'vol_asp',
        # For history Revenue formula: first tier's ASP row (simple or Active)
        'asp_h_r': asp_rows[0] if asp_rows else 0,
    }
    if yb:
        result['yb'] = yb
        result['ybs'] = ys
        result['ybe'] = ye
    return result
