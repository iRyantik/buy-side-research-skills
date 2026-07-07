"""ebitda module — EBITDA margin → EBITDA profit chain.

Contract:
  render(ws, R, ll, ctx) -> dict

  Renders EBITDA margin (assumption) + EBITDA = Rev × margin + Yoy/Y.
  Returns same shape as GM/GP: {next_R, gm_r, gp_r, rev_r, module}
  gm_r = EBITDA margin row, gp_r = EBITDA row.
"""

from openpyxl.utils import get_column_letter


def render(ws, R, ll, ctx):
    C = ctx['C']; I = ctx['I']; CF = ctx.get('CF', C)
    nf = ctx['nf']; itf = ctx['itf']
    NUM = ctx['NUM']; PCT = ctx['PCT']
    DS = ctx['DS']; FY0 = ctx['FY0']; LC = ctx['LC']
    proj_n = ctx['proj_n']; bfyr = ctx['bfyr']

    rev_r = ctx.get('_rev_r', 0)
    _br_li = ctx.get('_br_li', lambda fy: 0.3)
    _PROJ_FYS = ctx.get('_PROJ_FYS', [])
    FY0_KEY = f'FY{bfyr}'; FY1_KEY = f'FY{bfyr-1}'; FY2_KEY = f'FY{bfyr-2}'

    # ── EBITDA margin ──
    for fy_key, col in [(FY2_KEY, DS), (FY1_KEY, DS + 1)]:
        br_val = _br_li(fy_key)
        if br_val: I(ws, R, col, br_val, fmt=PCT)
        else: C(ws, R, col, '', fmt=PCT)
    br_fy0 = _br_li(FY0_KEY)
    if br_fy0: I(ws, R, FY0, br_fy0, fmt=PCT)
    for i in range(proj_n):
        if i < len(_PROJ_FYS):
            I(ws, R, FY0 + 1 + i, _br_li(_PROJ_FYS[i]), fmt=PCT)
    C(ws, R, 3, 'EBITDA margin')
    gm_r = R; R += 1

    # ── EBITDA = Rev × margin ──
    for ci in range(DS, LC + 1):
        cl = get_column_letter(ci)
        C(ws, R, ci, f'=IFERROR({cl}{rev_r}*{cl}{gm_r},"")', fmt=NUM)
    C(ws, R, 3, 'EBITDA')
    gp_r = R; R += 1

    # ── EBITDA YoY ──
    C(ws, R, DS, '', fmt=PCT)
    cl_e = get_column_letter(DS + 1); cl_d = get_column_letter(DS)
    C(ws, R, DS + 1, f'=IFERROR({cl_e}{gp_r}/{cl_d}{gp_r}-1,"")', fmt=PCT)
    f0 = get_column_letter(FY0); f_1 = get_column_letter(FY0 - 1)
    C(ws, R, FY0, f'=IFERROR({f0}{gp_r}/{f_1}{gp_r}-1,"")', fmt=PCT)
    for ci in range(FY0 + 1, LC + 1):
        cl = get_column_letter(ci); pl = get_column_letter(ci - 1)
        C(ws, R, ci, f'=IFERROR({cl}{gp_r}/{pl}{gp_r}-1,"")', fmt=PCT)
    C(ws, R, 3, 'EBITDA YoY')
    R += 1

    # ── EBITDA QoQ ──
    Q_START = ctx.get('Q_START', 0)
    Q_END = ctx.get('Q_END', 0)
    has_q = ctx.get('q_actual_n', 0) + ctx.get('q_proj_n', 0) > 0
    if has_q:
        for ci in range(DS, ctx.get('LC', DS + 5)):
            C(ws, R, ci, '', fmt=PCT)
        for qi in range(Q_START, Q_END + 1):
            cl = get_column_letter(qi); pl = get_column_letter(qi - 1)
            if qi == Q_START: C(ws, R, qi, '', fmt=PCT)
            else: C(ws, R, qi, f'=IFERROR({cl}{gp_r}/{pl}{gp_r}-1,"")', fmt=PCT)
        C(ws, R, 3, '  EBITDA QoQ', font=itf)
        R += 1

    return {
        'next_R': R,
        'gm_r': gm_r,
        'gp_r': gp_r,
        'rev_r': rev_r,
        'module': 'ebitda',
    }
