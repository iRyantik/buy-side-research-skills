"""backlog_burn module — Beginning Backlog × Burn Rate revenue model.

Contract:
  render(ws, R, ll, anchor_info, ctx) -> dict

  Chain:  End_Backlog_t = Beg_Backlog_t × (1 + Order_Rate_t − Burn_Rate_t)
          Beg_Backlog_{t+1} = End_Backlog_t  (cross-column chain link)

  Output: {next_R, rev_r, gm_r:None, gp_r:None, beg_r, end_r, order_r, burn_r, [yb,ybs,ybe]}
"""

from openpyxl.utils import get_column_letter


def _write_bb_rate(ws, R, line_idx, field, label, ctx):
    """Write Order Rate or Burn Rate rows using FY-keyed access. Returns (active_r, next_R, yb, ys, ye)."""
    C = ctx['C']; I = ctx['I']; CF = ctx.get('CF', C)
    nf = ctx['nf']; bf = ctx['bf']; itf = ctx['itf']
    PCT = ctx['PCT']; NUM = ctx['NUM']; DS = ctx['DS']; FY0 = ctx['FY0']; LC = ctx['LC']
    proj_n = ctx['proj_n']; bfyr = ctx['bfyr']
    _bb_li = ctx.get('_bb_li', lambda f, fy, sc=None: 0)
    _PROJ_FYS = ctx.get('_PROJ_FYS', [])
    FY0_KEY = f'FY{bfyr}'
    FY1_KEY = f'FY{bfyr-1}'; FY2_KEY = f'FY{bfyr-2}'

    # Check if BBE (bull/base/bear scenarios exist for this field)
    has_bull = _bb_li(field, FY0_KEY, 'bull') != 0 or any(
        _bb_li(field, fy, 'bull') for fy in _PROJ_FYS)
    is_bbe = has_bull
    yb = ys = ye = 0

    if is_bbe:
        # Active row
        for ci in range(DS, DS + 2):
            C(ws, R, ci, '', fmt=PCT)
        C(ws, R, FY0, '', fmt=PCT)
        for i in range(proj_n):
            C(ws, R, FY0 + 1 + i, 0, fmt=PCT)
        C(ws, R, 3, label)
        active_r = R; R += 1

        # Hidden BBE rows
        for scenario, lbl in [('bull', 'Bull'), ('base', 'Base'), ('bear', 'Bear')]:
            for ci in range(DS, DS + 2):
                C(ws, R, ci, '', fmt=PCT)
            C(ws, R, FY0, '', fmt=PCT)
            I(ws, R, FY0, _bb_li(field, FY0_KEY, scenario), fmt=PCT)
            for i in range(proj_n):
                if i < len(_PROJ_FYS):
                    I(ws, R, FY0 + 1 + i, _bb_li(field, _PROJ_FYS[i], scenario), fmt=PCT)
            C(ws, R, 3, f'    {lbl}', font=itf)
            ws.row_dimensions[R].hidden = True
            if lbl == 'Bull':   yb = R
            elif lbl == 'Base': ys = R
            elif lbl == 'Bear': ye = R
            R += 1

        ws.row_dimensions.group(yb, ye, outline_level=1, hidden=True)

        # Active formula
        for i in range(proj_n):
            ci = FY0 + 1 + i
            cl = get_column_letter(ci)
            CF(ws, active_r, ci,
               f'=IF(B1="Bull",{cl}{yb},IF(B1="Bear",{cl}{ye},{cl}{ys}))', fmt=PCT)
    else:
        # Simple rate — no BBE
        for ci in range(DS, DS + 2):
            C(ws, R, ci, '', fmt=PCT)
        I(ws, R, FY0, _bb_li(field, FY0_KEY), fmt=PCT)
        for i in range(proj_n):
            if i < len(_PROJ_FYS):
                I(ws, R, FY0 + 1 + i, _bb_li(field, _PROJ_FYS[i]), fmt=PCT)
        C(ws, R, 3, label)
        active_r = R; R += 1

    return active_r, R, yb, ys, ye


def render(ws, R, ll, anchor_info, ctx):
    C = ctx['C']; I = ctx['I']; CF = ctx.get('CF', C)
    nf = ctx['nf']; bf = ctx['bf']; itf = ctx['itf']
    NUM = ctx['NUM']; DEC = ctx['DEC']; PCT = ctx['PCT']
    DS = ctx['DS']; FY0 = ctx['FY0']; LC = ctx['LC']; SC = ctx['SC']
    proj_n = ctx['proj_n']; bfyr = ctx['bfyr']

    ln = ll['name']
    li = ctx.get('li', 0)
    _bb_li = ctx.get('_bb_li', lambda f, fy, sc=None: 0)
    _PROJ_FYS = ctx.get('_PROJ_FYS', [])
    FY0_KEY = f'FY{bfyr}'; FY1_KEY = f'FY{bfyr-1}'; FY2_KEY = f'FY{bfyr-2}'

    # BBE check: does order_rate have bull scenario?
    has_bbe = _bb_li('order_rate', FY0_KEY, 'bull') != 0 or any(
        _bb_li('order_rate', fy, 'bull') for fy in _PROJ_FYS)

    # ── Beginning Backlog ──
    for ci in range(DS, DS + 2):
        C(ws, R, ci, '', fmt=NUM)
    I(ws, R, FY0, _bb_li('beg_backlog', FY0_KEY), fmt=NUM)
    beg_unit = ll.get('beg_backlog', {}).get('unit', 'units')
    C(ws, R, 2, ln, font=bf)
    C(ws, R, 3, f'Beg Backlog ({beg_unit})')
    beg_r = R; R += 1

    # ── Order Rate (BBE or simple) ──
    order_r, R, ob, os, oe = _write_bb_rate(ws, R, li, 'order_rate', 'Order Rate (% backlog)', ctx)

    # ── Burn Rate (BBE or simple) ──
    burn_r, R, bb, bs, be = _write_bb_rate(ws, R, li, 'burn_rate', 'Burn Rate (% backlog)', ctx)

    yb = ob if ob else bb
    ys = os if os else bs
    ye = oe if oe else be

    # ── Revenue = Beg Backlog × Burn Rate ──
    rev_r = R
    for col_idx in [FY0] + [FY0 + 1 + i for i in range(proj_n)]:
        cl = get_column_letter(col_idx)
        C(ws, R, col_idx, f'={cl}{beg_r}*{cl}{burn_r}', fmt=NUM)
    for ci in range(DS, DS + 2):
        C(ws, R, ci, '', fmt=NUM)
    C(ws, R, 3, 'Revenue')
    R += 1

    # ── Scenario Revenue cache (if BBE) ──
    if has_bbe:
        sc_cl = get_column_letter(SC)
        for scenario, label in [('bull', 'Bull'), ('base', 'Base'), ('bear', 'Bear')]:
            burn_sc_row = {'bull': bb, 'base': bs, 'bear': be}.get(scenario, bs)
            if burn_sc_row:
                for ci in range(DS, DS + 2):
                    C(ws, R, ci, '', fmt=NUM)
                C(ws, R, 3, f'    {label} Rev @ SOTP', font=itf)
                C(ws, R, SC, f'={sc_cl}{beg_r}*{sc_cl}{burn_sc_row}', fmt=NUM)
                ws.row_dimensions[R].hidden = True
                if label == 'Bull':    yb = R
                elif label == 'Base':  ys = R
                elif label == 'Bear':  ye = R
                R += 1
        ws.row_dimensions.group(yb, ye, outline_level=1, hidden=True)

    # ── End Backlog = Beg × (1 + Order_Rate − Burn_Rate) ──
    end_r = R
    for col_idx in [FY0] + [FY0 + 1 + i for i in range(proj_n)]:
        cl = get_column_letter(col_idx)
        C(ws, R, col_idx,
          f'={cl}{beg_r}*(1+{cl}{order_r}-{cl}{burn_r})', fmt=NUM)
    for ci in range(DS, DS + 2):
        C(ws, R, ci, '', fmt=NUM)
    C(ws, R, 3, 'End Backlog')
    R += 1

    # ── Chain link: Beg Backlog FY26E+ = prior End Backlog ──
    for i in range(proj_n):
        ci = FY0 + 1 + i
        cl = get_column_letter(ci)
        pl = get_column_letter(ci - 1)
        CF(ws, beg_r, ci, f'={pl}{end_r}', fmt=NUM)

    # ── Implied YoY ──
    for ci in range(DS, DS + 2):
        C(ws, R, ci, '', fmt=PCT)
    f0 = get_column_letter(FY0); f_1 = get_column_letter(FY0 - 1)
    C(ws, R, FY0, f'=IFERROR({f0}{rev_r}/{f_1}{rev_r}-1,"")', fmt=PCT)
    for ci in range(FY0 + 1, LC + 1):
        cl = get_column_letter(ci); pl = get_column_letter(ci - 1)
        C(ws, R, ci, f'=IFERROR({cl}{rev_r}/{pl}{rev_r}-1,"")', fmt=PCT)
    C(ws, R, 3, 'Implied YoY')
    R += 1

    result = {
        'next_R': R,
        'rev_r': rev_r,
        'gm_r': None,
        'gp_r': None,
        'beg_r': beg_r,
        'end_r': end_r,
        'order_r': order_r,
        'burn_r': burn_r,
        'module': 'backlog_burn',
    }
    if yb:
        result['yb'] = yb
        result['ybs'] = ys
        result['ybe'] = ye
    return result
