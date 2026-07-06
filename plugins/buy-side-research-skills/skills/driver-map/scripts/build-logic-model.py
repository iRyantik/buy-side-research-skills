"""build-logic-model.py v4 — JSON-driven formula-linked Excel model.

Usage: python .scripts/driver-map/build-logic-model.py <driver-map.json> [-o output.xlsx]

Generates single-sheet Excel with 7 sections from JSON config. Zero hardcodes —
all company data, assumptions, and module choices from JSON.

Sections:
  §1 Reported Segments        Rev/Cost/GP/GM hardcode FY25A, FY26+ = Σ logic lines
  §2 Logic Lines              Module dispatch (yoy/vol_asp/capacity_util/backlog_burn)
                                → inputs + Revenue formula + Check + GM/GP
  §2→§1 Fill                  Back-link Section 1 from logic line results
  §3 P&L                      Total Rev/GP = Σ formulas + Check rows (collapsible)
                                Opex/Tax/NI hardcode FY23-25, formula FY26+
                                Full P&L always shown (SOTP needs all metrics)
                                D&A/EBITDA/EBIT per actuals.da + Rev YoY row
  §4 SOTP Logic               Per line: GP → metric alloc → multiple → MCap → SUM
  §5 SOTP Segments            Segment weighted multiple
  §6 Market Data              yfinance (mcap/price/shares/PE/52W) + implied ratios
  §7 Scenario Summary         3-scenario projection (yoy chain / vol_asp cache / single)

Revenue modules (modules/):
  yoy (built-in)              Rev = Prior × (1 + YoY Active), BBE group
  vol_asp                     Rev = Σ(Vol × Share% × ASP) / 100, multi-tier BBE
  capacity_util               Rev = Capacity × Util% × ASP
  backlog_burn                Rev = Beg_Backlog × Burn%, chain-linked

SOTP valuation methods: pe / ev_ebitda / ev_ebit / ev_sales (per line)
Backward compat: old ``sotp_pe: 40`` auto-converts to ``{method: pe, multiple: 40}``

Format: no gridlines, Calibri 11, yellow+blue inputs, selective C-column bold,
        ASP #,##0.0, Revenue #,##0, C1 (CUR millions), freeze D2.

validate_json() runs before build — checks depth, method, array lengths, required fields.
"""

import json, argparse, codecs, functools, os, time, datetime
import yfinance as yf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import win32com.client

# ── Row reference: wraps int row number for safety + traceability ──
class Ref:
    __slots__ = ('_r', '_label')
    def __init__(self, row, label=''):
        self._r = row; self._label = label
    def __int__(self): return self._r
    def __index__(self): return self._r
    def __str__(self): return str(self._r)
    def __repr__(self): return f'Ref({self._r}, {self._label!r})'

# ── Format constants ──
PCT = '0.0%'; NUM = '#,##0.0'; DEC = '#,##0.00'; INT = '#,##0'
PRICE_FMT = {'USD': '$#,##0.00', 'JPY': '¥#,##0', 'CNY': '¥#,##0.00', 'KRW': '₩#,##0',
             'TWD': 'NT$#,##0.00', 'HKD': 'HK$#,##0.00', 'SGD': 'S$#,##0.00',
             'EUR': '€#,##0.00', 'GBP': '£#,##0.00'}
DS = 4

# ── Fonts / fills ──
nf   = Font(name='Calibri', size=11)
bf   = Font(name='Calibri', bold=True, size=11)
bf12 = Font(name='Calibri', bold=True, size=12)
itf  = Font(name='Calibri', size=10, italic=True, color='808080')
inpf = Font(name='Calibri', size=11, color='0000CC')
inpfill = PatternFill('solid', fgColor='FFFFCC')
actfill = PatternFill('solid', fgColor='F0F0F0')
hlfill = PatternFill('solid', fgColor='963634')
hlfont = Font(name='Calibri', bold=True, size=11, color=Color(rgb='FFFFFF'))

# ── Shared cell helpers ──
def C(ws, r, c, v=None, font=None, fill=None, fmt=None):
    cl = ws.cell(row=r, column=c, value=v)
    cl.font = font or nf
    if fill: cl.fill = fill
    if fmt: cl.number_format = fmt
    cl.alignment = Alignment(
        horizontal='right' if c >= DS else 'left',
        vertical='center', wrap_text=False)

def I(ws, r, c, v, fmt=None):
    C(ws, r, c, v, font=inpf, fill=inpfill, fmt=fmt)

def A(ws, r, c, v, fmt=None):
    C(ws, r, c, v, fill=actfill, fmt=fmt)

def CF(ws, r, c, formula, fmt=None):
    """Write a formula cell — black font, no fill, guaranteed number_format."""
    C(ws, r, c, formula, fmt=fmt)

def HL(ws, r, c, v=None, fmt=None):
    """Highlight cell — deep red bg + white bold font."""
    C(ws, r, c, v, font=hlfont, fill=hlfill, fmt=fmt)

def BOLD(ws, r, c, v=None, fmt=None):
    """Bold key metric — black bold font."""
    C(ws, r, c, v, font=bf, fmt=fmt)

# ── Module contract decorator ──
RENDER_CONTRACT = {'next_R', 'rev_r', 'gm_r', 'gp_r', 'module'}

def validate_contract(fn):
    @functools.wraps(fn)
    def wrapper(ws, R, ll, anchor_info, ctx):
        result = fn(ws, R, ll, anchor_info, ctx)
        missing = RENDER_CONTRACT - set(result.keys())
        if missing:
            raise ValueError(f'{fn.__name__}: missing contract keys {missing}')
        return result
    return wrapper

# ── Context dict passed to modules ──
def make_ctx():
    return {
        'C': C, 'I': I, 'A': A, 'CF': CF, 'HL': HL,
        'nf': nf, 'bf': bf, 'itf': itf,
        'NUM': NUM, 'DEC': DEC, 'PCT': PCT, 'INT': INT,
        'DS': DS, 'FY0': 0, 'LC': 0,
        'proj_n': 0,
    }

# ═══════════════════════════════════════════════════════════════
# Module registry
# ═══════════════════════════════════════════════════════════════

MODULES = {}

# Lazy imports for external modules
def _load_module(name):
    if name not in MODULES:
        if name == 'yoy':
            from modules.yoy import render as fn
        elif name == 'vol_asp':
            from modules.vol_asp import render as fn
        elif name == 'capacity_util':
            from modules.capacity_util import render as fn
        elif name == 'ebitda':
            from modules.ebitda import render as fn
        elif name == 'backlog_burn':
            from modules.backlog_burn import render as fn
        else:
            raise ValueError(f'Unknown module: {name}')
        MODULES[name] = fn


# ═══════════════════════════════════════════════════════════════
# JSON validation
# ═══════════════════════════════════════════════════════════════

VALID_DEPTH = {'gp', 'op', 'ebitda'}
VALID_METHOD = {'pe', 'ps', 'ev_ebitda', 'ev_ebit', 'ev_sales'}
def validate_json_new(raw_cfg):
    """Validate research-model.json in new FY-inside format (reads raw before adapter)."""
    meta = raw_cfg.get('meta', {})
    asm = raw_cfg.get('assumptions', {})
    proj_n = meta.get('proj_years', 5)
    bfyr = meta.get('base_fy', 2025)

    depth = meta.get('p&l_depth', 'ni')
    if depth not in VALID_DEPTH:
        raise ValueError(f'p&l_depth must be one of {VALID_DEPTH}, got {depth}')

    # Actuals: warn if core fields are missing for all history FYs
    gaap_is = raw_cfg.get('actuals', {}).get('gaap', {}).get('is', {})
    non_is = raw_cfg.get('actuals', {}).get('non_gaap', {}).get('is', {})
    history_fys = [f'FY{bfyr-2}', f'FY{bfyr-1}', f'FY{bfyr}']

    core_gaap = ['rev']
    if depth != 'ebitda': core_gaap.append('gp')
    for fy_key in history_fys:
        for field in core_gaap:
            val = gaap_is.get(field, {}).get(fy_key, {}).get('annual')
            if val is None:
                print(f'  [warn] actuals.gaap.is.{field}.{fy_key}.annual missing')

    # Logic lines
    lines = asm.get('lines', [])
    seg_names = set()
    for ll in lines:
        ln = ll.get('name', '?')
        seg = ll.get('segment', '')
        if seg:
            seg_names.add(seg)

        # SOTP — required, no empty defaults
        sotp = ll.get('sotp', {})
        if not sotp or 'method' not in sotp:
            raise ValueError(f'{ln}: sotp is required (method + multiple)')
        m = sotp.get('method', 'pe')
        if m not in VALID_METHOD:
            raise ValueError(f'{ln}: sotp.method {m} not in {VALID_METHOD}')
        if 'multiple' not in sotp:
            raise ValueError(f'{ln}: sotp.multiple required')

        # vol_asp: check projection year coverage
        if ll.get('module') == 'vol_asp':
            vol = ll.get('volume', {})
            for i in range(1, proj_n + 1):
                proj_fy = f'FY{bfyr + i}E'
                v = vol.get(proj_fy)
                if isinstance(v, dict): v = v.get('annual')
                if v is None:
                    print(f'  [warn] {ln}: volume.{proj_fy} missing')
            for t in ll.get('tiers', []):
                tn = t.get('name', '?')
                for k in ('asp', 'asp_base'):
                    asp = t.get(k, {})
                    if not asp: continue
                    for i in range(1, proj_n + 1):
                        proj_fy = f'FY{bfyr + i}E'
                        a = asp.get(proj_fy)
                        if isinstance(a, dict): a = a.get('annual')
                        if a is None:
                            print(f'  [warn] {ln}/{tn} {k}.{proj_fy} missing')

    # Segment residuals: warn if references unknown segment names
    residuals = asm.get('segment_residuals', {})
    for seg_name in residuals:
        if seg_name not in seg_names:
            print(f'  [warn] segment_residuals "{seg_name}" not referenced by any line.segment')

    print(f'  Validated: depth={depth}, {len(lines)} logic lines, {len(seg_names)} segments')


# ═══════════════════════════════════════════════════════════════
# Main build function
# ═══════════════════════════════════════════════════════════════

def _build_segments(raw_cfg):
    """Build minimal segment list + company Q data from raw (replaces old adapter)."""
    lines = raw_cfg['assumptions']['lines']
    gaap_segs = raw_cfg['actuals']['gaap']['segments']
    non_segs = raw_cfg['actuals']['non_gaap'].get('segments', [])
    residuals = raw_cfg['assumptions'].get('segment_residuals', {})
    bfyr = raw_cfg['meta'].get('base_fy', 2025)
    fy_keys = [f'FY{bfyr-2}', f'FY{bfyr-1}', f'FY{bfyr}']
    q_start_yr = int(raw_cfg['meta'].get('q_start_yr', bfyr))
    q_start_q = int(raw_cfg['meta'].get('q_start_q', 1))
    q_actual_n = int(raw_cfg['meta'].get('q_actual_count', 4))

    seg_to_lines = {}
    for ll in lines:
        sn = ll.get('segment', '')
        if sn:
            seg_to_lines.setdefault(sn, []).append(ll)

    segments = []
    for gs in gaap_segs:
        sn = gs['name']
        matched = seg_to_lines.get(sn, [])
        if not matched:
            continue
        seg_q = {}
        for qi in range(q_actual_n):
            cal_qk = f'Q{((q_start_q - 1 + qi) % 4) + 1}'
            rel_qk = f'q{qi+1}'
            fy_ofs = qi // 4
            q_fy = f'FY{q_start_yr + fy_ofs}'
            q_entry = {}
            for fld, old_key in [('rev', 'rev'), ('gp', 'gp'), ('oi', 'op')]:
                v = gs.get(fld, {}).get(q_fy, {}).get(cal_qk, 0) or gs.get(fld, {}).get(f'FY{q_start_yr}', {}).get(cal_qk, 0)
                if v: q_entry[old_key] = v
            s_short = sn.replace(' Segment', '')
            ns = next((s for s in non_segs if s['name'] == s_short), None)
            if ns:
                eb_q = ns.get('ebitda', {}).get(q_fy, {}).get(cal_qk, 0)
                if eb_q: q_entry['ebitda'] = eb_q
            if q_entry:
                seg_q[rel_qk] = q_entry
        seg_entry = {
            'name': sn,
            'logic_lines': [{'name': ll['name'],
                           'split': 1.0 if ll.get('one_to_one') else ll.get('split', 1.0 / max(1, len(matched)))}
                          for ll in matched],
            'quarters': seg_q,
            'residual': {},
        }
        res = residuals.get(sn, {}) or residuals.get(sn.replace(' Segment', ''), {})
        if res:
            seg_entry['residual'] = {'gm': res.get('base_rate', 0)}
        segments.append(seg_entry)

    # Company-level Q data
    quarters = {}
    gaap_is = raw_cfg['actuals']['gaap']['is']
    non_is = raw_cfg['actuals']['non_gaap']['is']
    for qi in range(q_actual_n):
        cal_qk = f'Q{((q_start_q - 1 + qi) % 4) + 1}'
        rel_qk = f'q{qi+1}'
        fy_ofs = qi // 4
        q_fy = f'FY{q_start_yr + fy_ofs}'
        quarters[rel_qk] = {
            'rev': gaap_is.get('rev', {}).get(q_fy, {}).get(cal_qk, 0),
            'gp': gaap_is.get('gp', {}).get(q_fy, {}).get(cal_qk, 0),
            'op': gaap_is.get('oi', {}).get(q_fy, {}).get(cal_qk, 0),
            'ni': gaap_is.get('ni', {}).get(q_fy, {}).get(cal_qk, 0),
            'tax': gaap_is.get('tax', {}).get(q_fy, {}).get(cal_qk, 0),
            'ebitda': non_is.get('ebitda', {}).get(q_fy, {}).get(cal_qk, 0),
        }
    return segments, quarters



def build(json_path, output_path=None):
    with codecs.open(json_path, 'r', 'utf-8') as f:
        cfg = json.load(f)

    # ── Validate, build compat structures from raw ──
    validate_json_new(cfg)
    raw = cfg                                             # no deep copy needed
    segments, cfg_quarters = _build_segments(raw)         # thin compat layer
    cfg['segments'] = segments
    cfg['quarters'] = cfg_quarters
    logic_lines = raw['assumptions']['lines']             # new-format lines directly

    meta = raw['meta']
    # ── /Compat ──

    is_ebitda_depth = meta.get('p&l_depth') == 'ebitda'
    is_op_depth = cfg['meta'].get('p&l_depth') == 'op'
    is_gp_depth = cfg['meta'].get('p&l_depth') == 'gp'

    # ── New-format direct accessors (FY-inside structure) ──
    bfyr = cfg['meta'].get('base_fy', 2025)
    FY0_KEY = f'FY{bfyr}'
    FY1_KEY = f'FY{bfyr-1}'
    FY2_KEY = f'FY{bfyr-2}'
    _R = raw['actuals']  # shortcut to flipped actuals

    def _gaap(field, fy, period='annual'):
        return _R['gaap']['is'].get(field, {}).get(fy, {}).get(period, 0)

    def _non(field, fy, period='annual'):
        return _R['non_gaap']['is'].get(field, {}).get(fy, {}).get(period, 0)

    def _gaap_seg(name, field, fy, period='annual'):
        for s in _R['gaap'].get('segments', []):
            if s['name'] == name:
                return s.get(field, {}).get(fy, {}).get(period)
        return None

    def _non_seg(name, field, fy, period='annual'):
        s_short = name.replace(' Segment', '')
        for s in _R['non_gaap'].get('segments', []):
            if s['name'] == s_short:
                return s.get(field, {}).get(fy, {}).get(period)
        return None

    def _non_seg(name, field, fy, period='annual'):
        s_short = name.replace(' Segment', '')
        for s in _R['non_gaap'].get('segments', []):
            if s['name'] == s_short:
                return s.get(field, {}).get(fy, {}).get(period, 0)
        return 0

    def _br(line_idx, fy):
        return raw['assumptions']['lines'][line_idx].get('base_rate', {}).get(fy, {}).get('annual', 0)

    def _yoy(line_idx, scenario, fy):
        return raw['assumptions']['lines'][line_idx].get('yoy', {}).get(scenario, {}).get(fy, {}).get('annual', 0)

    def _gl(field, fy):
        return raw['assumptions']['global'].get(field, {}).get(fy, {}).get('annual', 0)

    def _seg_name(line_idx):
        return raw['assumptions']['lines'][line_idx].get('segment', '')

    # ── Mutable OPM cache (Phase 1.3 blend modifies in-place) ──
    # Compute proj_keys dynamically (matches adapter's scan for FY keys with 'E')
    _proj_keys = set()
    for _ll in raw['assumptions']['lines']:
        for _field in ['base_rate', 'volume']:
            for _k in _ll.get(_field, {}):
                if isinstance(_k, str) and 'E' in _k: _proj_keys.add(_k)
        for _sc in ['bull', 'base', 'bear']:
            for _k in _ll.get('yoy', {}).get(_sc, {}):
                if isinstance(_k, str) and 'E' in _k: _proj_keys.add(_k)
        for _t in _ll.get('tiers', []):
            for _ak in ['asp', 'asp_base', 'asp_bull', 'asp_bear']:
                for _k in _t.get(_ak, {}):
                    if isinstance(_k, str) and 'E' in _k: _proj_keys.add(_k)
    _PROJ_FYS = sorted(_proj_keys)
    _ALL_FYS = [FY2_KEY, FY1_KEY, FY0_KEY] + _PROJ_FYS
    _opm_cache = {}
    for _fy in _ALL_FYS:
        _v = _gl('opex_rev', _fy)
        _opm_cache[_fy] = round(_v, 4) if _v else 0.25

    def _opm(fy):
        """Mutable Opex/Rev rate cache — use _opm(fy_key) instead of gl['opm'][idx]."""
        if fy not in _opm_cache:
            _v = _gl('opex_rev', fy)
            _opm_cache[fy] = round(_v, 4) if _v else 0.25
        return _opm_cache[fy]

    # ── Mutable GM cache (Phase 1.3 blend modifies in-place) ──
    _gm_cache = {}

    def _gm(line_idx, fy):
        """Mutable base_rate cache — use _gm(line_idx, fy) instead of ll['gm'][...]."""
        key = (line_idx, fy)
        if key not in _gm_cache:
            _gm_cache[key] = _br(line_idx, fy)
        return _gm_cache[key]

    # ── Line-level helpers (replace ll['...'] old-format access) ──
    def _vol(line_idx, fy):
        """Volume for a line+FY — direct access to raw assumptions (FY-inside: {FY: {annual: N}})."""
        v = raw['assumptions']['lines'][line_idx].get('volume', {}).get(fy, 0)
        return v.get('annual', 0) if isinstance(v, dict) else v

    def _asp(line_idx, fy, tier_idx=0, scenario=None):
        """ASP for a line+FY+tier — direct access to raw assumptions. scenario='base'|'bull'|'bear' for BBE."""
        tiers = raw['assumptions']['lines'][line_idx].get('tiers', [])
        if tier_idx < len(tiers):
            t = tiers[tier_idx]
            if scenario:
                v = t.get(f'asp_{scenario}', {}).get(fy, 0)
                if isinstance(v, dict): v = v.get('annual', 0)
                if v: return v
            for key in ('asp_base', 'asp', 'asp_bull', 'asp_bear'):
                v = t.get(key, {}).get(fy, 0)
                if isinstance(v, dict): v = v.get('annual', 0)
                if v: return v
        return 0

    def _has_bb(line_idx, tier_idx=0):
        """Check if tier has BBE scenario ASP arrays."""
        tiers = raw['assumptions']['lines'][line_idx].get('tiers', [])
        if tier_idx < len(tiers):
            t = tiers[tier_idx]
            return any(k in t for k in ('asp_bull', 'asp_base', 'asp_bear'))
        return False

    def _share(line_idx, fy, tier_idx=0):
        """Tier share%% for a line+FY — {FY: {annual: N}}."""
        tiers = raw['assumptions']['lines'][line_idx].get('tiers', [])
        if tier_idx < len(tiers):
            v = tiers[tier_idx].get('share', {}).get(fy, 0)
            return v.get('annual', 0) if isinstance(v, dict) else v
        return 0

    def _capacity(line_idx, fy):
        """Capacity for a line+FY — {FY: {annual: N}}."""
        v = raw['assumptions']['lines'][line_idx].get('capacity', {}).get(fy, 0)
        return v.get('annual', 0) if isinstance(v, dict) else v

    def _utilization(line_idx, fy):
        """Utilization%% for a line+FY — {FY: {annual: N}}."""
        v = raw['assumptions']['lines'][line_idx].get('utilization', {}).get(fy, 0)
        return v.get('annual', 0) if isinstance(v, dict) else v

    def _bb_field(line_idx, field, fy, scenario=None):
        """Generic backlog_burn field reader — {FY: {annual: N}} or {scenario: {FY: {annual: N}}}."""
        data = raw['assumptions']['lines'][line_idx].get(field, {})
        if scenario and isinstance(data.get(scenario), dict):
            v = data[scenario].get(fy, 0)
        else:
            v = data.get(fy, 0)
        return v.get('annual', 0) if isinstance(v, dict) else v

    def _has_module(line_idx, module_name):
        """Check if a line uses a specific module."""
        return raw['assumptions']['lines'][line_idx].get('module') == module_name

    def _unit_scale(line_idx):
        """Unit scale for a vol_asp line (default 100 if not set)."""
        return raw['assumptions']['lines'][line_idx].get('unit_scale', 100)

    # ═══ Phase 1.1: Reconcile — scale Q→FY for M=4 complete actual FYs ═══
    meta_tmp = cfg['meta']
    q_actual_n = meta_tmp.get('q_actual_count', 0)
    if q_actual_n > 0:
        bfyr = meta_tmp['base_fy']; proj_n = meta_tmp['proj_years']
        q_start_yr = meta_tmp.get('q_start_yr', bfyr)
        q_start_q = meta_tmp.get('q_start_q', 1)
        cur_yr, cur_q, qi = q_start_yr, q_start_q, 0
        total_q = q_actual_n + meta_tmp.get('q_proj_count', 0)
        while qi < total_q:
            rem = 4 - cur_q + 1; fyc = min(rem, total_q - qi)
            if fyc == 4:  # complete 4Q FY
                fy_idx = cur_yr - bfyr + 2
                fy_rev = 0
                if fy_idx < 3:
                    fy_rev = _gaap('rev', f'FY{cur_yr}')
                elif fy_idx < 3 + proj_n:
                    proj_i = fy_idx - 3
                    proj_fy = f'FY{bfyr + proj_i + 1}E'
                    for line_idx, ll in enumerate(logic_lines):
                        if _has_module(line_idx, 'vol_asp'):
                            vol_v = _vol(line_idx, proj_fy)
                            asp_v = _asp(line_idx, proj_fy)
                            if vol_v:
                                fy_rev += vol_v * asp_v / _unit_scale(line_idx)
                        else:
                            for seg in cfg.get('segments',[]):
                                for l in seg.get('logic_lines',[]):
                                    if l['name'] == ll['name']:
                                        f0 = _gaap_seg(seg['name'], 'rev', FY0_KEY) * l['split']
                                        cum = 1.0
                                        for bi in range(proj_i+1):
                                            prior_fy = f'FY{bfyr + bi + 1}E'
                                            cum *= (1 + _yoy(line_idx, 'base', prior_fy))
                                        fy_rev += round(f0*cum)
                if fy_rev == 0: qi += fyc; cur_yr += 1; cur_q = 1; continue
                qdata = cfg.get('quarters',{})
                # Guard: only reconcile if ALL 4 Qs have actual data (skip mixed actual+proj FYs)
                all_q_present = all(qdata.get(f'q{qi+j+1}',{}).get('rev',0) > 0 for j in range(4))
                if not all_q_present: qi += fyc; cur_yr += 1; cur_q = 1; continue
                q_rev_sum = sum(qdata.get(f'q{qi+j+1}',{}).get('rev',0) for j in range(4))
                if q_rev_sum > 0 and abs(q_rev_sum/fy_rev-1) > 0.005:
                    s = fy_rev / q_rev_sum
                    print(f'  [reconcile] FY{cur_yr}: Q sum={q_rev_sum:.0f} FY={fy_rev:.0f} scale={s:.3f}')
                    for field in ['rev','gp','op','ni','opex','tax','da']:
                        for j in range(4):
                            qk=f'q{qi+j+1}'; qd=qdata.get(qk,{})
                            if field in qd: qdata[qk][field]=round(qd[field]*s)
                    for seg in cfg.get('segments',[]):
                        sq=seg.get('quarters',{})
                        for field in ['rev','gp','op','ni','opex','tax','da']:
                            for j in range(4):
                                qk=f'q{qi+j+1}'; qd=sq.get(qk,{})
                                if field in qd: sq[qk][field]=round(qd[field]*s)
                    for ll in cfg.get('logic_lines',[]):
                        qh=ll.setdefault('q_history',{})
                        for j in range(4):
                            qk=f'q{qi+j+1}'; qd=qh.get(qk,{})
                            if 'rev' in qd: qh[qk]['rev']=round(qd['rev']*s)
                            if 'volume' in qd and 'asp' in qd and qd.get('asp',0)>0:
                                qh[qk]['volume']=round(qh[qk]['rev']*ll.get('unit_scale',100)/qd['asp'])
            qi += fyc; cur_yr += 1; cur_q = 1

        # ═══ Phase 1.3: Blend — actual Q profit rates → annual model assumptions ═══
        # For projection FYs with M∈{1,2,3}, blend actual Q margins with model
        _line_idx_by_name = {}
        for _i, _ll in enumerate(logic_lines):
            _line_idx_by_name[_ll['name']] = _i
        cur_yr, cur_q, qi_b = q_start_yr, q_start_q, 0
        while qi_b < total_q:
            rem = 4 - cur_q + 1; fyc = min(rem, total_q - qi_b)
            if fyc == 4:
                fy_idx = cur_yr - bfyr + 2; proj_i = fy_idx - 3
                if 0 <= proj_i < proj_n:
                    for seg in cfg.get('segments', []):
                        sq = seg.get('quarters', {})
                        # Count actual Qs for this segment in this FY
                        seg_q_revs = []; seg_q_gps = []; seg_q_ops = []
                        for j in range(4):
                            qk = f'q{qi_b+j+1}'; qd = sq.get(qk, {})
                            rv = qd.get('rev', 0); gv = qd.get('gp', 0)
                            if rv and rv > 0:
                                seg_q_revs.append(rv)
                                if gv: seg_q_gps.append(gv)
                                ov = qd.get('op')
                                if ov is not None: seg_q_ops.append(ov)
                        M_seg = len(seg_q_revs)
                        if M_seg not in (1, 2, 3): continue
                        w_act = M_seg / 4; w_mod = 1 - w_act

                        # Blend per-line gm + opm for each logic line in this segment
                        for ll_cfg in seg.get('logic_lines', []):
                            ln_name = ll_cfg['name']
                            ll_obj = None
                            for ll in logic_lines:
                                if ll['name'] == ln_name: ll_obj = ll; break
                            if not ll_obj: continue

                            # GM blend
                            if seg_q_gps and sum(seg_q_revs) > 0:
                                gm_actual = sum(seg_q_gps) / sum(seg_q_revs)
                                _blend_idx = _line_idx_by_name.get(ln_name)
                                if _blend_idx is not None and proj_i < len(_PROJ_FYS):
                                    proj_fy_blend = _PROJ_FYS[proj_i]
                                    gm_model = _gm(_blend_idx, proj_fy_blend)
                                    gm_blend = w_act * gm_actual + w_mod * gm_model
                                    _gm_cache[(_blend_idx, proj_fy_blend)] = round(gm_blend, 4)

                            # Opex/Rev blend (requires OP data)
                            if seg_q_ops and seg_q_gps and sum(seg_q_revs) > 0:
                                opex_actual = sum(seg_q_gps) - sum(seg_q_ops)
                                om_actual = opex_actual / sum(seg_q_revs) if sum(seg_q_revs) > 0 else 0
                                line_opex = ll_obj.get('opm')
                                proj_fy_om = _PROJ_FYS[proj_i] if proj_i < len(_PROJ_FYS) else _PROJ_FYS[-1]
                                if line_opex:
                                    idx_o = 3 + proj_i
                                    om_model = line_opex[idx_o] if idx_o < len(line_opex) else 0.25
                                    om_blend = w_act * om_actual + w_mod * om_model
                                    if idx_o < len(line_opex):
                                        ll_obj['opm'][idx_o] = round(om_blend, 4)
                                else:
                                    om_model = _opm(proj_fy_om)
                                    om_blend = w_act * om_actual + w_mod * om_model
                                    _opm_cache[proj_fy_om] = round(om_blend, 4)
                    # Blend global opm (company-level, from all segments' actual Qs)
                    if M_seg > 0:
                        all_act_rev = 0; all_act_gp = 0; all_act_op = 0
                        for seg in cfg.get('segments', []):
                            sq = seg.get('quarters', {})
                            for j in range(4):
                                qk = f'q{qi_b+j+1}'; qd = sq.get(qk, {})
                                rv = qd.get('rev', 0)
                                if rv and rv > 0:
                                    all_act_rev += rv
                                    gv = qd.get('gp', 0)
                                    if gv: all_act_gp += gv
                                    ov = qd.get('op')
                                    if ov is not None: all_act_op += ov
                        if all_act_rev > 0 and all_act_gp > 0 and all_act_op > 0:
                            co_om_act = (all_act_gp - all_act_op) / all_act_rev
                            proj_fy_om = _PROJ_FYS[proj_i] if proj_i < len(_PROJ_FYS) else _PROJ_FYS[-1]
                            co_om_mod = _opm(proj_fy_om)
                            co_blend = (M_seg / 4) * co_om_act + (1 - M_seg / 4) * co_om_mod
                            _opm_cache[proj_fy_om] = round(co_blend, 4)
            qi_b += fyc; cur_yr += 1; cur_q = 1

        # ═══ Phase 1.4: Q Driver Distribution — annual drivers → Q projections ═══
        # For each complete 4Q FY, distribute annual drivers to Qs using seasonal weights.
        # vol_asp: Q_Vol = Vol_Y × w_i, Q_ASP = ASP_Y × s_i (Σ(w×s)=1)
        # yoy: Newton solve r s.t. Σ Q_1×(1+r)^k = Annual
        # backlog_burn: Q_Burn = Burn_Y × w_i, Q_ASP = ASP_Y × s_i

        # Pre-compute total_residual (Non-core absorbs all segment residuals)
        total_residual = 0; gp_residual = 0
        for seg in cfg.get('segments', []):
            r = round(_gaap_seg(seg['name'], 'rev', FY0_KEY) - sum(_gaap_seg(seg['name'], 'rev', FY0_KEY) * l['split'] for l in seg.get('logic_lines', [])))
            if r > 0:
                total_residual += r
                gp_residual += round(r * seg.get('residual', {}).get('gm', 0))
        cur_yr, cur_q, qi = q_start_yr, q_start_q, 0
        while qi < total_q:
            rem = 4 - cur_q + 1; fyc = min(rem, total_q - qi)
            if fyc == 4:
                fy_idx = cur_yr - bfyr + 2; proj_i = fy_idx - 3
                proj_fy = f'FY{bfyr + proj_i + 1}E'
                for line_idx, ll in enumerate(logic_lines):
                    ln = ll['name']; qh = ll.setdefault('q_history', {})
                    module = ll.get('module', 'yoy')
                    us = _unit_scale(line_idx)

                    # ── Compute annual revenue & drivers ──
                    ann = 0; ann_vol = 0; ann_asp = 0
                    if fy_idx < 3:
                        for seg in cfg.get('segments', []):
                            for l in seg.get('logic_lines', []):
                                if l['name'] == ln:
                                    seg_fy_key = ['fy-2', 'fy-1', 'fy0'][fy_idx]
                                    yr_data = seg.get(seg_fy_key, {})
                                    if yr_data and yr_data.get('rev'):
                                        ann += yr_data['rev'] * l['split']
                    elif 0 <= proj_i < proj_n:
                        if module == 'vol_asp':
                            us = _unit_scale(line_idx)
                            ann_vol = _vol(line_idx, proj_fy)
                            ann_asp = _asp(line_idx, proj_fy)
                            ann = ann_vol * ann_asp / us if ann_vol else 0
                        elif module == 'backlog_burn':
                            ann_vol = _bb_field(line_idx, 'burn_rate', proj_fy)
                            ann_asp = _asp(line_idx, proj_fy, 0)
                            ann = ann_vol * ann_asp / _unit_scale(line_idx) if ann_asp else 0
                        else:  # yoy
                            for seg in cfg.get('segments', []):
                                for l in seg.get('logic_lines', []):
                                    if l['name'] == ln:
                                        f0 = _gaap_seg(seg['name'], 'rev', FY0_KEY) * l['split']
                                        cum = 1.0
                                        for bi in range(proj_i + 1):
                                            prior_fy = f'FY{bfyr + bi + 1}E'
                                            cum *= (1 + _yoy(line_idx, 'base', prior_fy))
                                        ann += round(f0 * cum)
                            # Non-core absorbs all segment residuals
                            if ln == 'Non-core':
                                ann += total_residual
                    if ann <= 0: continue

                    # ── Count M and read actual Q driver data ──
                    actual_q = []  # list of (offset, rev, vol, asp) for Qs with data
                    for j in range(4):
                        qk = f'q{qi+j+1}'; qd = qh.get(qk, {})
                        rv = qd.get('rev', 0)
                        # Fallback: segment quarters (skip for vol_asp/backlog with volume+asp)
                        if not rv and qd.get('volume') and qd.get('asp') and module in ('vol_asp', 'backlog_burn'):
                            rv = round(qd['volume'] * qd['asp'] / us, 2)  # compute from vol×asp
                        elif not rv:
                            for seg in cfg.get('segments', []):
                                for l in seg.get('logic_lines', []):
                                    if l['name'] == ln:
                                        sq_r = seg.get('quarters', {}).get(qk, {}).get('rev', 0)
                                        if sq_r:
                                            rv = sq_r * l['split']
                        if rv and rv > 0:
                            vv = qd.get('volume', 0) or (rv * us / ann_asp if ann_asp else 0)
                            av = qd.get('asp', 0) or ann_asp
                            actual_q.append((j, rv, vv, av))
                    M = len(actual_q)
                    if M == 4: continue  # all-actual FY, reconcile handled

                    # ── Compute QoQ rate r ──
                    if M >= 2:
                        qoq_rates = []
                        for mi in range(1, len(actual_q)):
                            r_prev = actual_q[mi - 1][1]; r_curr = actual_q[mi][1]
                            if r_prev > 0: qoq_rates.append(r_curr / r_prev - 1)
                        _r = sum(qoq_rates) / len(qoq_rates) if qoq_rates else 0.02
                    else:
                        _yoy_rate = 0
                        if fy_idx < 3:
                            if fy_idx > 0:
                                fy_curr = f'FY{bfyr + fy_idx - 2}'
                                fy_prev = f'FY{bfyr + fy_idx - 3}'
                                _yoy_rate = _gaap('rev', fy_curr) / _gaap('rev', fy_prev) - 1
                        elif module == 'yoy':
                            _yoy_rate = _yoy(line_idx, 'base', proj_fy)
                        elif module == 'vol_asp':
                            if proj_i > 0:
                                prev_fy = f'FY{bfyr + proj_i}E'
                                vol_prev = _vol(line_idx, prev_fy)
                                if vol_prev > 0:
                                    _yoy_rate = _vol(line_idx, proj_fy) / vol_prev - 1
                        elif module == 'backlog_burn':
                            if proj_i > 0:
                                prev_fy_bb = f'FY{bfyr + proj_i}E'
                                vol_prev_bb = _bb_field(line_idx, 'burn_rate', prev_fy_bb)
                                if vol_prev_bb > 0:
                                    _yoy_rate = _bb_field(line_idx, 'burn_rate', proj_fy) / vol_prev_bb - 1
                        _r = (1 + _yoy_rate) ** (1 / 4) - 1 if _yoy_rate > -0.99 else 0.02
                    _r = max(-0.2, min(0.2, _r))  # cap for vol_asp/backlog weights only

                    # ── Compute seasonal weights ──
                    actual_set = set(a[0] for a in actual_q)
                    first_actual = actual_q[0][0] if actual_q else 0
                    last_actual = actual_q[-1][0] if actual_q else -1

                    if module == 'yoy':
                        # ── yoy: binary search for r such that Σ Q_i = ann ──
                        a_first = actual_q[0][1] if actual_q else ann / 4
                        a_last = actual_q[-1][1] if actual_q else ann / 4
                        def yoy_sum(_r):
                            s = 0
                            for j in range(4):
                                if j in actual_set:
                                    for a in actual_q:
                                        if a[0] == j: s += a[1]
                                elif j < first_actual:
                                    s += a_first / ((1 + _r) ** (first_actual - j))
                                else:  # j > last_actual
                                    s += a_last * ((1 + _r) ** (j - last_actual))
                            return s
                        lo, hi = -0.5, 1.0
                        for _ in range(30):
                            mid = (lo + hi) / 2
                            s = yoy_sum(mid)
                            if abs(s - ann) < 0.5:
                                r_q = mid; break
                            if s > ann: hi = mid
                            else: lo = mid
                        else:
                            r_q = mid
                        r_q = max(-0.5, min(1.0, r_q))

                        # Write YoY_Q rates to q_history
                        for j in range(4):
                            qk = f'q{qi+j+1}'
                            if j in actual_set: continue
                            qh[qk] = qh.get(qk, {})
                            qh[qk]['yoy_q'] = round(r_q, 6)

                        # Non-core: write full Q revs + GP (base + residual) to q_history
                        if ln == 'Non-core' and total_residual > 0:
                            residual_per_q = round(total_residual / 4)
                            gp_residual_per_q = round(gp_residual / 4)
                            nc_gm = _br(line_idx, proj_fy) or 0.22
                            if not isinstance(nc_gm, (int, float)): nc_gm = 0.22
                            for j in range(4):
                                qk = f'q{qi+j+1}'
                                qh[qk] = qh.get(qk, {})
                                if j in actual_set:
                                    for a in actual_q:
                                        if a[0] == j:
                                            qh[qk]['rev'] = a[1] + residual_per_q
                                            qh[qk]['gp'] = round(a[1] * nc_gm) + gp_residual_per_q
                                elif j < first_actual:
                                    base_rev = round(a_first / ((1 + r_q) ** (first_actual - j)))
                                    qh[qk]['rev'] = base_rev + residual_per_q
                                    qh[qk]['gp'] = round(base_rev * nc_gm) + gp_residual_per_q
                                else:
                                    base_rev = round(a_last * ((1 + r_q) ** (j - last_actual)))
                                    qh[qk]['rev'] = base_rev + residual_per_q
                                    qh[qk]['gp'] = round(base_rev * nc_gm) + gp_residual_per_q

                        # Compute Q Rev targets for debug
                        q_revs = []
                        for j in range(4):
                            if j in actual_set:
                                for a in actual_q:
                                    if a[0] == j: q_revs.append(a[1])
                            elif j < first_actual:
                                q_revs.append(a_first / ((1 + r_q) ** (first_actual - j)))
                            else:
                                q_revs.append(a_last * ((1 + r_q) ** (j - last_actual)))
                        print(f'  [Q driver] {ln} FY{cur_yr} M={M} (yoy): r={r_q:.4f} Q revs={[round(x) for x in q_revs]} sum={round(sum(q_revs))} ann={round(ann)}')

                    elif module in ('vol_asp', 'backlog_burn'):
                        # ── vol_asp / backlog_burn: Vol/Burn + ASP weights ──
                        # Locked actual Qs consume part of annual Vol and Rev budget
                        locked_vol = sum(a[2] for a in actual_q)
                        locked_rev = sum(a[1] for a in actual_q)
                        remaining_vol = max(0.01, ann_vol - locked_vol)
                        remaining_rev = max(0.01, ann - locked_rev)
                        remaining_asp = remaining_rev * us / remaining_vol  # natural ASP for formula Vol*ASP/us

                        # Compute Vol/Burn weights from actual Q data
                        if M >= 2 and len([a for a in actual_q if a[2] > 0]) >= 2:
                            vol_actuals = [a for a in actual_q if a[2] > 0]
                            vol_sum = sum(a[2] for a in vol_actuals)
                            w_actual = {a[0]: a[2] / vol_sum for a in vol_actuals}
                        else:
                            w_actual = {}

                        # Compute ASP seasonal s_i_raw from actual Q data
                        if M >= 2 and len([a for a in actual_q if a[3] > 0]) >= 2:
                            asp_actuals = [a for a in actual_q if a[3] > 0]
                            s_raw_vals = {a[0]: a[3] / ann_asp for a in asp_actuals} if ann_asp else {}
                        else:
                            s_raw_vals = {}

                        # Build w and s for projection Qs only
                        n_proj = 4 - M
                        w = [0.0] * 4; s = [1.0] * 4
                        proj_indices = [j for j in range(4) if j not in actual_set]
                        for j in proj_indices:
                            if j in w_actual:
                                w[j] = w_actual[j]
                                s[j] = s_raw_vals.get(j, 1.0)
                            elif j < first_actual:
                                steps = first_actual - j
                                w[j] = w_actual.get(first_actual, 1.0/n_proj) / ((1 + _r) ** steps)
                                s[j] = s_raw_vals.get(first_actual, 1.0)
                            elif j > last_actual:
                                steps = j - last_actual
                                w[j] = w_actual.get(last_actual, 1.0/n_proj) * ((1 + _r) ** steps)
                                s[j] = s_raw_vals.get(last_actual, 1.0)
                            else:
                                # M=0: uniform with r growth
                                w[j] = (1.0 / n_proj) * ((1 + _r) ** (j + 1))
                                s[j] = 1.0

                        # Normalize w for projection Qs to sum to 1
                        w_proj_sum = sum(w[j] for j in proj_indices)
                        if w_proj_sum > 0:
                            for j in proj_indices:
                                w[j] /= w_proj_sum

                        # Normalize s for projection Qs so Σ(w × s) = 1
                        inner = sum(w[j] * s[j] for j in proj_indices)
                        if inner > 0:
                            for j in proj_indices:
                                s[j] /= inner

                        # Write q_history
                        for j in range(4):
                            qk = f'q{qi+j+1}'
                            qh[qk] = qh.get(qk, {})
                            if j in actual_set:
                                # Lock actual Qs: preserve Vol, set ASP so Vol×ASP=actual Rev
                                for a in actual_q:
                                    if a[0] == j:
                                        qh[qk]['rev'] = a[1]
                                        if a[2] > 0 and a[1] > 0 and module == 'vol_asp':
                                            qh[qk]['volume'] = a[2]
                                            qh[qk]['asp'] = a[1] * us / a[2]  # natural ASP for Vol×ASP/us formula
                                continue
                            if module == 'vol_asp':
                                qh[qk]['volume'] = remaining_vol * w[j]
                                qh[qk]['asp'] = remaining_asp * s[j]
                            else:  # backlog_burn
                                qh[qk]['burn'] = remaining_vol * w[j]  # remaining_vol = Burn_budget
                                qh[qk]['asp'] = remaining_asp * s[j]
                                order_yr = _bb_field(line_idx, 'order_rate', proj_fy)
                                qh[qk]['order'] = order_yr * w[j]

                        # Debug output
                        _revs = []
                        for j in range(4):
                            if j in actual_set:
                                for a in actual_q:
                                    if a[0] == j: _revs.append(a[1])
                            else:
                                _revs.append(remaining_vol * w[j] * remaining_asp * s[j] / us)
                        print(f'  [Q driver] {ln} FY{cur_yr} M={M} ({module}): '
                              f'w={[round(w[j],3) if j in proj_indices else 0 for j in range(4)]} '
                              f's={[round(s[j],3) if j in proj_indices else 0 for j in range(4)]} '
                              f'revs={[round(x) for x in _revs]} sum={round(sum(_revs))} ann={round(ann)}')

            qi += fyc; cur_yr += 1; cur_q = 1

    # ═══ Phase 1.5: Market Data — yfinance snapshot ═══
    try:
        yf_ticker = meta.get('yf_ticker', meta['ticker'])
        info = yf.Ticker(yf_ticker).info
    except Exception:
        info = {}
    mcap_raw = info.get('marketCap', 0) or 0
    price = info.get('currentPrice', 0) or 0
    shares = int(info.get('sharesOutstanding', 0) / 1e6) or 0
    ttm_pe = info.get('trailingPE', 0) or 0
    fwd_pe = info.get('forwardPE', 0) or 0
    hi52 = info.get('fiftyTwoWeekHigh', 0) or 0
    lo52 = info.get('fiftyTwoWeekLow', 0) or 0
    # Manual overrides from meta take priority over yfinance
    if meta.get('mcap_m'): mcap_M = meta['mcap_m']
    else: mcap_M = int(mcap_raw / 1e6) if mcap_raw else 0
    if meta.get('price'): price = meta['price']
    if meta.get('shares_m'): shares = meta['shares_m']
    # Unit: display_unit ("M"=millions) + display_decimals from JSON meta
    display_unit = meta.get('display_unit', 'M')
    display_decimals = meta.get('display_decimals', 1)
    # B-mode: JPY/KRW/TWD display in billions (extra ÷1000)
    use_B = meta.get('currency', '') in ('JPY', 'KRW', 'TWD')
    div = 1000 if use_B else 1
    mcap_d = round(mcap_M / div, display_decimals)
    price_fmt = PRICE_FMT.get(meta.get('currency', 'USD'), '$#,##0.00')

    def sc(v):
        return round(v / div, display_decimals)

    bfyr = meta['base_fy']; proj_n = meta['proj_years']; s_off = meta['sotp_offset']
    q_actual_n = meta.get('q_actual_count', 0)
    q_proj_n = meta.get('q_proj_count', 0)
    has_q = q_actual_n + q_proj_n > 0
    COLS = 3 + proj_n
    FY0 = DS + 2; SC = FY0 + s_off; LC_ANNUAL = DS + COLS - 1
    LC = LC_ANNUAL  # always the last annual column (never polluted by Q)
    if has_q:
        Q_START = LC_ANNUAL + 3  # 2 blank columns between Y and Q
        Q_END = Q_START + q_actual_n + q_proj_n - 1
        ALL_END = Q_END  # last column including Q, for loops that span both
        # Q labels: 4Q25A, 1Q26A, 2Q26E, 3Q26E, ...
        q_start_yr = meta.get('q_start_yr', bfyr)
        q_start_q = meta.get('q_start_q', 1)
        yr, q = q_start_yr, q_start_q
        QL = []
        for i in range(q_actual_n):
            QL.append(f'{q}Q{str(yr)[-2:]}A')
            q += 1
            if q > 4: q = 1; yr += 1
        for i in range(q_proj_n):
            QL.append(f'{q}Q{str(yr)[-2:]}E')
            q += 1
            if q > 4: q = 1; yr += 1
    else:
        Q_START = Q_END = 0
        ALL_END = LC_ANNUAL
        QL = []

    YR = [f'FY{bfyr - 2}A', f'FY{bfyr - 1}A', f'FY{bfyr}A'] + \
         [f'FY{bfyr + i}E' for i in range(1, proj_n + 1)]

    # ═══════════════ Phase 2: Render Excel ═══════════════
    # ── Workbook setup ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = meta['company'][:31]
    ws.sheet_view.showGridLines = False
    C(ws, 1, 1, 'Scenario:', font=bf)
    dv = DataValidation(type='list', formula1='"Bull,Base,Bear"')
    ws.add_data_validation(dv)
    C(ws, 1, 2, 'Base', font=Font(name='Calibri', bold=True, size=12, color='2F5496'))
    dv.add('B1')
    for ci, y in enumerate(YR, DS):
        C(ws, 1, ci, y, font=bf)
    for ci, ql in enumerate(QL, Q_START):
        C(ws, 1, ci, ql, font=bf)
    unit_label = 'bn' if use_B else 'millions'
    C(ws, 1, 3, f'({meta.get("currency","CNY")} {unit_label})', font=itf)

    # module context (modules see annual range only)
    ctx = make_ctx()
    ctx['FY0'] = FY0; ctx['LC'] = LC_ANNUAL; ctx['SC'] = SC; ctx['proj_n'] = proj_n
    ctx['COLS'] = COLS; ctx['bfyr'] = bfyr
    ctx['Q_START'] = Q_START; ctx['Q_END'] = Q_END
    ctx['q_actual_n'] = q_actual_n; ctx['q_proj_n'] = q_proj_n

    print(f'  Cols: D=DS({DS}) FY0={get_column_letter(FY0)}({FY0}) '
          f'LC={get_column_letter(LC_ANNUAL)}({LC_ANNUAL}) SC={get_column_letter(SC)}({SC}) '
          f'proj_n={proj_n} unit={display_unit} div={div}'
          + (f' Q={get_column_letter(Q_START)}({Q_START})-{get_column_letter(Q_END)}({Q_END})' if has_q else ''))

    # ═══ Phase 2.1: §1 Reported Segments ═══
    R = 3
    C(ws, R, 1, 'Reported Segments', font=bf12)
    s1_start = R
    R += 1
    C(ws, R, 1, '(Segments)', font=itf)
    # Basis label (below Reported Segments, same as Logic Lines)
    basis = meta.get('basis', 'gaap')
    basis_note = meta.get('basis_note', '')
    basis_label = f'Basis: {basis.upper()}'
    if basis_note:
        basis_label += f' — {basis_note[:120]}'
    C(ws, R, 1, basis_label, font=itf)
    R = 6
    seg_info = {}
    anchor_info = {}  # {ln: (Section1_Rev_row, value_in_M)}
    one_to_one = set()  # logic lines where segment=line (split=1.0, no residual)
    line_to_seg = {}    # {ln: seg_name}
    line_to_split = {}  # {ln: split_row}

    for seg in segments:
        sn = seg['name']; lls = seg['logic_lines']
        srev = _gaap_seg(sn, 'rev', FY0_KEY) or 0
        if is_ebitda_depth:
            sebitda = _non_seg(sn, 'ebitda', FY0_KEY) or 0
        else:
            sgp = _gaap_seg(sn, 'gp', FY0_KEY) or 0; sgm = sgp / srev if srev else 0; scost = srev - sgp
        # History layer: FY-2 and FY-1. Optional — leave empty if segment didn't exist.
        hist_years = [(FY2_KEY, DS), (FY1_KEY, DS + 1)]

        # Revenue
        for yr_key, col in hist_years:
            yr_val = _gaap_seg(sn, 'rev', yr_key)
            if yr_val: A(ws, R, col, sc(yr_val), fmt=NUM)
            else: C(ws, R, col, '', fmt=NUM)
        A(ws, R, FY0, sc(srev), fmt=NUM)
        # Q columns: segment Q actuals
        seg_quarters = seg.get('quarters', {})
        for qi in range(q_actual_n):
            qk = f'q{qi+1}'; qd = seg_quarters.get(qk, {})
            if qd.get('rev'): A(ws, R, Q_START + qi, sc(qd['rev']), fmt=NUM)
        for qi in range(q_proj_n):
            C(ws, R, Q_START + q_actual_n + qi, '', fmt=NUM)
        rev_r = R
        C(ws, R, 2, sn, font=bf)
        C(ws, R, 3, 'Revenue', font=bf)
        R += 1
        # name_cn line (if exists)
        name_cn = seg.get('name_cn', '')
        if name_cn:
            C(ws, R, 2, name_cn, font=itf)
            ws.cell(row=R, column=2).font = Font(color='808080', italic=True)
            R += 1

        # Implied YoY (annual) + QoQ (quarterly)
        for yr_key, col in hist_years:
            C(ws, R, col, '', fmt=PCT)
        cl_e = get_column_letter(DS + 1); cl_d = get_column_letter(DS)
        C(ws, R, DS + 1, f'=IFERROR({cl_e}{rev_r}/{cl_d}{rev_r}-1,"")', fmt=PCT)
        f0 = get_column_letter(FY0); f_1 = get_column_letter(FY0 - 1)
        C(ws, R, FY0, f'=IFERROR({f0}{rev_r}/{f_1}{rev_r}-1,"")', fmt=PCT)
        for ci in range(FY0 + 1, LC_ANNUAL + 1):
            cl = get_column_letter(ci); pl = get_column_letter(ci - 1)
            C(ws, R, ci, f'=IFERROR({cl}{rev_r}/{pl}{rev_r}-1,"")', fmt=PCT)
        # Q columns: YoY (vs same Q 4 quarters ago)
        if has_q:
            for qi in range(Q_START, Q_START + q_actual_n + q_proj_n):
                cl = get_column_letter(qi)
                if qi - 4 >= Q_START:
                    pl = get_column_letter(qi - 4)
                    C(ws, R, qi, f'=IFERROR({cl}{rev_r}/{pl}{rev_r}-1,"")', fmt=PCT)
                else:
                    C(ws, R, qi, '', fmt=PCT)
        C(ws, R, 3, 'Rev YoY', font=nf)
        R += 1
        # QoQ row (collapsed, Q columns only)
        if has_q:
            for ci in range(DS, LC_ANNUAL + 1):
                C(ws, R, ci, '', fmt=PCT)
            for qi in range(Q_START, Q_END + 1):
                cl = get_column_letter(qi); pl = get_column_letter(qi - 1)
                if qi == Q_START: C(ws, R, qi, '', fmt=PCT)
                else: C(ws, R, qi, f'=IFERROR({cl}{rev_r}/{pl}{rev_r}-1,"")', fmt=PCT)
            C(ws, R, 3, '  Rev QoQ', font=itf)
            R += 1

        if is_ebitda_depth:
            # ── EBITDA block (EBITDA first, margin formula after) ──
            for yr_key, col in hist_years:
                yr_val = _non_seg(sn, 'ebitda', yr_key)
                if yr_val: A(ws, R, col, sc(yr_val), fmt=NUM)
                else: C(ws, R, col, '', fmt=NUM)
            A(ws, R, FY0, sc(sebitda), fmt=NUM)
            if has_q:
                for qi in range(q_actual_n):
                    qk = f'q{qi+1}'
                    q_ebitda = seg.get('quarters', {}).get(qk, {}).get('ebitda', 0)
                    if q_ebitda: A(ws, R, Q_START + qi, sc(q_ebitda), fmt=NUM)
            ebitda_r_s1 = R
            C(ws, R, 3, 'EBITDA', font=bf)
            R += 1
            # EBITDA YoY
            C(ws, R, DS, '', fmt=PCT)
            _cle = get_column_letter(DS + 1); _cld = get_column_letter(DS)
            C(ws, R, DS + 1, f'=IFERROR({_cle}{ebitda_r_s1}/{_cld}{ebitda_r_s1}-1,"")', fmt=PCT)
            _f0 = get_column_letter(FY0); _f1 = get_column_letter(FY0 - 1)
            C(ws, R, FY0, f'=IFERROR({_f0}{ebitda_r_s1}/{_f1}{ebitda_r_s1}-1,"")', fmt=PCT)
            for ci in range(FY0 + 1, LC_ANNUAL + 1):
                cl = get_column_letter(ci); pl = get_column_letter(ci - 1)
                C(ws, R, ci, f'=IFERROR({cl}{ebitda_r_s1}/{pl}{ebitda_r_s1}-1,"")', fmt=PCT)
            C(ws, R, 3, 'EBITDA YoY', font=nf)
            R += 1
            # EBITDA QoQ
            if has_q:
                for ci in range(DS, LC_ANNUAL + 1):
                    C(ws, R, ci, '', fmt=PCT)
                for qi in range(Q_START, Q_END + 1):
                    cl = get_column_letter(qi); pl = get_column_letter(qi - 1)
                    if qi == Q_START: C(ws, R, qi, '', fmt=PCT)
                    else: C(ws, R, qi, f'=IFERROR({cl}{ebitda_r_s1}/{pl}{ebitda_r_s1}-1,"")', fmt=PCT)
                C(ws, R, 3, '  EBITDA QoQ', font=itf)
                R += 1
            # EBITDA margin = EBITDA / Revenue (formula)
            for ci in range(DS, ALL_END + 1):
                cl = get_column_letter(ci)
                C(ws, R, ci, f'=IFERROR({cl}{ebitda_r_s1}/{cl}{rev_r},"")', fmt=PCT)
            C(ws, R, 3, 'EBITDA margin', font=nf)
            margin_r = R; R += 1
            cost_r = 0
            gp_r = ebitda_r_s1  # alias for seg_info
        else:
            # ── GP block ──
            _gp_r = R + 2
            for ci in range(DS, ALL_END + 1):
                cl = get_column_letter(ci)
                C(ws, R, ci, f'=IFERROR({cl}{rev_r}-{cl}{_gp_r},"")', fmt=NUM)
            cost_r = R
            C(ws, R, 3, 'Cost', font=bf)
            R += 1

            for ci in range(DS, ALL_END + 1):
                cl = get_column_letter(ci)
                CF(ws, R, ci, f'=IFERROR({cl}{R + 1}/{cl}{rev_r},"")', fmt=PCT)
            margin_r = R
            C(ws, R, 3, 'GM', font=nf)
            R += 1

            for yr_key, col in hist_years:
                yr_val = _gaap_seg(sn, 'gp', yr_key)
                if yr_val: A(ws, R, col, sc(yr_val), fmt=NUM)
                else: C(ws, R, col, '', fmt=NUM)
            A(ws, R, FY0, sc(sgp), fmt=NUM)
            for qi in range(q_actual_n):
                qk = f'q{qi+1}'; qd = seg_quarters.get(qk, {})
                if qd.get('gp'): A(ws, R, Q_START + qi, sc(qd['gp']), fmt=NUM)
            for qi in range(q_proj_n):
                C(ws, R, Q_START + q_actual_n + qi, '', fmt=NUM)
            gp_r = R
            C(ws, R, 3, 'GP', font=bf)
            R += 1
            # GP YoY
            C(ws, R, DS, '', fmt=PCT)
            _cle = get_column_letter(DS + 1); _cld = get_column_letter(DS)
            C(ws, R, DS + 1, f'=IFERROR({_cle}{gp_r}/{_cld}{gp_r}-1,"")', fmt=PCT)
            _f0 = get_column_letter(FY0); _f1 = get_column_letter(FY0 - 1)
            C(ws, R, FY0, f'=IFERROR({_f0}{gp_r}/{_f1}{gp_r}-1,"")', fmt=PCT)
            for ci in range(FY0 + 1, LC_ANNUAL + 1):
                cl = get_column_letter(ci); pl = get_column_letter(ci - 1)
                C(ws, R, ci, f'=IFERROR({cl}{gp_r}/{pl}{gp_r}-1,"")', fmt=PCT)
            C(ws, R, 3, 'GP YoY', font=nf)
            R += 1
            # GP QoQ
            if has_q:
                for ci in range(DS, LC_ANNUAL + 1):
                    C(ws, R, ci, '', fmt=PCT)
                for qi in range(Q_START, Q_END + 1):
                    cl = get_column_letter(qi); pl = get_column_letter(qi - 1)
                    if qi == Q_START: C(ws, R, qi, '', fmt=PCT)
                    else: C(ws, R, qi, f'=IFERROR({cl}{gp_r}/{pl}{gp_r}-1,"")', fmt=PCT)
                C(ws, R, 3, '  GP QoQ', font=itf)
                R += 1
        # Opex = GP - OI (OP depth only, OI = R + 1)
        op_r_est = R + 1
        if not is_ebitda_depth and _gaap_seg(sn, 'oi', FY0_KEY) is not None:
            for ci in range(DS, ALL_END + 1):
                cl = get_column_letter(ci)
                C(ws, R, ci, f'=IFERROR({cl}{gp_r}-{cl}{op_r_est},"")', fmt=NUM)
            C(ws, R, 3, 'Opex', font=bf)
            R += 1

        # OP row (if segment discloses operating profit)
        fy0_op = _gaap_seg(sn, 'oi', FY0_KEY)
        op_r = 0
        if fy0_op is not None:
            for yr_key, col in hist_years:
                yr_val = _gaap_seg(sn, 'oi', yr_key)
                if yr_val is not None: A(ws, R, col, sc(yr_val), fmt=NUM)
                else: C(ws, R, col, '', fmt=NUM)
            A(ws, R, FY0, sc(fy0_op), fmt=NUM)
            for qi in range(q_actual_n):
                qk = f'q{qi+1}'; qd = seg_quarters.get(qk, {})
                _op = qd.get('op')
                if not _op and qd.get('gp') and qd.get('rev'):
                    _op_rate = _gl('opex_rev', FY0_KEY)  # FY0 rate
                    _op = round(qd['gp'] - qd['rev'] * _op_rate)
                if _op is not None: A(ws, R, Q_START + qi, sc(_op), fmt=NUM)
            for qi in range(q_proj_n):
                C(ws, R, Q_START + q_actual_n + qi, '', fmt=NUM)
            op_r = R
            C(ws, R, 3, 'OI', font=bf)
            R += 1
            # OI YoY
            for yr_key, col in hist_years:
                C(ws, R, col, '', fmt=PCT)
            CF(ws, R, DS + 1, f'=IFERROR({get_column_letter(DS+1)}{op_r}/{get_column_letter(DS)}{op_r}-1,"")', fmt=PCT)
            CF(ws, R, FY0, f'=IFERROR({get_column_letter(FY0)}{op_r}/{get_column_letter(FY0-1)}{op_r}-1,"")', fmt=PCT)
            for ci in range(FY0 + 1, LC_ANNUAL + 1):
                cl = get_column_letter(ci); pl = get_column_letter(ci - 1)
                C(ws, R, ci, f'=IFERROR({cl}{op_r}/{pl}{op_r}-1,"")', fmt=PCT)
            # Q YoY (vs 4Q back)
            if has_q:
                for qi in range(Q_START, Q_END + 1):
                    cl = get_column_letter(qi)
                    if qi - 4 >= Q_START:
                        pl = get_column_letter(qi - 4)
                        C(ws, R, qi, f'=IFERROR({cl}{op_r}/{pl}{op_r}-1,"")', fmt=PCT)
                    else:
                        C(ws, R, qi, '', fmt=PCT)
            C(ws, R, 3, 'OI YoY', font=nf)
            R += 1
            # OI QoQ (collapsed, Q columns only)
            if has_q:
                for ci in range(DS, LC_ANNUAL + 1):
                    C(ws, R, ci, '', fmt=PCT)
                for qi in range(Q_START, Q_END + 1):
                    cl = get_column_letter(qi); pl = get_column_letter(qi - 1)
                    if qi == Q_START: C(ws, R, qi, '', fmt=PCT)
                    else: C(ws, R, qi, f'=IFERROR({cl}{op_r}/{pl}{op_r}-1,"")', fmt=PCT)
                C(ws, R, 3, '  OI QoQ', font=itf)
                R += 1
            # OPM
            for ci in range(DS, ALL_END + 1):
                cl = get_column_letter(ci)
                C(ws, R, ci, f'=IFERROR({cl}{op_r}/{cl}{rev_r},"")', fmt=PCT)
            C(ws, R, 3, 'OPM', font=nf)
            R += 1


        # Line Revenue Split
        R += 1
        C(ws, R, 3, 'Line Revenue Split', font=bf)
        R += 1

        srows = []; lrevs = {}
        for l in lls:
            ln = l['name']; sp = l['split']
            # Split% — same value across all historical years + FY0 + Q
            for col in (DS, DS + 1, FY0):
                I(ws, R, col, sp, fmt=PCT)
            if has_q:
                for qi in range(Q_START, Q_END + 1):
                    I(ws, R, qi, sp, fmt=PCT)
            C(ws, R, 3, f'  {ln} %', font=itf)
            srows.append(R); R += 1
            lr = round(srev * sp)
            anchor_info[ln] = (R, sc(lr))
            C(ws, R, 3, f'  {ln} FY{bfyr}A Rev', font=itf)
            split_row = srows[-1]
            line_to_seg[ln] = sn
            line_to_split[ln] = split_row
            for col in (DS, DS + 1, FY0):
                cl = get_column_letter(col)
                CF(ws, R, col, f'={cl}{rev_r}*{cl}{split_row}', fmt=NUM)
            if has_q:
                for qi in range(Q_START, Q_END + 1):
                    cl = get_column_letter(qi)
                    CF(ws, R, qi, f'={cl}{rev_r}*{cl}{split_row}', fmt=NUM)
            lrevs[ln] = R; R += 1

        if srows:
            for col in (DS, DS + 1, FY0):
                cl = get_column_letter(col)
                refs = '+'.join([f'{cl}{sr}' for sr in srows])
                CF(ws, R, col, f'=1-({refs})', fmt=PCT)
            if has_q:
                for qi in range(Q_START, Q_END + 1):
                    cl = get_column_letter(qi)
                    refs = '+'.join([f'{cl}{sr}' for sr in srows])
                    CF(ws, R, qi, f'=1-({refs})', fmt=PCT)
            C(ws, R, 3, '  residual %', font=itf)
            res_row = R; R += 1
        else:
            res_row = 0

        # ── Segment Check rows (multi-line segments only, below residual) ──
        seg_check_gp_r = 0; seg_check_op_r = 0
        if len(lls) >= 2:
            for ci in range(DS, ALL_END + 1):
                C(ws, R, ci, '', fmt=PCT)
            C(ws, R, 3, '  Check Seg GP', font=itf)
            ws.row_dimensions.group(R, R, outline_level=1, hidden=True)
            seg_check_gp_r = R; R += 1
            if op_r:
                for ci in range(DS, ALL_END + 1):
                    C(ws, R, ci, '', fmt=PCT)
                C(ws, R, 3, '  Check Seg OP', font=itf)
                ws.row_dimensions.group(R, R, outline_level=1, hidden=True)
                seg_check_op_r = R; R += 1

        seg_info[sn] = {
            'rev': rev_r, 'cost': cost_r, 'gp': gp_r, 'gm': margin_r,
            'split_rows': srows, 'lrev_rows': lrevs, 'res_row': res_row,
            'seg_check_gp_r': seg_check_gp_r, 'seg_check_op_r': seg_check_op_r,
        }
        if op_r: seg_info[sn]['op'] = op_r
        # Blank row between segments
        R += 1
        # Mark yoy 1:1 lines (segment=line, no residual, no vol_asp fit)
        if len(lls) == 1 and lls[0]['split'] == 1.0:
            one_to_one.add(lls[0]['name'])

    # Detect deepest profit level disclosed across all segments (numeric ordering)
    DEPTH_RANK = {'gp': 0, 'op': 1, 'ni': 2}
    max_seg_depth = 0
    for seg in segments:
        fy0 = seg.get('fy0', {})
        if _gaap_seg(sn, 'oi', FY0_KEY) is not None and DEPTH_RANK['op'] > max_seg_depth: max_seg_depth = DEPTH_RANK['op']
        if _gaap_seg(sn, 'ni', FY0_KEY) is not None and DEPTH_RANK['ni'] > max_seg_depth: max_seg_depth = DEPTH_RANK['ni']

    # ═══════════════ §2 Logic Lines ═══════════════
    s1_end = R  # Section 1 rows end here
    R += 1
    C(ws, R, 1, 'Logic Lines', font=bf12)
    R += 1
    # Basis label (below Logic Lines header)
    basis = meta.get('basis', 'gaap')
    basis_note = meta.get('basis_note', '')
    basis_label = f'Basis: {basis.upper()}'
    if basis_note:
        basis_label += f' — {basis_note[:120]}'
    C(ws, R, 1, basis_label, font=itf)
    R += 1
    L = {}
    protected_rows = set()
    for line_idx, ll in enumerate(logic_lines):
        ln = ll['name']
        module_name = ll.get('module', 'yoy')

        # Dispatch to module
        _load_module(module_name)
        render_fn = MODULES[module_name]

        # ── Inject per-line helpers into ctx (module renderers use these instead of ll['...']) ──
        ctx['li'] = line_idx
        ctx['_PROJ_FYS'] = _PROJ_FYS  # dynamic projection FY keys
        ctx['_br_li'] = lambda fy: _br(line_idx, fy)
        ctx['_yoy_li'] = lambda sc, fy: _yoy(line_idx, sc, fy)
        ctx['_vol_li'] = lambda fy: _vol(line_idx, fy)
        ctx['_asp_li'] = lambda fy, ti=0, sc=None: _asp(line_idx, fy, ti, sc)
        ctx['_share_li'] = lambda fy, ti=0: _share(line_idx, fy, ti)
        ctx['_cap_li'] = lambda fy: _capacity(line_idx, fy)
        ctx['_util_li'] = lambda fy: _utilization(line_idx, fy)
        ctx['_bb_li'] = lambda field, fy, sc=None: _bb_field(line_idx, field, fy, sc)
        ctx['_has_bb_li'] = lambda ti=0: _has_bb(line_idx, ti)
        ctx['_us_li'] = _unit_scale(line_idx)
        ctx['_has_module_li'] = lambda m: _has_module(line_idx, m)

        result = render_fn(ws, R, ll, anchor_info, ctx)
        R = result['next_R']
        line_op_r = 0  # may be set by per-line profit chain below

        # ── Resolve line→segment (O(1) via dict) ──
        seg_name = line_to_seg.get(ln, '')
        si = seg_info.get(seg_name, {})
        s1_gp_row = si.get('gp', 0)
        s1_rev_row = si.get('rev', 0)
        s1_op_row = si.get('op', 0)
        split_r = line_to_split.get(ln, 0)
        lrev_row = si.get('lrev_rows', {}).get(ln, 0)
        seg_obj = None
        for s in segments:
            if s['name'] == seg_name:
                seg_obj = s
                break

        # ── Common: GM + GP (all modules) ──
        # Cost = Rev - GP (GP/OP depth only, GP = R + 2: Cost→GM→GP)
        if not is_ebitda_depth:
            for ci in range(DS, ALL_END + 1):
                cl = get_column_letter(ci)
                C(ws, R, ci, f'=IFERROR({cl}{result["rev_r"]}-{cl}{R + 2},"")', fmt=NUM)
            C(ws, R, 3, 'Cost', font=bf)
            cost_s2_r = R; R += 1
        else:
            cost_s2_r = 0

        if ln in one_to_one:
            # 1:1 → reference S1 margin row directly
            for yr_fy, col in [(FY2_KEY, DS), (FY1_KEY, DS + 1), (FY0_KEY, FY0)]:
                if col == FY0 or _gaap_seg(seg_name, 'rev', yr_fy):
                    cl = get_column_letter(col)
                    C(ws, R, col, f'={cl}{si["gm"]}', fmt=PCT)
        else:
            # Non-1:1 → I() assumption (blend-aware via _gm cache)
            for yr_fy, col in [(FY2_KEY, DS), (FY1_KEY, DS + 1)]:
                br_val = _gm(line_idx, yr_fy)
                if br_val: I(ws, R, col, br_val, fmt=PCT)
                else: C(ws, R, col, '', fmt=PCT)
            br_fy0 = _gm(line_idx, FY0_KEY)
            if br_fy0: I(ws, R, FY0, br_fy0, fmt=PCT)
        for i, proj_fy in enumerate(_PROJ_FYS):
            I(ws, R, FY0 + 1 + i, _gm(line_idx, proj_fy), fmt=PCT)
        C(ws, R, 3, 'EBITDA margin' if is_ebitda_depth else 'GM')
        margin_r = R; R += 1

        for ci in range(DS, ALL_END + 1):
            cl = get_column_letter(ci)
            C(ws, R, ci, f'=IFERROR({cl}{result["rev_r"]}*{cl}{margin_r},"")', fmt=NUM)
        C(ws, R, 3, 'EBITDA' if is_ebitda_depth else 'GP')
        gp_r = R; R += 1
        # GP/EBITDA YoY
        C(ws, R, DS, '', fmt=PCT)
        _cl2e = get_column_letter(DS + 1); _cl2d = get_column_letter(DS)
        C(ws, R, DS + 1, f'=IFERROR({_cl2e}{gp_r}/{_cl2d}{gp_r}-1,"")', fmt=PCT)
        _cf2_0 = get_column_letter(FY0); _cf2_1 = get_column_letter(FY0 - 1)
        C(ws, R, FY0, f'=IFERROR({_cf2_0}{gp_r}/{_cf2_1}{gp_r}-1,"")', fmt=PCT)
        for ci in range(FY0 + 1, LC_ANNUAL + 1):
            cl = get_column_letter(ci); pl = get_column_letter(ci - 1)
            C(ws, R, ci, f'=IFERROR({cl}{gp_r}/{pl}{gp_r}-1,"")', fmt=PCT)
        C(ws, R, 3, ('EBITDA' if is_ebitda_depth else 'GP') + ' YoY')
        R += 1
        # GP QoQ
        if has_q:
            for ci in range(DS, LC_ANNUAL + 1):
                C(ws, R, ci, '', fmt=PCT)
            for qi in range(Q_START, Q_END + 1):
                cl = get_column_letter(qi); pl = get_column_letter(qi - 1)
                if qi == Q_START: C(ws, R, qi, '', fmt=PCT)
                else: C(ws, R, qi, f'=IFERROR({cl}{gp_r}/{pl}{gp_r}-1,"")', fmt=PCT)
            C(ws, R, 3, '  ' + ('EBITDA' if is_ebitda_depth else 'GP') + ' QoQ', font=itf)
            R += 1

        # ── Check rows: anchor reference for agent-driven diagnosis ──
        # Each check cell = gap% between model and anchor, agent scans pattern.
        # Format: Check Rev = (ModelRev - AnchorRev) / ABS(AnchorRev) → PCT
        rev_module = ll.get('module', 'yoy')
        needs_rev_check = True  # all lines get Rev check
        needs_gp_check = ln not in one_to_one
        hist_rev_ok = lrev_row and s1_rev_row and split_r
        hist_gp_ok = lrev_row and s1_gp_row and split_r

        def _chk_gap(ws, r, col, model_r, anchor_r, fmt=PCT):
            """Write gap% formula: (Model - Anchor) / ABS(Anchor)"""
            cl = get_column_letter(col)
            CF(ws, r, col,
               f'=IFERROR(({cl}{model_r}-{cl}{anchor_r})/ABS({cl}{anchor_r}),"")', fmt=fmt)

        # Check GP/GM removed from Section 2 — validated in Section 1 via Segment Aggregation

        # ── Check Util: for vol_asp lines with capacity ──
        cap_r = result.get('cap_r', 0)
        if cap_r and result.get('util_r', 0):
            # Util % is already Vol/Cap at result['util_r']
            # Check: is Util > 100%? Flag pattern for agent.
            # Re-use util_r as the check — agent reads it directly.
            check_util_r = result['util_r']
        else:
            check_util_r = 0

        # ── Check Util: for vol_asp lines with capacity ──
        cap_r = result.get('cap_r', 0)
        if cap_r and result.get('util_r', 0):
            check_util_r = result['util_r']
        else:
            check_util_r = 0

        # ── Per-line profit chain (gated by segment disclosure depth) ──
        line_ni_r = 0
        if max_seg_depth >= DEPTH_RANK['op']:
            # Opex = GP - OI (formula, hidden), OI = R+2 (Opex→OPM→OI)
            _oi_r_expected = R + 2
            for ci in range(DS, ALL_END + 1):
                cl = get_column_letter(ci)
                C(ws, R, ci, f'=IFERROR({cl}{gp_r}-{cl}{_oi_r_expected},"")', fmt=NUM)
            C(ws, R, 3, 'Opex', font=bf)
            R += 1

            # OPM = OI/Rev (operating margin assumption)
            line_om_arr = ll.get('opm')
            if line_om_arr:
                om_rates = line_om_arr
            else:
                om_rates = [_opm(fy) for fy in _ALL_FYS]
            _om_r = R
            for yr_i, col in [(0, DS), (1, DS + 1), (2, FY0)]:
                cl = get_column_letter(col)
                if (ln in one_to_one) and s1_gp_row and s1_op_row and s1_rev_row:
                    CF(ws, R, col, f'=IFERROR(({cl}{s1_gp_row}-{cl}{s1_op_row})/{cl}{s1_rev_row},"")', fmt=PCT)
                elif yr_i < len(om_rates):
                    I(ws, R, col, om_rates[yr_i], fmt=PCT)
                else:
                    C(ws, R, col, '', fmt=PCT)
            for i in range(proj_n):
                ri = 3 + i
                if ri < len(om_rates):
                    I(ws, R, FY0 + 1 + i, om_rates[ri], fmt=PCT)
            C(ws, R, 3, 'OPM', font=nf)
            R += 1
            # OI = Rev × OPM (historical: S1 OP×split for actual years)
            for ci in range(DS, ALL_END + 1):
                cl = get_column_letter(ci)
                is_hist = ci <= FY0 or (has_q and Q_START <= ci < Q_START + q_actual_n)
                if is_hist and s1_op_row and split_r:
                    if ln in one_to_one:
                        CF(ws, R, ci, '=IFERROR(%s%d,"")' % (cl, s1_op_row), fmt=NUM)
                    else:
                        CF(ws, R, ci, '=IFERROR(%s%d*%s%d,"")' % (cl, s1_op_row, cl, split_r), fmt=NUM)
                else:
                    CF(ws, R, ci, '=IFERROR(%s%d*%s%d,"")' % (cl, result['rev_r'], cl, _om_r), fmt=NUM)
            C(ws, R, 3, 'OI', font=bf)
            line_op_r = R; R += 1
            if _om_r: protected_rows.add(_om_r)
            # OI YoY
            C(ws, R, DS, '', fmt=PCT)
            _cl3e = get_column_letter(DS + 1); _cl3d = get_column_letter(DS)
            C(ws, R, DS + 1, f'=IFERROR({_cl3e}{line_op_r}/{_cl3d}{line_op_r}-1,"")', fmt=PCT)
            _cf3_0 = get_column_letter(FY0); _cf3_1 = get_column_letter(FY0 - 1)
            C(ws, R, FY0, f'=IFERROR({_cf3_0}{line_op_r}/{_cf3_1}{line_op_r}-1,"")', fmt=PCT)
            for ci in range(FY0 + 1, LC_ANNUAL + 1):
                cl = get_column_letter(ci); pl = get_column_letter(ci - 1)
                C(ws, R, ci, f'=IFERROR({cl}{line_op_r}/{pl}{line_op_r}-1,"")', fmt=PCT)
            C(ws, R, 3, 'OI YoY', font=nf)
            R += 1
            # OI QoQ
            if has_q:
                for ci in range(DS, LC_ANNUAL + 1):
                    C(ws, R, ci, '', fmt=PCT)
                for qi in range(Q_START, Q_END + 1):
                    cl = get_column_letter(qi); pl = get_column_letter(qi - 1)
                    if qi == Q_START: C(ws, R, qi, '', fmt=PCT)
                    else: C(ws, R, qi, f'=IFERROR({cl}{line_op_r}/{pl}{line_op_r}-1,"")', fmt=PCT)
                C(ws, R, 3, '  OI QoQ', font=itf)
                R += 1
        if max_seg_depth >= DEPTH_RANK['ni']:
            # NM assumed → NI = Rev × NM → Tax = OI - NI (derived)
            line_nm = ll.get('tax_rate')  # JSON field: NM array
            if line_nm and isinstance(line_nm, list):
                nm_rates = line_nm
            else:
                nm_rates = [_gl('tax_rate', fy) for fy in _ALL_FYS]
            _nm_r = R
            for yr_i, col in [(0, DS), (1, DS + 1), (2, FY0)]:
                cl = get_column_letter(col)
                if yr_i < len(nm_rates):
                    I(ws, R, col, nm_rates[yr_i], fmt=PCT)
                else:
                    C(ws, R, col, '', fmt=PCT)
            for i in range(proj_n):
                ri = 3 + i
                if ri < len(nm_rates):
                    I(ws, R, FY0 + 1 + i, nm_rates[ri], fmt=PCT)
            C(ws, R, 3, '  NM', font=itf)
            R += 1
            # NI = Rev × NM (same pattern as GP = Rev × GM)
            for ci in range(DS, ALL_END + 1):
                cl = get_column_letter(ci)
                CF(ws, R, ci, f'=IFERROR({cl}{result["rev_r"]}*{cl}{_nm_r},"")', fmt=NUM)
            C(ws, R, 3, '  NI', font=bf)
            line_ni_r = R; R += 1

        # ── Check Rev: at end of line section, Q→Y + split verification ──
        check_rev_r = 0
        if lrev_row and s1_rev_row and split_r:
            for col in (DS, DS + 1, FY0):
                _chk_gap(ws, R, col, result['rev_r'], lrev_row)
            for ci in range(FY0 + 1, LC_ANNUAL + 1):
                _chk_gap(ws, R, ci, result['rev_r'], lrev_row)
            gm_fy0 = _br(line_idx, FY0_KEY)
            gm_label = f' ({gm_fy0:.0%} ' + ('EBITDA margin' if is_ebitda_depth else 'GM') + ')' if gm_fy0 else ''
            C(ws, R, 3, f'  Check Rev{gm_label}', font=itf)
            ws.row_dimensions.group(R, R, outline_level=1, hidden=True)
            check_rev_r = R; R += 1

        result['gm_r'] = margin_r
        result['gp_r'] = gp_r
        result['op_r'] = line_op_r
        result['ni_r'] = 0  # NI depth removed
        result['next_R'] = R
        L[ln] = result
        # Protect Rev, GM, GP, OI, NI from D/E clear
        protected_rows.update([result['rev_r'], margin_r, gp_r])
        if line_op_r: protected_rows.add(line_op_r)
        # line_ni_r protection removed
        # Store check row refs for agent diagnosis (Section 2: Check Rev + Util only)
        result['check_rev_r'] = check_rev_r
        result['check_util_r'] = check_util_r
        if check_rev_r: protected_rows.add(check_rev_r)
        for scan_r in range(result['rev_r'], R):
            cv = str(ws.cell(row=scan_r, column=3).value or '')
            if any(kw in cv for kw in ('Check', 'YoY', 'Opex', 'OP', 'Tax', 'NI')):
                protected_rows.add(scan_r)

        # ── Q Columns: mirror Y logic for actual + projection ──
        if has_q:
            q_hist = ll.get('q_history', {})
            rev_module = ll.get('module', 'yoy')
            is_vol_asp = rev_module == 'vol_asp'

            if is_vol_asp:
                # Q Volume row (mirrors annual Volume) — pre-fill from annual or q_history
                vol_fy0 = _vol(line_idx, FY0_KEY)
                for qi in range(q_actual_n + q_proj_n):
                    col = Q_START + qi
                    qv = q_hist.get(f'q{qi+1}', {}).get('volume')  # check all Qs
                    if qv is not None:
                        I(ws, result['vol_r'], col, qv, fmt=INT)
                    elif qi < q_actual_n:
                        I(ws, result['vol_r'], col, round(vol_fy0 / 4), fmt=INT)
                    else:
                        # Fallback: use annual volume for this Q's FY (divided evenly across 4 Qs)
                        fy_year = q_start_yr + (q_start_q - 1 + qi) // 4
                        fy_ofs = fy_year - bfyr - 1  # 0=FY26, 1=FY27, ...
                        proj_fy_v = _PROJ_FYS[fy_ofs] if 0 <= fy_ofs < len(_PROJ_FYS) else _PROJ_FYS[-1]
                        ann_vol = _vol(line_idx, proj_fy_v)
                        I(ws, result['vol_r'], col, round(ann_vol / 4), fmt=INT)
                # Q ASP row (first tier, mirrors annual ASP) — pre-fill from annual or q_history
                asp_r = result.get('asp_rows', [0])[0] if result.get('asp_rows') else 0
                if asp_r:
                    is_bb = _has_bb(line_idx, 0)
                    # BBE tiers: use asp_base for Q distribution (same as driver)
                    if is_bb:
                        asp_fy0 = _asp(line_idx, FY0_KEY, 0, 'base')
                    else:
                        asp_fy0 = _asp(line_idx, FY0_KEY, 0)
                    _asp_fy0_val = _asp(line_idx, FY0_KEY, 0)
                    for qi in range(q_actual_n + q_proj_n):
                        col = Q_START + qi
                        q_asp = q_hist.get(f'q{qi+1}', {}).get('asp')  # check all Qs
                        if q_asp is not None:
                            I(ws, asp_r, col, q_asp, fmt=DEC)
                        elif qi < q_actual_n:
                            I(ws, asp_r, col, _asp_fy0_val, fmt=DEC)
                        else:
                            # Fallback: use annual ASP for the Q's fiscal year
                            fy_year = q_start_yr + (q_start_q - 1 + qi) // 4
                            asp_fy = f'FY{fy_year}E' if fy_year > bfyr else f'FY{fy_year}'
                            I(ws, asp_r, col, _asp(line_idx, asp_fy, 0), fmt=DEC)
                # Q Revenue: =Q_Vol × Q_ASP / scale (mirrors annual formula)
                for qi in range(q_actual_n + q_proj_n):
                    col = Q_START + qi; cl = get_column_letter(col)
                    scale = ll.get('unit_scale', 100)
                    if asp_r:
                        CF(ws, result['rev_r'], col, f'=({cl}{result["vol_r"]}*{cl}{asp_r})/{scale}', fmt=NUM)
                    else:
                        CF(ws, result['rev_r'], col, '', fmt=NUM)
            else:
                # yoy: Q Revenue = prior_Q × (1+YoY_Active) for proj, I() for actual
                # Extend YoY Active row to Q columns (scenario-driven)
                ya = result.get('ya', 0)
                if ya and result.get('yb'):
                    for qi in range(q_actual_n + q_proj_n):
                        col = Q_START + qi; cl = get_column_letter(col)
                        CF(ws, ya, col,
                           f'=IF(B1="Bull",{cl}{result["yb"]},IF(B1="Bear",{cl}{result["ybe"]},{cl}{result["ybs"]}))',
                           fmt=PCT)
                for qi in range(q_actual_n + q_proj_n):
                    col = Q_START + qi; cl = get_column_letter(col)
                    is_q_actual = qi < q_actual_n
                    s1r = anchor_info.get(ln, (0, 0))[0]
                    anchor_row = s1r if ln in one_to_one else lrev_row
                    # 1:1 actual Q: use S1 reference like GM/OpexRev (real data)
                    # Non-core: use q_history hardcodes for all Qs (residual distribution avoids chain compounding)
                    if ln == 'Non-core' and qi < q_actual_n:
                        qv = q_hist.get(f'q{qi+1}', {}).get('rev')
                        if qv is not None:
                            I(ws, result['rev_r'], col, sc(qv), fmt=NUM)
                        elif anchor_row:
                            C(ws, result['rev_r'], col, f'={cl}{anchor_row}', fmt=NUM)
                    elif ln in one_to_one and anchor_row and is_q_actual:
                        C(ws, result['rev_r'], col, f'={cl}{anchor_row}', fmt=NUM)
                    elif is_q_actual:
                        # non-1:1 actual Q: q_history rev or S1 anchor
                        qv = q_hist.get(f'q{qi+1}', {}).get('rev')
                        if qv is not None:
                            I(ws, result['rev_r'], col, sc(qv), fmt=NUM)
                        elif anchor_row:
                            C(ws, result['rev_r'], col, f'={cl}{anchor_row}', fmt=NUM)
                    elif ln == 'Non-core':
                        # Non-core proj Q: use q_history hardcode (residual avoids chain compounding)
                        qv = q_hist.get(f'q{qi+1}', {}).get('rev')
                        if qv is not None:
                            I(ws, result['rev_r'], col, sc(qv), fmt=NUM)
                        else:
                            # fallback to chain formula
                            pl = get_column_letter(col - 1)
                            CF(ws, result['rev_r'], col,
                               f'={pl}{result["rev_r"]}*(1+{cl}{ya})', fmt=NUM)
                    else:
                        # proj Q: chain formula (never use stale q_history rev)
                        pl = get_column_letter(col - 1)
                        CF(ws, result['rev_r'], col,
                           f'={pl}{result["rev_r"]}*(1+{cl}{ya})', fmt=NUM)
                # Extend BBE rows to Q columns (q_history yoy_q if available, else annual rate → QoQ)
                if result.get('yb'):
                    yoy_keys = [('bull', 'yb'), ('base', 'ybs'), ('bear', 'ybe')]
                    cur_yr, cur_q = q_start_yr, q_start_q
                    for qi in range(q_actual_n + q_proj_n):
                        proj_idx = cur_yr - bfyr - 1
                        qk = f'q{qi+1}'
                        # q_history yoy_q takes priority (per-Q rate from driver distribution)
                        q_yoy = q_hist.get(qk, {}).get('yoy_q')
                        if q_yoy is not None:
                            for arr_key, rk in yoy_keys:
                                I(ws, result[rk], Q_START + qi, q_yoy, fmt=PCT)
                        elif 0 <= proj_idx < proj_n:
                            # Fallback: convert annual rate to QoQ
                            _ann_rate = _yoy(line_idx, 'base', _PROJ_FYS[proj_idx]) if proj_idx < len(_PROJ_FYS) else 0
                            _q_rate = (1 + _ann_rate) ** (1 / 4) - 1 if _ann_rate > -0.99 else 0.02
                            for arr_key, rk in yoy_keys:
                                I(ws, result[rk], Q_START + qi, _q_rate, fmt=PCT)
                        cur_q += 1
                        if cur_q > 4: cur_q = 1; cur_yr += 1
            # Q GP/OP cascade (same as annual)
            for qi in range(Q_START, Q_END + 1):
                cl = get_column_letter(qi)
                CF(ws, gp_r, qi, f'=IFERROR({cl}{result["rev_r"]}*{cl}{margin_r},"")', fmt=NUM)
            if line_op_r:
                for qi in range(Q_START, Q_END + 1):
                    cl = get_column_letter(qi)
                    CF(ws, line_op_r, qi, f'=IFERROR({cl}{result["rev_r"]}*{cl}{_om_r},"")', fmt=NUM)
        # ── Check rows: extend Check Rev to Q columns ──
        if has_q:
            for scan_r in range(result['rev_r'], R):
                cv = str(ws.cell(row=scan_r, column=3).value or '')
                if 'Check Rev' in cv and lrev_row:
                    for qi in range(q_actual_n + q_proj_n):
                        col = Q_START + qi; cl = get_column_letter(col)
                        _chk_gap(ws, scan_r, col, result['rev_r'], lrev_row)

    # ═══════════════ §2→§1 Fill ─ Section 1 FY26E+ ═══════════════
    for seg in segments:
        sn = seg['name']; lls = seg['logic_lines']
        res_val = round(_gaap_seg(seg['name'], 'rev', FY0_KEY) -
                        sum(_gaap_seg(seg['name'], 'rev', FY0_KEY) * l['split'] for l in lls))
        res_val = res_val if res_val > 0 else 0
        res_gm = seg.get('residual', {}).get('gm', 0)
        si = seg_info.get(sn, {})
        s1r = si.get('rev', 0); s1c = si.get('cost', 0)
        s1g = si.get('gp', 0); s1gm = si.get('gm', 0)
        if not s1r or not s1g:
            continue
        srows = si.get('split_rows', [])
        lrevs = si.get('lrev_rows', {})
        res_row = si.get('res_row', 0)

        for ci in range(FY0 + 1, LC_ANNUAL + 1):
            cl = get_column_letter(ci)
            CF(ws, s1r, ci, '=' + '+'.join(
                [f'{cl}{L[ln["name"]]["rev_r"]}' for ln in lls] +
                ([str(sc(res_val))] if res_val else [])), fmt=NUM)
            if s1c:
                CF(ws, s1c, ci, f'=IFERROR({cl}{s1r}-{cl}{s1g},"")', fmt=NUM)
            CF(ws, s1g, ci, '=' + '+'.join(
                [f'{cl}{L[ln["name"]]["gp_r"]}' for ln in lls] +
                ([f'{sc(res_val)}*{res_gm}'] if res_val else [])), fmt=NUM)
            if s1gm:
                CF(ws, s1gm, ci, f'=IFERROR({cl}{s1g}/{cl}{s1r},"")', fmt=PCT)
            # OP fill (if segment discloses OP)
            s1_op = si.get('op', 0)
            if s1_op and max_seg_depth >= DEPTH_RANK['op']:
                op_terms = [f'{cl}{L[ln["name"]]["op_r"]}' for ln in lls if L[ln["name"]].get('op_r')]
                if op_terms:
                    CF(ws, s1_op, ci, '=' + '+'.join(op_terms), fmt=NUM)

        # Q proj fill (same logic as annual, for Q projection columns)
        if has_q:
            for ci in range(Q_START + q_actual_n, Q_END + 1):
                cl = get_column_letter(ci)
                CF(ws, s1r, ci, '=' + '+'.join(
                    [f'{cl}{L[ln["name"]]["rev_r"]}' for ln in lls] +
                    ([str(sc(res_val))] if res_val else [])), fmt=NUM)
                if s1c:
                    CF(ws, s1c, ci, f'=IFERROR({cl}{s1r}-{cl}{s1g},"")', fmt=NUM)
                CF(ws, s1g, ci, '=' + '+'.join(
                    [f'{cl}{L[ln["name"]]["gp_r"]}' for ln in lls] +
                    ([f'{sc(res_val)}*{res_gm}'] if res_val else [])), fmt=NUM)
                if s1gm:
                    CF(ws, s1gm, ci, f'=IFERROR({cl}{s1g}/{cl}{s1r},"")', fmt=PCT)
                s1_op = si.get('op', 0)
                if s1_op and max_seg_depth >= DEPTH_RANK['op']:
                    op_terms = [f'{cl}{L[ln["name"]]["op_r"]}' for ln in lls if L[ln["name"]].get('op_r')]
                    if op_terms:
                        CF(ws, s1_op, ci, '=' + '+'.join(op_terms), fmt=NUM)

        for ln_name in [l['name'] for l in lls]:
            for sr_row in srows:
                cell_val = ws.cell(row=sr_row, column=3).value
                if cell_val and ln_name in str(cell_val):
                    for ci in range(FY0 + 1, LC_ANNUAL + 1):
                        cl = get_column_letter(ci)
                        CF(ws, sr_row, ci, f'=IFERROR({cl}{L[ln_name]["rev_r"]}/{cl}{s1r},"")', fmt=PCT)
            lr_row = lrevs.get(ln_name, 0)
            if lr_row:
                for ci in range(FY0 + 1, LC_ANNUAL + 1):
                    cl = get_column_letter(ci)
                    CF(ws, lr_row, ci, f'={cl}{L[ln_name]["rev_r"]}', fmt=NUM)
        if res_row and srows:
            for ci in range(FY0 + 1, LC_ANNUAL + 1):
                cl = get_column_letter(ci)
                refs = '+'.join([f'{cl}{r}' for r in srows])
                CF(ws, res_row, ci, f'=1-({refs})', fmt=PCT)

        # Write Segment Check formulas (multi-line segments only)
        seg_check_gp_r = si.get('seg_check_gp_r', 0)
        seg_check_op_r = si.get('seg_check_op_r', 0)
        if seg_check_gp_r:
            for ci in range(DS, ALL_END + 1):
                cl = get_column_letter(ci)
                line_sum = '+'.join([f'{cl}{L[l["name"]]["gp_r"]}' for l in lls])
                if res_val:
                    line_sum += f'+{sc(round(res_val * res_gm))}'
                CF(ws, seg_check_gp_r, ci,
                   f'=IFERROR(({line_sum}-{cl}{s1g})/ABS({cl}{s1g}),"")', fmt=PCT)
        if seg_check_op_r:
            op_lls = [l for l in lls if L[l['name']].get('op_r')]
            if op_lls:
                for ci in range(DS, ALL_END + 1):
                    cl = get_column_letter(ci)
                    line_sum = '+'.join([f'{cl}{L[l["name"]]["op_r"]}' for l in op_lls])
                    CF(ws, seg_check_op_r, ci,
                       f'=IFERROR(({line_sum}-{cl}{s1_op})/ABS({cl}{s1_op}),"")', fmt=PCT)

    # ── Non-core absorbs all segment residuals ──
    total_residual = 0; gp_residual = 0
    for seg in segments:
        r = round(_gaap_seg(seg['name'], 'rev', FY0_KEY) - sum(_gaap_seg(seg['name'], 'rev', FY0_KEY) * l['split'] for l in seg.get('logic_lines', [])))
        if r > 0:
            total_residual += r
            gp_residual += round(r * seg.get('residual', {}).get('gm', 0))
    # Patch Non-core Revenue + GP to include residuals
    if 'Non-core' in L:
        nc = L['Non-core']
        nc_rev_r = nc['rev_r']
        nc_gp_r = nc['gp_r']
        # Annual columns: add full annual residual
        for ci in range(FY0 + 1, LC_ANNUAL + 1):
            cl = get_column_letter(ci)
            old_rev = ws.cell(row=nc_rev_r, column=ci).value or ''
            old_gp = ws.cell(row=nc_gp_r, column=ci).value or ''
            if isinstance(old_rev, str) and old_rev.startswith('='):
                CF(ws, nc_rev_r, ci, old_rev + f'+{sc(total_residual)}', fmt=NUM)
            if isinstance(old_gp, str) and old_gp.startswith('='):
                CF(ws, nc_gp_r, ci, old_gp + f'+{sc(gp_residual)}', fmt=NUM)
        # Q columns: add residual_per_q to chain formulas only (hardcoded values already include it)
        if has_q:
            residual_per_q = round(total_residual / 4)
            gp_residual_per_q = round(gp_residual / 4)
            for qi in range(q_actual_n + q_proj_n):
                col = Q_START + qi; cl = get_column_letter(col)
                old_rev = ws.cell(row=nc_rev_r, column=col).value or ''
                old_gp = ws.cell(row=nc_gp_r, column=col).value or ''
                if isinstance(old_rev, str) and old_rev.startswith('='):
                    CF(ws, nc_rev_r, col, old_rev + f'+{sc(residual_per_q)}', fmt=NUM)
                if isinstance(old_gp, str) and old_gp.startswith('='):
                    CF(ws, nc_gp_r, col, old_gp + f'+{sc(gp_residual_per_q)}', fmt=NUM)

    # EBITDA depth: Non-core Corporate - zero revenue + absorb company ebitda gap
    nc = None
    if is_ebitda_depth:
        _nc_key = next((k for k in L.keys() if 'Non-core' in k), None)
        nc = L[_nc_key] if _nc_key else None
    if nc is not None:
        nc_rev_r = nc.get('rev_r', 0)
        nc_gp_r = nc.get('gp_r', 0)
        # Zero out Revenue (Non-core has no revenue)
        if nc_rev_r:
            for ci in range(DS, ALL_END + 1):
                A(ws, nc_rev_r, ci, 0, fmt=NUM)
        # Write EBITDA = company gap (all years, including 0)
        if nc_gp_r:
            gap_fy0 = 0
            _asm_nc = raw.get('assumptions', {}).get('noncore_gap', {})
            for new_fy, col in [(FY2_KEY, DS), (FY1_KEY, DS + 1), (FY0_KEY, FY0)]:
                seg_sum = sum((_non_seg(s['name'], 'ebitda', new_fy) or 0) for s in cfg.get('segments', []))
                cv = _non('ebitda', new_fy)
                gap_val = round(cv - seg_sum)
                if new_fy == FY0_KEY: gap_fy0 = gap_val
                A(ws, nc_gp_r, col, sc(gap_val), fmt=NUM)
            # Projected FY columns: carry FY0 gap (agent override via assumptions.noncore_gap)
            for ci in range(FY0 + 1, LC_ANNUAL + 1):
                fy_proj = f'FY{bfyr + (ci - FY0)}E'
                proj_gap = _asm_nc.get(fy_proj, {}).get('annual')
                if proj_gap is not None:
                    A(ws, nc_gp_r, ci, sc(proj_gap), fmt=NUM)
                else:
                    A(ws, nc_gp_r, ci, sc(gap_fy0), fmt=NUM)
            # Q gap = company Q ebitda - Σ seg Q ebitda
            if has_q and 'quarters' in cfg:
                for i, qk in enumerate(['q1','q2','q3','q4']):
                    cq = cfg['quarters'].get(qk, {})
                    q_company = cq.get('ebitda', 0)
                    q_seg = sum(s.get('quarters',{}).get(qk,{}).get('ebitda',0) for s in cfg.get('segments',[]))
                    q_gap = round(q_company - q_seg)
                    A(ws, nc_gp_r, Q_START + i, sc(q_gap), fmt=NUM)

    # ── Q Columns: post-Fill GM + opm extension ──
    if has_q:
        for line_idx, ll in enumerate(logic_lines):
            ln = ll['name']; rows = L.get(ln)
            if not rows: continue
            rev_r = rows['rev_r']; margin_r = rows['gm_r']
            seg_name = line_to_seg.get(ln, ''); si = seg_info.get(seg_name, {})
            s1_gp_row = si.get('gp', 0); s1_rev_row = si.get('rev', 0)
            is_1to1 = ln in one_to_one
            cur_yr, cur_q = q_start_yr, q_start_q
            for qi in range(Q_START, Q_END + 1):
                cl = get_column_letter(qi)
                proj_idx = cur_yr - bfyr - 1
                is_q_actual = qi < Q_START + q_actual_n
                # GM: 1:1 actual Qs use S1 GP/Rev (real quarterly margin)
                if is_1to1 and s1_gp_row and s1_rev_row and is_q_actual:
                    CF(ws, margin_r, qi, f'=IFERROR({cl}{s1_gp_row}/{cl}{s1_rev_row},"")', fmt=PCT)
                elif is_1to1 and not is_q_actual and proj_idx < len(_PROJ_FYS):
                    I(ws, margin_r, qi, _gm(line_idx, _PROJ_FYS[proj_idx]), fmt=PCT)
                else:
                    I(ws, margin_r, qi, _gm(line_idx, FY0_KEY), fmt=PCT)
                cur_q += 1
                if cur_q > 4: cur_q = 1; cur_yr += 1
            # OPM → Q columns (1:1 actual Qs use S1 (GP−OP)/Rev)
            if ll.get('opm'):
                om_arr = ll['opm']
            else:
                om_arr = [_opm(fy) for fy in _ALL_FYS]
            cur_yr, cur_q = q_start_yr, q_start_q
            for scan_r in range(rev_r, rows.get('op_r', rev_r + 6)):
                cv = str(ws.cell(row=scan_r, column=3).value or '')
                if 'OPM' in cv:
                    for qi in range(Q_START, Q_END + 1):
                        cl = get_column_letter(qi)
                        is_q_actual = qi < Q_START + q_actual_n
                        if (ln in one_to_one) and s1_gp_row and s1_op_row and s1_rev_row and is_q_actual:
                            CF(ws, scan_r, qi, f'=IFERROR(({cl}{s1_gp_row}-{cl}{s1_op_row})/{cl}{s1_rev_row},"")', fmt=PCT)
                        else:
                            _py = cur_yr - bfyr - 1
                            _idx = max(0, min(2 + _py + 1, len(om_arr) - 1))
                            I(ws, scan_r, qi, om_arr[_idx], fmt=PCT)
                        cur_q += 1
                        if cur_q > 4: cur_q = 1; cur_yr += 1
                    break

    # Collapse Section 1 (segment rows only)
    ws.row_dimensions.group(s1_start, s1_end, outline_level=1, hidden=True)

    # ═══════════════ Rate Bridge (unified: actuals + rates + gaps, all depths) ═══════════════
    s2_end = R  # Section 2 ends before Bridge; D/E clear stops here
    R += 1  # blank row
    C(ws, R, 1, 'Rate Bridge', font=bf12)
    R += 1
    _ds_col = get_column_letter(DS)
    _fy0_col = get_column_letter(FY0)
    gap_gp_ref = gap_oi_ref = gap_ni_ref = tax_rate_ref = '0'
    rev_act_r = gp_act_r = op_act_r = ebitda_act_r = ni_act_r = tax_act_r = da_act_r = 0
    rev_act_cells = gp_act_cells = op_act_cells = ebitda_act_cells = ni_act_cells = tax_act_cells = da_act_cells = {}
    q_act_cells = {}

    # Helper for Q actuals — reads new-format actuals (FY-inside)
    def _q_act(mkey, qkey):
        if not has_q: return
        q_act_cells[qkey] = {}
        NON_GAAP_FIELDS = {'ebitda'}
        for _qi, _qk in enumerate(['Q1','Q2','Q3','Q4'][:min(q_actual_n,4)]):
            if mkey in NON_GAAP_FIELDS:
                _v = _non(mkey, FY0_KEY, _qk)
            else:
                _v = _gaap(mkey, FY0_KEY, _qk)
            if _v: A(ws, R, Q_START+_qi, sc(_v), fmt=NUM)
            q_act_cells[qkey][_qk.lower()] = R

    # ── Actuals (all depths, collapsed) — A() = gray bg actuals style ──
    for ci, new_fy in [(DS, FY2_KEY), (DS + 1, FY1_KEY), (FY0, FY0_KEY)]:
        A(ws, R, ci, sc(_gaap('rev', new_fy)), fmt=NUM)
    C(ws, R, 3, '  actuals Rev', font=itf)
    _q_act('rev', 'Revenue')
    ws.row_dimensions.group(R, R, outline_level=1, hidden=True)
    rev_act_r = R
    rev_act_cells = {fy: f'{get_column_letter(col)}{R}' for fy, col in [('fy-2', DS), ('fy-1', DS + 1), ('fy0', FY0)]}
    R += 1

    for ci, new_fy in [(DS, FY2_KEY), (DS + 1, FY1_KEY), (FY0, FY0_KEY)]:
        A(ws, R, ci, sc(_gaap('gp', new_fy)), fmt=NUM)
    C(ws, R, 3, '  actuals GP', font=itf)
    _q_act('gp', 'GP')
    ws.row_dimensions.group(R, R, outline_level=1, hidden=True)
    gp_act_r = R
    gp_act_cells = {fy: f'{get_column_letter(col)}{R}' for fy, col in [('fy-2', DS), ('fy-1', DS + 1), ('fy0', FY0)]}
    R += 1

    if not is_gp_depth:
        for ci, new_fy in [(DS, FY2_KEY), (DS + 1, FY1_KEY), (FY0, FY0_KEY)]:
            A(ws, R, ci, sc(_gaap('oi', new_fy)), fmt=NUM)
        C(ws, R, 3, '  actuals OI', font=itf)
        _q_act('op', 'OI')
        ws.row_dimensions.group(R, R, outline_level=1, hidden=True)
        op_act_r = R
        op_act_cells = {fy: f'{get_column_letter(col)}{R}' for fy, col in [('fy-2', DS), ('fy-1', DS + 1), ('fy0', FY0)]}
        R += 1

    # EBITDA depth extras: ebitda, ni, tax, da actuals
    if is_ebitda_depth:
        for ci, new_fy in [(DS, FY2_KEY), (DS + 1, FY1_KEY), (FY0, FY0_KEY)]:
            A(ws, R, ci, sc(_non('ebitda', new_fy)), fmt=NUM)
        C(ws, R, 3, '  actuals EBITDA', font=itf)
        _q_act('ebitda', 'EBITDA')
        ws.row_dimensions.group(R, R, outline_level=1, hidden=True)
        ebitda_act_r = R
        ebitda_act_cells = {fy: f'{get_column_letter(col)}{R}' for fy, col in [('fy-2', DS), ('fy-1', DS + 1), ('fy0', FY0)]}
        R += 1

        for ci, new_fy in [(DS, FY2_KEY), (DS + 1, FY1_KEY), (FY0, FY0_KEY)]:
            A(ws, R, ci, sc(_gaap('ni', new_fy)), fmt=NUM)
        C(ws, R, 3, '  actuals NI', font=itf)
        _q_act('ni', 'NI')
        ws.row_dimensions.group(R, R, outline_level=1, hidden=True)
        ni_act_r = R
        ni_act_cells = {fy: f'{get_column_letter(col)}{R}' for fy, col in [('fy-2', DS), ('fy-1', DS + 1), ('fy0', FY0)]}
        R += 1

        for ci, new_fy in [(DS, FY2_KEY), (DS + 1, FY1_KEY), (FY0, FY0_KEY)]:
            A(ws, R, ci, sc(_gaap('tax', new_fy)), fmt=NUM)
        C(ws, R, 3, '  actuals Tax', font=itf)
        _q_act('tax', 'Tax')
        ws.row_dimensions.group(R, R, outline_level=1, hidden=True)
        tax_act_r = R
        tax_act_cells = {fy: f'{get_column_letter(col)}{R}' for fy, col in [('fy-2', DS), ('fy-1', DS + 1), ('fy0', FY0)]}
        R += 1

        for ci, new_fy in [(DS, FY2_KEY), (DS + 1, FY1_KEY), (FY0, FY0_KEY)]:
            A(ws, R, ci, sc(_gaap('da', new_fy)), fmt=NUM)
        C(ws, R, 3, '  actuals D&A', font=itf)
        _q_act('da', 'D&A')
        ws.row_dimensions.group(R, R, outline_level=1, hidden=True)
        da_act_r = R
        da_act_cells = {fy: f'{get_column_letter(col)}{R}' for fy, col in [('fy-2', DS), ('fy-1', DS + 1), ('fy0', FY0)]}
        R += 1

    # ── Opex/Rev (visible, all depths) — filled in P&L tail ──
    for ci in range(DS, ALL_END + 1):
        C(ws, R, ci, 0, fmt=PCT)
    C(ws, R, 3, 'Opex/Rev', font=nf)
    opex_r = R; R += 1

    # ── (OI-NI)/Rev (visible, all depths) — filled in P&L tail ──
    for ci in range(DS, ALL_END + 1):
        C(ws, R, ci, 0, fmt=PCT)
    C(ws, R, 3, '(OI-NI)/Rev', font=nf)
    tax_r = R; R += 1

    # ── Gap rows (collapsed, EBITDA depth only) ──
    if is_ebitda_depth:
        C(ws, R, DS, f'=({ebitda_act_cells["fy0"]}-{gp_act_cells["fy0"]})/{rev_act_cells["fy0"]}', fmt=PCT)
        C(ws, R, 3, '  gap_gp (EBITDA→GP anchor)', font=itf)
        ws.row_dimensions.group(R, R, outline_level=1, hidden=True)
        gap_gp_ref = f'${_ds_col}${R}'; R += 1

        C(ws, R, DS, f'=({ebitda_act_cells["fy0"]}-{op_act_cells["fy0"]})/{rev_act_cells["fy0"]}', fmt=PCT)
        C(ws, R, 3, '  gap_oi (EBITDA→OI anchor)', font=itf)
        ws.row_dimensions.group(R, R, outline_level=1, hidden=True)
        gap_oi_ref = f'${_ds_col}${R}'; R += 1

        C(ws, R, DS, f'=({op_act_cells["fy0"]}-{ni_act_cells["fy0"]})/{rev_act_cells["fy0"]}', fmt=PCT)
        C(ws, R, 3, '  gap_ni (OI→NI anchor)', font=itf)
        ws.row_dimensions.group(R, R, outline_level=1, hidden=True)
        gap_ni_ref = f'${_ds_col}${R}'; R += 1

        C(ws, R, DS, f'={tax_act_cells["fy0"]}/{op_act_cells["fy0"]}', fmt=PCT)
        C(ws, R, 3, '  tax_rate (Tax/OI anchor)', font=itf)
        ws.row_dimensions.group(R, R, outline_level=1, hidden=True)
        tax_rate_ref = f'${_ds_col}${R}'; R += 1



    # ═══════════════ §3 P&L ═══════════════
    R += 1
    C(ws, R, 1, 'P&L', font=bf12)
    R += 1
    # Basis label (below P&L, same as Sections 1+2)
    _basis = meta.get('basis', 'gaap')
    _bnote = meta.get('basis_note', '')
    _blabel = f'Basis: {_basis.upper()}'
    if _bnote:
        _blabel += f' — {_bnote[:120]}'
    C(ws, R, 1, _blabel, font=itf)
    R += 1
    LN = [ln['name'] for ln in logic_lines]

    residual_term = f'+{sc(total_residual)}' if total_residual else ''
    gp_residual_term = f'+{sc(gp_residual)}' if gp_residual else ''
    # Q columns: Non-core already absorbs residual_per_q, no extra term needed
    q_residual_term = ''
    q_gp_residual_term = ''

    # Total Revenue (all years Σ line rev)
    for ci in range(DS, LC_ANNUAL + 1):
        cl = get_column_letter(ci)
        C(ws, R, ci, '=' + '+'.join([f'{cl}{L[ln]["rev_r"]}' for ln in LN]) + residual_term,
          fmt=NUM)
    if has_q:
        for qi in range(Q_START, Q_END + 1):
            cl = get_column_letter(qi)
            C(ws, R, qi, '=' + '+'.join([f'{cl}{L[ln]["rev_r"]}' for ln in LN]) + q_residual_term,
              fmt=NUM)
    C(ws, R, 3, 'Revenue', font=bf)
    trev = R; R += 1

    # Check Rev: (model sum - actual total rev) / actual total rev → gap%
    for ci in range(DS, ALL_END + 1):
        cl = get_column_letter(ci)
        term = q_residual_term if ci >= Q_START else residual_term
        model_sum = '+'.join([f'{cl}{L[ln]["rev_r"]}' for ln in LN]) + term
        CF(ws, R, ci,
           f'=IFERROR(({model_sum}-{cl}{trev})/ABS({cl}{trev}),"")', fmt=PCT)
    # Revenue YoY
    C(ws, R, DS, '', fmt=PCT)
    cl_e = get_column_letter(DS + 1); cl_d = get_column_letter(DS)
    C(ws, R, DS + 1, f'=IFERROR({cl_e}{trev}/{cl_d}{trev}-1,"")', fmt=PCT)
    for ci in range(FY0, ALL_END + 1):
        cl = get_column_letter(ci)
        if has_q and ci >= Q_START:
            # Q columns: YoY (4Q back) if year-ago exists, else blank
            if ci - 4 >= Q_START:
                pl = get_column_letter(ci - 4)
            else:
                continue
        else:
            # Annual columns: YoY = current/prior year - 1
            pl = get_column_letter(ci - 1)
        C(ws, R, ci, f'=IFERROR({cl}{trev}/{pl}{trev}-1,"")', fmt=PCT)
    C(ws, R, 3, 'Rev YoY', font=nf)
    R += 1
    # QoQ (collapsed, Q columns only)
    if has_q:
        for ci in range(DS, LC_ANNUAL + 1):
            C(ws, R, ci, '', fmt=PCT)
        for qi in range(Q_START, Q_END + 1):
            cl = get_column_letter(qi); pl = get_column_letter(qi - 1)
            if qi == Q_START: C(ws, R, qi, '', fmt=PCT)
            else: C(ws, R, qi, f'=IFERROR({cl}{trev}/{pl}{trev}-1,"")', fmt=PCT)
        C(ws, R, 3, '  Rev QoQ', font=itf)
        R += 1

    # ═══ P&L: Cost → GM → GP (GP row pre-computed for forward ref) ═══
    # GP will be at R+2 (after Cost, GM)
    _gp_future = R + 2

    # Cost = Rev - GP
    for ci in range(DS, ALL_END + 1):
        cl = get_column_letter(ci)
        C(ws, R, ci, f'=IFERROR({cl}{trev}-{cl}{_gp_future},"")', fmt=NUM)
    C(ws, R, 3, 'Cost', font=bf)
    R += 1

    # GM = GP / Rev
    for ci in range(DS, ALL_END + 1):
        cl = get_column_letter(ci)
        C(ws, R, ci, f'=IFERROR({cl}{_gp_future}/{cl}{trev},"")', fmt=PCT)
    C(ws, R, 3, 'GM', font=nf)
    R += 1

    # GP (all depths, all years formula)
    for ci in range(DS, LC_ANNUAL + 1):
        cl = get_column_letter(ci)
        if is_ebitda_depth:
            line_sum = '+'.join([f'{cl}{L[ln]["gp_r"]}' for ln in LN]) + gp_residual_term
            C(ws, R, ci, f'={line_sum}-{cl}{trev}*{gap_gp_ref}', fmt=NUM)
        else:
            C(ws, R, ci, '=' + '+'.join([f'{cl}{L[ln]["gp_r"]}' for ln in LN]) + gp_residual_term, fmt=NUM)
    if has_q:
        for qi in range(Q_START, Q_END + 1):
            cl = get_column_letter(qi)
            if is_ebitda_depth:
                line_sum = '+'.join([f'{cl}{L[ln]["gp_r"]}' for ln in LN]) + q_gp_residual_term
                C(ws, R, qi, f'={line_sum}-{cl}{trev}*{gap_gp_ref}', fmt=NUM)
            else:
                C(ws, R, qi, '=' + '+'.join([f'{cl}{L[ln]["gp_r"]}' for ln in LN]) + q_gp_residual_term, fmt=NUM)
    C(ws, R, 3, 'GP', font=bf)
    tgp = R; gp_row = R if is_ebitda_depth else 0
    R += 1

    # GP YoY
    _gp_yoy_ref = gp_row if is_ebitda_depth else tgp
    C(ws, R, DS, '', fmt=PCT)
    _gpe = get_column_letter(DS + 1); _gpd = get_column_letter(DS)
    C(ws, R, DS + 1, f'=IFERROR({_gpe}{_gp_yoy_ref}/{_gpd}{_gp_yoy_ref}-1,"")', fmt=PCT)
    _gf0 = get_column_letter(FY0); _gf1 = get_column_letter(FY0 - 1)
    C(ws, R, FY0, f'=IFERROR({_gf0}{_gp_yoy_ref}/{_gf1}{_gp_yoy_ref}-1,"")', fmt=PCT)
    for ci in range(FY0 + 1, LC_ANNUAL + 1):
        cl = get_column_letter(ci); pl = get_column_letter(ci - 1)
        C(ws, R, ci, f'=IFERROR({cl}{_gp_yoy_ref}/{pl}{_gp_yoy_ref}-1,"")', fmt=PCT)
    C(ws, R, 3, 'GP YoY', font=nf)
    R += 1
    # GP QoQ
    if has_q:
        for ci in range(DS, LC_ANNUAL + 1):
            C(ws, R, ci, '', fmt=PCT)
        for qi in range(Q_START, Q_END + 1):
            cl = get_column_letter(qi); pl = get_column_letter(qi - 1)
            if qi == Q_START: C(ws, R, qi, '', fmt=PCT)
            else: C(ws, R, qi, f'=IFERROR({cl}{_gp_yoy_ref}/{pl}{_gp_yoy_ref}-1,"")', fmt=PCT)
        C(ws, R, 3, '  GP QoQ', font=itf)
        R += 1

    _gp_ref = tgp  # GP row for Cost/GM/Opex formulas
    nci_rate = meta.get('nci_rate', 0)
    net_debt = meta.get('net_debt', 0)

    # Opex = GP - OI (OI = R + 1)
    for ci in range(DS, ALL_END + 1):
        cl = get_column_letter(ci)
        C(ws, R, ci, f'=IFERROR({cl}{_gp_ref}-{cl}{R + 1},"")', fmt=NUM)
    C(ws, R, 3, 'Opex', font=bf)
    R += 1

    # OI (depth-aware: GP=A/F, OP=F/F, EBITDA=F/F)
    if is_gp_depth:
        # GP depth: OI not from lines — keep actuals for historical, global OPM for projected
        A(ws, R, DS, sc(_gaap('oi', FY2_KEY)), fmt=NUM)
        A(ws, R, DS + 1, sc(_gaap('oi', FY1_KEY)), fmt=NUM)
        A(ws, R, FY0, sc(_gaap('oi', FY0_KEY)), fmt=NUM)
        for ci in range(FY0 + 1, LC_ANNUAL + 1):
            cl = get_column_letter(ci)
            C(ws, R, ci, f'={cl}{tgp}-{cl}{trev}*{cl}{opex_r}', fmt=NUM)
        if has_q:
            for qi in range(q_actual_n):
                qk = f'q{qi+1}'
                qv = sum(seg.get('quarters', {}).get(qk, {}).get('op', 0) for seg in segments)
                if not qv: qv = cfg.get('quarters', {}).get(qk, {}).get('op', 0)
                if qv: A(ws, R, Q_START + qi, sc(qv), fmt=NUM)
            for qi in range(Q_START + q_actual_n, Q_END + 1):
                cl = get_column_letter(qi)
                C(ws, R, qi, f'={cl}{tgp}-{cl}{trev}*{cl}{opex_r}', fmt=NUM)
    elif is_op_depth:
        # OP depth: OI = Σ line OI (all years formula, =ΣQ for complete 4Q years)
        for ci in range(DS, LC_ANNUAL + 1):
            cl = get_column_letter(ci)
            C(ws, R, ci, '=' + '+'.join([f'{cl}{L[ln]["op_r"]}' for ln in LN if L[ln].get('op_r')]), fmt=NUM)
        if has_q:
            for qi in range(Q_START, Q_END + 1):
                cl = get_column_letter(qi)
                C(ws, R, qi, '=' + '+'.join([f'{cl}{L[ln]["op_r"]}' for ln in LN if L[ln].get('op_r')]), fmt=NUM)
    else:
        # EBITDA depth: OI = Σ line EBITDA − Rev × gap_oi (=ΣQ for complete 4Q years)
        for ci in range(DS, LC_ANNUAL + 1):
            cl = get_column_letter(ci)
            line_sum = '+'.join([f'{cl}{L[ln]["gp_r"]}' for ln in LN]) + gp_residual_term
            C(ws, R, ci, f'={line_sum}-{cl}{trev}*{gap_oi_ref}', fmt=NUM)
        if has_q:
            for qi in range(Q_START, Q_END + 1):
                cl = get_column_letter(qi)
                line_sum = '+'.join([f'{cl}{L[ln]["gp_r"]}' for ln in LN]) + q_gp_residual_term
                C(ws, R, qi, f'={line_sum}-{cl}{trev}*{gap_oi_ref}', fmt=NUM)
    C(ws, R, 3, 'OI', font=bf)
    op = R; R += 1

    # OI YoY
    C(ws, R, DS, '', fmt=PCT)
    _cl_e = get_column_letter(DS + 1); _cl_d = get_column_letter(DS)
    C(ws, R, DS + 1, f'=IFERROR({_cl_e}{op}/{_cl_d}{op}-1,"")', fmt=PCT)
    _f0 = get_column_letter(FY0); _f_1 = get_column_letter(FY0 - 1)
    C(ws, R, FY0, f'=IFERROR({_f0}{op}/{_f_1}{op}-1,"")', fmt=PCT)
    for ci in range(FY0 + 1, LC_ANNUAL + 1):
        cl = get_column_letter(ci); pl = get_column_letter(ci - 1)
        C(ws, R, ci, f'=IFERROR({cl}{op}/{pl}{op}-1,"")', fmt=PCT)
    C(ws, R, 3, 'OI YoY', font=nf)
    R += 1
    # OI QoQ
    if has_q:
        for ci in range(DS, LC_ANNUAL + 1):
            C(ws, R, ci, '', fmt=PCT)
        for qi in range(Q_START, Q_END + 1):
            cl = get_column_letter(qi); pl = get_column_letter(qi - 1)
            if qi == Q_START: C(ws, R, qi, '', fmt=PCT)
            else: C(ws, R, qi, f'=IFERROR({cl}{op}/{pl}{op}-1,"")', fmt=PCT)
        C(ws, R, 3, '  OI QoQ', font=itf)
        R += 1

    for ci in range(DS, ALL_END + 1):
        cl = get_column_letter(ci)
        C(ws, R, ci, f'=IFERROR({cl}{op}/{cl}{trev},"")', fmt=PCT)
    C(ws, R, 3, 'OPM', font=nf)
    _opex_end = R; R += 1

    # D&A (EBITDA depth: F/F = Rev×gap_oi; GP/OP depth: A/F = OI-ratio)
    _ebitda_start = R
    if is_ebitda_depth:
        for ci in range(DS, LC_ANNUAL + 1):
            cl = get_column_letter(ci)
            C(ws, R, ci, f'={cl}{trev}*{gap_oi_ref}', fmt=NUM)
        if has_q:
            for qi in range(Q_START, Q_END + 1):
                cl = get_column_letter(qi)
                C(ws, R, qi, f'={cl}{trev}*{gap_oi_ref}', fmt=NUM)
        da_actuals_r = 0
    else:
        da_fy2 = _gaap('da', FY2_KEY); da_fy1 = _gaap('da', FY1_KEY)
        da_fy0 = _gaap('da', FY0_KEY)
        A(ws, R, DS, sc(da_fy2), fmt=NUM)
        A(ws, R, DS + 1, sc(da_fy1), fmt=NUM)
        A(ws, R, FY0, sc(da_fy0), fmt=NUM)
        da_actuals_r = R
        for ci in range(FY0 + 1, LC_ANNUAL + 1):
            cl = get_column_letter(ci)
            C(ws, R, ci, f'={cl}{op}*{get_column_letter(FY0)}{da_actuals_r}/{get_column_letter(FY0)}{trev}', fmt=NUM)
        if has_q:
            da_fy0_col = get_column_letter(FY0)
            for qi in range(q_actual_n):
                qk = f'q{qi+1}'
                qv = sum(seg.get('quarters', {}).get(qk, {}).get('da', 0) for seg in segments)
                if not qv: qv = cfg.get('quarters', {}).get(qk, {}).get('da', 0)
                cl = get_column_letter(Q_START + qi)
                if qv: A(ws, R, Q_START + qi, sc(qv), fmt=NUM)
                else: C(ws, R, Q_START + qi, f'=({da_fy0_col}{op}*{da_fy0_col}{da_actuals_r}/{da_fy0_col}{trev})/4', fmt=NUM)
            for qi in range(Q_START + q_actual_n, Q_END + 1):
                cl = get_column_letter(qi)
                C(ws, R, qi, f'=({da_fy0_col}{op}*{da_fy0_col}{da_actuals_r}/{da_fy0_col}{trev})/4', fmt=NUM)
    C(ws, R, 3, 'D&A', font=bf)
    da_r = R; R += 1

    ebitda_r = R
    if is_ebitda_depth:
        # EBITDA depth: = Σ line EBITDA (all years formula, =ΣQ for complete 4Q)
        for ci in range(DS, LC_ANNUAL + 1):
            cl = get_column_letter(ci)
            line_sum = '+'.join([f'{cl}{L[ln]["gp_r"]}' for ln in LN]) + gp_residual_term
            C(ws, R, ci, f'={line_sum}', fmt=NUM)
        if has_q:
            for qi in range(Q_START, Q_END + 1):
                cl = get_column_letter(qi)
                line_sum = '+'.join([f'{cl}{L[ln]["gp_r"]}' for ln in LN]) + q_gp_residual_term
                C(ws, R, qi, f'={line_sum}', fmt=NUM)
    else:
        # GP/OP depth: EBITDA = OI + D&A
        for ci in range(DS, ALL_END + 1):
            cl = get_column_letter(ci)
            C(ws, R, ci, f'={cl}{op}+{cl}{da_r}', fmt=NUM)
    C(ws, R, 3, 'EBITDA', font=bf)
    R += 1
    # EBITDA YoY (all depths)
    C(ws, R, DS, '', fmt=PCT)
    _ebe = get_column_letter(DS + 1); _ebd = get_column_letter(DS)
    C(ws, R, DS + 1, f'=IFERROR({_ebe}{ebitda_r}/{_ebd}{ebitda_r}-1,"")', fmt=PCT)
    _ebf0 = get_column_letter(FY0); _ebf1 = get_column_letter(FY0 - 1)
    C(ws, R, FY0, f'=IFERROR({_ebf0}{ebitda_r}/{_ebf1}{ebitda_r}-1,"")', fmt=PCT)
    for ci in range(FY0 + 1, LC_ANNUAL + 1):
        cl = get_column_letter(ci); pl = get_column_letter(ci - 1)
        C(ws, R, ci, f'=IFERROR({cl}{ebitda_r}/{pl}{ebitda_r}-1,"")', fmt=PCT)
    C(ws, R, 3, 'EBITDA YoY', font=nf)
    R += 1
    # EBITDA QoQ
    if has_q:
        for ci in range(DS, LC_ANNUAL + 1):
            C(ws, R, ci, '', fmt=PCT)
        for qi in range(Q_START, Q_END + 1):
            cl = get_column_letter(qi); pl = get_column_letter(qi - 1)
            if qi == Q_START: C(ws, R, qi, '', fmt=PCT)
            else: C(ws, R, qi, f'=IFERROR({cl}{ebitda_r}/{pl}{ebitda_r}-1,"")', fmt=PCT)
        C(ws, R, 3, '  EBITDA QoQ', font=itf)
        R += 1

    for ci in range(DS, ALL_END + 1):
        cl = get_column_letter(ci)
        C(ws, R, ci, f'=IFERROR({cl}{ebitda_r}/{cl}{trev},"")', fmt=PCT)
    C(ws, R, 3, 'EBITDA margin', font=nf)
    _ebitda_end = R; R += 1

    # Tax + NI (EBITDA depth: F/F gap-derived; GP/OP depth: A/F)
    # gap_val = (OI-NI)/Rev — the gap rate between operating income and net income
    _ni_start = R
    if is_ebitda_depth:
        # EBITDA depth: Tax = OI × tax_rate (all years formula)
        for ci in range(DS, LC_ANNUAL + 1):
            cl = get_column_letter(ci)
            C(ws, R, ci, f'={cl}{op}*{tax_rate_ref}', fmt=NUM)
        if has_q:
            for qi in range(Q_START, Q_END + 1):
                cl = get_column_letter(qi)
                C(ws, R, qi, f'={cl}{op}*{tax_rate_ref}', fmt=NUM)
    else:
        # GP/OP depth: Tax = Rev × (OI-NI)/Rev (formula reference, historical actuals)
        A(ws, R, DS, sc(_gaap('tax', FY2_KEY)), fmt=NUM)
        A(ws, R, DS + 1, sc(_gaap('tax', FY1_KEY)), fmt=NUM)
        A(ws, R, FY0, sc(_gaap('tax', FY0_KEY)), fmt=NUM)
        for ci in range(FY0 + 1, LC_ANNUAL + 1):
            cl = get_column_letter(ci)
            C(ws, R, ci, f'={cl}{trev}*{cl}{tax_r}', fmt=NUM)
        if has_q:
            for qi in range(q_actual_n):
                qk = f'q{qi+1}'
                qv = sum(seg.get('quarters', {}).get(qk, {}).get('tax', 0) for seg in segments)
                if not qv: qv = cfg.get('quarters', {}).get(qk, {}).get('tax', 0)
                cl = get_column_letter(Q_START + qi)
                if qv: A(ws, R, Q_START + qi, sc(qv), fmt=NUM)
                else: C(ws, R, Q_START + qi, f'={cl}{trev}*{cl}{tax_r}', fmt=NUM)
            for qi in range(Q_START + q_actual_n, Q_END + 1):
                cl = get_column_letter(qi)
                C(ws, R, qi, f'={cl}{trev}*{cl}{tax_r}', fmt=NUM)
    C(ws, R, 3, 'Tax', font=bf)
    tv = R; R += 1

    if is_ebitda_depth:
        # EBITDA depth: NI = OI − Rev × gap_ni (all years formula)
        for ci in range(DS, LC_ANNUAL + 1):
            cl = get_column_letter(ci)
            C(ws, R, ci, f'={cl}{op}-{cl}{trev}*{gap_ni_ref}', fmt=NUM)
        if has_q:
            for qi in range(Q_START, Q_END + 1):
                cl = get_column_letter(qi)
                C(ws, R, qi, f'={cl}{op}-{cl}{trev}*{gap_ni_ref}', fmt=NUM)
    else:
        # GP/OP depth: NI = OI − Rev × (OI-NI)/Rev (formula reference, historical actuals)
        A(ws, R, DS, sc(_gaap('ni', FY2_KEY)), fmt=NUM)
        A(ws, R, DS + 1, sc(_gaap('ni', FY1_KEY)), fmt=NUM)
        A(ws, R, FY0, sc(_gaap('ni', FY0_KEY)), fmt=NUM)
        for ci in range(FY0 + 1, LC_ANNUAL + 1):
            cl = get_column_letter(ci)
            C(ws, R, ci, f'={cl}{op}-{cl}{tv}', fmt=NUM)
        if has_q:
            for qi in range(q_actual_n):
                qk = f'q{qi+1}'
                qv = sum(seg.get('quarters', {}).get(qk, {}).get('ni', 0) for seg in segments)
                if not qv: qv = cfg.get('quarters', {}).get(qk, {}).get('ni', 0)
                cl = get_column_letter(Q_START + qi)
                if qv: A(ws, R, Q_START + qi, sc(qv), fmt=NUM)
                else: C(ws, R, Q_START + qi, f'={cl}{op}-{cl}{tv}', fmt=NUM)
            for qi in range(Q_START + q_actual_n, Q_END + 1):
                cl = get_column_letter(qi)
                C(ws, R, qi, f'={cl}{op}-{cl}{tv}', fmt=NUM)
    C(ws, R, 3, 'Net Income', font=bf)
    ni_r = R; R += 1

    if nci_rate > 0:
        for ci in range(DS, ALL_END + 1):
            cl = get_column_letter(ci)
            C(ws, R, ci, f'={cl}{ni_r}*(1-{nci_rate})', fmt=NUM)
        C(ws, R, 3, 'NI attributable')
        ni_r = R; R += 1

    C(ws, R, DS, '', fmt=PCT)
    cl_e = get_column_letter(DS + 1); cl_d = get_column_letter(DS)
    C(ws, R, DS + 1, f'=IFERROR({cl_e}{ni_r}/{cl_d}{ni_r}-1,"")', fmt=PCT)
    for ci in range(FY0, ALL_END + 1):
        cl = get_column_letter(ci)
        if has_q and ci >= Q_START and ci - 4 >= Q_START:
            pl = get_column_letter(ci - 4)
        elif has_q and ci >= Q_START:
            continue
        else:
            pl = get_column_letter(ci - 1)
        C(ws, R, ci, f'=IFERROR({cl}{ni_r}/{pl}{ni_r}-1,"")', fmt=PCT)
    C(ws, R, 3, 'NI YoY', font=nf)
    R += 1
    # QoQ (collapsed, Q columns only)
    if has_q:
        for ci in range(DS, LC_ANNUAL + 1):
            C(ws, R, ci, '', fmt=PCT)
        for qi in range(Q_START, Q_END + 1):
            cl = get_column_letter(qi); pl = get_column_letter(qi - 1)
            if qi == Q_START: C(ws, R, qi, '', fmt=PCT)
            else: C(ws, R, qi, f'=IFERROR({cl}{ni_r}/{pl}{ni_r}-1,"")', fmt=PCT)
        C(ws, R, 3, '  NI QoQ', font=itf)
        R += 1
    # NPM (after NI YoY+QoQ, consistent with OPM placement)
    for ci in range(DS, ALL_END + 1):
        cl = get_column_letter(ci)
        C(ws, R, ci, f'=IFERROR({cl}{ni_r}/{cl}{trev},"")', fmt=PCT)
    C(ws, R, 3, 'NPM', font=nf)
    R += 1
    _ni_end = R; R += 1

    # ═══ P&L Checks (consolidated at bottom, collapsed) ═══
    # Each Check: (P&L formula row - bridge actuals) / ABS(bridge actuals)
    # Historical years have actuals; projected years blank
    def _write_check(ref_r, act_cells, label):
        # FY checks (D/E/F columns)
        for ci in range(DS, ALL_END + 1):
            cl = get_column_letter(ci)
            if ci == DS: cell = act_cells.get('fy-2', '')
            elif ci == DS + 1: cell = act_cells.get('fy-1', '')
            elif ci == FY0: cell = act_cells.get('fy0', '')
            else: cell = ''
            if cell:
                C(ws, R, ci, f'=IFERROR(({cl}{ref_r}-{cell})/ABS({cell}),"")', fmt=PCT)
            else:
                C(ws, R, ci, '', fmt=PCT)
        # Q checks (N/Q_START... columns, when Q actuals exist)
        q_ok = has_q and label in q_act_cells and q_actual_n > 0
        if q_ok:
            for qi in range(min(q_actual_n, 4)):
                qk = f'q{qi+1}'; qc = Q_START + qi; cl = get_column_letter(qc)
                br = q_act_cells.get(label, {}).get(qk, 0)
                if br:
                    C(ws, R, qc, f'=IFERROR(({cl}{ref_r}-{cl}{br})/ABS({cl}{br}),"")', fmt=PCT)
                else:
                    C(ws, R, qc, '', fmt=PCT)
        C(ws, R, 3, f'  Check {label}', font=itf)
        ws.row_dimensions.group(R, R, outline_level=1, hidden=True)
        return R + 1

    R = _write_check(trev, rev_act_cells, 'Revenue')
    R = _write_check(tgp, gp_act_cells, 'GP')
    if not is_gp_depth:
        R = _write_check(op, op_act_cells, 'OI')
    if is_ebitda_depth:
        R = _write_check(ebitda_r, ebitda_act_cells, 'EBITDA')
        R = _write_check(da_r, da_act_cells, 'D&A')
        R = _write_check(tv, tax_act_cells, 'Tax')
        R = _write_check(ni_r, ni_act_cells, 'NI')


    # ── Inline Check columns: per-row Annual−QSum at existing rows (S1-S3) ──
    if has_q:
        cur_yr, cur_q, qi = q_start_yr, q_start_q, 0
        total_q = q_actual_n + q_proj_n
        SKIP_KW = ('YoY', 'QoQ', 'GM', 'OPM', 'NPM', 'NM', 'margin', '/ Rev', '%', 'Check', 'Implied', 'Bull', 'Base', 'Bear', 'ASP', 'Shares', '(OI', 'NI)/', 'gap_', 'tax_rate', 'Line Revenue', 'actuals')
        while qi < total_q:
            rem_in_fy = 4 - cur_q + 1; fyc = min(rem_in_fy, total_q - qi)
            if fyc == 4:
                ann_col = DS + (cur_yr - bfyr + 2)
                chk_col = Q_END + 3  # after Q + 2 gap
                if ann_col <= LC_ANNUAL:
                    ac = get_column_letter(ann_col); cc = get_column_letter(chk_col)
                    q_letters = [get_column_letter(Q_START + qi + j) for j in range(4)]
                    q_sum_f = '+'.join(f'{ql}{{r}}' for ql in q_letters)
                    ws.cell(row=1, column=chk_col).value = f'FY{cur_yr} Check'
                    ws.cell(row=1, column=chk_col).font = bf
                    for row in range(s1_start, _ni_end + 1):
                        cv = str(ws.cell(row=row, column=3).value or '').strip()
                        if not cv or any(kw in cv for kw in SKIP_KW): continue
                        ws.cell(row=row, column=chk_col).value = f'=IFERROR(({ac}{row}-({q_sum_f.format(r=row)}))/ABS({ac}{row}),"")'
                        ws.cell(row=row, column=chk_col).number_format = PCT
                        ws.cell(row=row, column=chk_col).font = nf
                    ws.column_dimensions[cc].width = 12
                    ws.column_dimensions.group(cc, cc, outline_level=1, hidden=True)
            qi += fyc; cur_yr += 1; cur_q = 1

    # ── Rate Bridge: fill Opex/Rev + (OI-NI)/Rev (all depths) ──
    if meta.get('p&l_depth') == 'gp':
        # GP depth: historical = formula, projected = assumption
        for ci in [DS, DS + 1, FY0]:
            cl = get_column_letter(ci)
            CF(ws, opex_r, ci, f'=IFERROR(({cl}{tgp}-{cl}{op})/{cl}{trev},"")', fmt=PCT)
        for i, proj_fy in enumerate(_PROJ_FYS, FY0 + 1):
            I(ws, opex_r, i, _opm(proj_fy), fmt=PCT)
        if has_q:
            _opms = [_opm(fy) for fy in _ALL_FYS]
            cur_yr, cur_q = q_start_yr, q_start_q
            for qi in range(q_actual_n + q_proj_n):
                _py = cur_yr - bfyr - 1
                if _py < 0: _py = 0
                _idx = min(3 + _py, len(_opms) - 1)
                I(ws, opex_r, Q_START + qi, _opms[_idx], fmt=PCT)
            cur_q += 1
            if cur_q > 4: cur_q = 1; cur_yr += 1
    else:
        # OP/EBITDA depth: Opex/Rev = (GP−OI)/Rev (derived from P&L)
        for ci in range(DS, ALL_END + 1):
            cl = get_column_letter(ci)
            CF(ws, opex_r, ci, f'=IFERROR(({cl}{tgp}-{cl}{op})/{cl}{trev},"")', fmt=PCT)

    # (OI-NI)/Rev: historical = formula, projected depends on depth
    for ci in [DS, DS + 1, FY0]:
        cl = get_column_letter(ci)
        CF(ws, tax_r, ci, f'=IFERROR(({cl}{op}-{cl}{ni_r})/{cl}{trev},"")', fmt=PCT)
    if is_ebitda_depth:
        for ci in range(FY0 + 1, LC_ANNUAL + 1):
            C(ws, tax_r, ci, f'={gap_ni_ref}', fmt=PCT)
        if has_q:
            for qi in range(q_actual_n + q_proj_n):
                C(ws, tax_r, Q_START + qi, f'={gap_ni_ref}', fmt=PCT)
    else:
        _gap_hist = []
        for _fy in [FY2_KEY, FY1_KEY, FY0_KEY]:
            _oi_v = _gaap('oi', _fy); _ni_v = _gaap('ni', _fy); _rev_v = _gaap('rev', _fy)
            if _oi_v and _ni_v and _rev_v and _rev_v > 0:
                _gap_hist.append((_oi_v - _ni_v) / _rev_v)
        _avg_gap = round(sum(_gap_hist) / len(_gap_hist), 4) if _gap_hist else _gl('gap_oi_ni', FY0_KEY)
        for ci in range(FY0 + 1, LC_ANNUAL + 1):
            proj_idx = ci - FY0 - 1
            rate_val = _gl('gap_oi_ni', _PROJ_FYS[proj_idx]) if 0 <= proj_idx < len(_PROJ_FYS) else 0
            I(ws, tax_r, ci, rate_val or _avg_gap, fmt=PCT)
        if has_q:
            _q_rate = _gl('gap_oi_ni', FY0_KEY) or _avg_gap
            for qi in range(q_actual_n + q_proj_n):
                I(ws, tax_r, Q_START + qi, _q_rate, fmt=PCT)

    # ═══════════════ §4 SOTP - Logic ═══════════════
    R += 1
    C(ws, R, 1, 'SOTP - Logic', font=bf12)
    R += 1
    sc_l = get_column_letter(SC)
    mc_rows = []

    def _sotp_info(ll):
        """Resolve sotp method + multiple, backward-compat with sotp_pe."""
        s = ll.get('sotp', {})
        if not s and 'sotp_pe' in ll:
            s = {'method': 'pe', 'multiple': ll['sotp_pe']}
        default_method = 'ev_ebitda' if is_ebitda_depth else 'pe'
        return s.get('method', default_method), s.get('multiple', 10)

    def _sotp_metric_ref(method):
        """Return (metric_row, metric_label) for a given valuation method.
        All metrics always exist. P&L depth set by segment disclosure (max_seg_depth)."""
        if method == 'pe':
            return ni_r, 'NI'
        if method == 'ev_ebitda':
            return ebitda_r, 'EBITDA'
        if method == 'ev_ebit':
            return op, 'EBIT'
        if method in ('ev_sales', 'ps'):
            return trev, 'Revenue'
        return tgp, 'GP'

    # Detect whether all lines use EV methods (for TOTAL directional logic)
    _all_methods = set()
    for ll in logic_lines:
        _m, _ = _sotp_info(ll)
        _all_methods.add(_m)
    is_ev_method = _all_methods and all(m.startswith('ev_') for m in _all_methods)

    for ll in logic_lines:
        ln = ll['name']
        method, mult = _sotp_info(ll)
        metric_r, metric_label = _sotp_metric_ref(method)
        gc = f'{sc_l}{L[ln]["gp_r"]}'        # line profit (allocation key)
        lr = f'{sc_l}{L[ln]["rev_r"]}'        # line revenue
        mc = f'{sc_l}{metric_r}'              # total metric
        tc_gp = f'{sc_l}{tgp}'                # total profit
        is_ev = method.startswith('ev_')
        is_ps = method in ('ev_sales', 'ps')

        # Revenue (all methods)
        C(ws, R, 2, ln, font=bf)
        C(ws, R, 3, 'Revenue', font=bf)
        C(ws, R, SC, f'={lr}', fmt=NUM)
        R += 1

        # Allocated metric
        if is_ps:
            alloc_formula = f'={lr}'
            alloc_ref = lr
        else:
            alloc_formula = f'=IFERROR({mc}*{gc}/{tc_gp},"")'
            alloc_ref = f'({mc}*{gc}/{tc_gp})'
        C(ws, R, 3, metric_label)
        C(ws, R, SC, alloc_formula, fmt=NUM)
        R += 1

        # Multiple
        if method == 'pe':      label_m = 'PE'
        elif method == 'ps':    label_m = 'P/S'
        else:                   label_m = method.replace('_', '/').upper()
        I(ws, R, SC, mult, fmt='0.0x')
        C(ws, R, 3, label_m, font=nf)
        mult_row = R; R += 1

        # EV or Mkt Cap (no per-line Net Debt)
        value_f = f'=IFERROR({alloc_ref}*{sc_l}{mult_row},"")'
        if is_ev:
            C(ws, R, 3, 'EV', font=bf)
        else:
            C(ws, R, 3, 'Mkt Cap', font=bf)
        C(ws, R, SC, value_f, fmt=DEC)
        mc_rows.append(R); R += 1

    # Blank row before Total section + Enterprise Value → Net Debt → Mkt Cap
    R += 1
    nd_total_r = R + 1  # Net Debt row (forward ref for PE/PS)
    if is_ev_method:
        ev_f = '=' + '+'.join([f'{sc_l}{mr}' for mr in mc_rows])
    else:
        ev_f = '=(' + '+'.join([f'{sc_l}{mr}' for mr in mc_rows]) + f')+{sc_l}{nd_total_r}'
    C(ws, R, 2, 'Total', font=bf)
    C(ws, R, SC, ev_f, fmt=DEC)
    C(ws, R, 3, 'Enterprise Value', font=bf)
    sotp_ev_r = R; R += 1

    # Net Debt
    I(ws, R, SC, net_debt, fmt=DEC)
    C(ws, R, 3, 'Net Debt', font=bf)
    nd_total_r = R; R += 1

    # Mkt Cap (EV method: EV − Net Debt; PE/PS: Σ Mkt Cap)
    if is_ev_method:
        mcap_f = f'=IFERROR({sc_l}{sotp_ev_r}-{sc_l}{nd_total_r},"")'
    else:
        mcap_f = '=' + '+'.join([f'{sc_l}{mr}' for mr in mc_rows])
    C(ws, R, SC, mcap_f, fmt=DEC)
    C(ws, R, 3, 'Mkt Cap', font=bf)
    sotp_mcap_r = R; R += 1

    # ═══════════════ §5 SOTP - Segments ═══════════════
    R += 1
    C(ws, R, 1, 'SOTP - Segments', font=bf12)
    sotp_seg_start = R
    R += 1
    smc_rows = []
    LL_SOTP = {}
    for ll in logic_lines:
        method, mult = _sotp_info(ll)
        LL_SOTP[ll['name']] = (method, mult)

    for seg in segments:
        sn = seg['name']; lls = seg['logic_lines']
        lmethods = [LL_SOTP[l['name']][1] for l in lls]
        w_mult = sum(lmethods[i] * _gaap_seg(seg['name'], 'rev', FY0_KEY) * lls[i]['split']
                     for i in range(len(lls))) / _gaap_seg(seg['name'], 'rev', FY0_KEY) if _gaap_seg(seg['name'], 'rev', FY0_KEY) > 0 else 10
        mult_s = round(w_mult)
        method_s = LL_SOTP[lls[0]['name']][0]
        metric_r_s, metric_label_s = _sotp_metric_ref(method_s)
        gc = f'{sc_l}{seg_info[sn]["gp"]}'
        sr = f'{sc_l}{seg_info[sn]["rev"]}'
        mc = f'{sc_l}{metric_r_s}'
        is_ev_s = method_s.startswith('ev_')
        is_ps_s = method_s in ('ev_sales', 'ps')

        # Revenue
        C(ws, R, 2, sn, font=bf)
        C(ws, R, 3, 'Revenue', font=bf)
        C(ws, R, SC, f'={sr}', fmt=NUM)
        R += 1

        # Allocated metric
        if is_ps_s:
            alloc_f_s = f'={sr}'
            alloc_ref_s = sr
        else:
            alloc_f_s = f'=IFERROR({mc}*{gc}/{sc_l}{tgp},"")'
            alloc_ref_s = f'({mc}*{gc}/{sc_l}{tgp})'
        C(ws, R, 3, metric_label_s)
        C(ws, R, SC, alloc_f_s, fmt=NUM)
        R += 1

        # Multiple
        if method_s == 'pe':      label_ms = 'PE'
        elif method_s == 'ps':    label_ms = 'P/S'
        else:                     label_ms = method_s.replace('_', '/').upper()
        I(ws, R, SC, mult_s, fmt='0.0x')
        C(ws, R, 3, label_ms, font=nf)
        pe_row = R; R += 1

        # EV or Mkt Cap
        value_f_s = f'=IFERROR({alloc_ref_s}*{sc_l}{pe_row},"")'
        if is_ev_s:
            C(ws, R, 3, 'EV', font=bf)
        else:
            C(ws, R, 3, 'Mkt Cap', font=bf)
        C(ws, R, SC, value_f_s, fmt=DEC)
        smc_rows.append(R); R += 1

    # Blank row before Total + Enterprise Value → Net Debt → Mkt Cap
    R += 1
    nd_seg_r = R + 1
    if is_ev_method:
        ev_f_s = '=' + '+'.join([f'{sc_l}{mr}' for mr in smc_rows])
    else:
        ev_f_s = '=(' + '+'.join([f'{sc_l}{mr}' for mr in smc_rows]) + f')+{sc_l}{nd_seg_r}'
    C(ws, R, 2, 'Total', font=bf)
    C(ws, R, SC, ev_f_s, fmt=DEC)
    C(ws, R, 3, 'Enterprise Value', font=bf)
    sotp_seg_ev_r = R; R += 1

    I(ws, R, SC, net_debt, fmt=DEC)
    C(ws, R, 3, 'Net Debt', font=bf)
    nd_seg_r = R; R += 1

    if is_ev_method:
        mcap_f_s = f'=IFERROR({sc_l}{sotp_seg_ev_r}-{sc_l}{nd_seg_r},"")'
    else:
        mcap_f_s = '=' + '+'.join([f'{sc_l}{mr}' for mr in smc_rows])
    C(ws, R, SC, mcap_f_s, fmt=DEC)
    C(ws, R, 3, 'Mkt Cap', font=bf)
    sotp_seg_mcap_r = R; R += 1
    ws.row_dimensions.group(sotp_seg_start, R - 1, outline_level=1, hidden=True)

    # ═══════════════ §6 Market Data ═══════════════
    R += 1
    C(ws, R, 1, 'Market Cap', font=bf12)
    R += 1
    # Key metrics — highlighted (actual market data)
    C(ws, R, 3, 'MCap (Actual)', font=hlfont, fill=hlfill)
    C(ws, R, SC, mcap_d, fmt=NUM)
    mcap_data_r = R; R += 1
    if shares:
        C(ws, R, 3, 'Shares (M)', font=hlfont, fill=hlfill)
        C(ws, R, SC, shares, fmt='#,##0.0')
        shares_data_r = R; R += 1
    if price:
        C(ws, R, 3, 'Price', font=hlfont, fill=hlfill)
        C(ws, R, SC, price, fmt=price_fmt)
        price_data_r = R; R += 1
    R += 1
    mref = f'{sc_l}{mcap_data_r}'

    # Implied valuation (references SOTP-derived Mkt Cap, not EV)
    if shares:
        implied_logic = f'=IFERROR({sc_l}{sotp_mcap_r}*{div}/{sc_l}{shares_data_r},"")'
        implied_seg = f'=IFERROR({sc_l}{sotp_seg_mcap_r}*{div}/{sc_l}{shares_data_r},"")'
        C(ws, R, 3, 'Implied Price')
        C(ws, R, SC, implied_logic, fmt=price_fmt)
        imp_row = R; R += 1
        C(ws, R, 3, 'Implied Price (Seg)')
        C(ws, R, SC, implied_seg, fmt=price_fmt)
        R += 1
        C(ws, R, 3, 'Upside / Downside')
        C(ws, R, SC, f'=IFERROR({sc_l}{imp_row}/{sc_l}{price_data_r}-1,"")', fmt='0.0%')
        R += 1
    if ttm_pe:
        C(ws, R, 3, 'TTM PE')
        C(ws, R, SC, round(ttm_pe, 1), fmt='0.0x')
        R += 1
    if fwd_pe:
        C(ws, R, 3, 'Fwd PE')
        C(ws, R, SC, round(fwd_pe, 1), fmt='0.0x')
        R += 1
    if hi52:
        currency_prefix = PRICE_FMT.get(meta.get('currency', 'USD'), '$#,##0.00').replace('#,##0.00','').replace('#,##0','').rstrip()
        C(ws, R, 3, '52W Range')
        C(ws, R, SC, f'{currency_prefix}{lo52:.0f} - {currency_prefix}{hi52:.0f}')
        R += 1

    # ═══════════════ §7 Scenario Summary ═══════════════
    R += 2
    C(ws, R, 1, 'Scenario Summary', font=bf12)
    R += 1
    syl = YR[2 + s_off].replace('A', 'E')
    _gp_label = 'EBITDA' if is_ebitda_depth else 'GP'
    C(ws, R, DS, f'{syl} Rev', font=bf)
    C(ws, R, DS + 1, f'{syl} {_gp_label}', font=bf)
    # Determine dominant metric/multiple labels from SOTP methods in use
    methods = set()
    for ll in logic_lines:
        if ll.get('sotp'):
            methods.add(ll['sotp'].get('method', 'ev_ebitda' if is_ebitda_depth else 'pe'))
        elif 'sotp_pe' in ll:
            methods.add('pe')
    if not methods: methods = {'pe'}
    ev_only = all(m.startswith('ev_') for m in methods)
    metric_label = 'EBITDA' if ev_only else ('Revenue' if methods == {'ps'} else 'NI')
    mult_label = 'Implied EV/EBITDA' if ev_only else ('Implied P/S' if methods == {'ps'} else 'Implied PE')
    C(ws, R, DS + 2, f'{syl} {metric_label}', font=bf)
    C(ws, R, DS + 3, mult_label, font=bf)
    R += 1
    syc = get_column_letter(FY0 + s_off)

    for label, yk in [('Bull', 'yb'), ('Base', 'ybs'), ('Bear', 'ybe')]:
        rp = []; gp = []
        for ll in logic_lines:
            ln = ll['name']
            rows = L.get(ln)
            if not rows:
                continue
            is_yoy = rows.get('module') == 'yoy'
            if is_yoy and yk in rows:
                # yoy: chain projection FY0 → FY0+s_off
                yr = rows[yk]
                comp = f'{get_column_letter(FY0)}{rows["rev_r"]}'
                for off in range(1, s_off + 1):
                    comp = f'{comp}*(1+{get_column_letter(FY0 + off)}{yr})'
            elif yk in rows:
                # non-yoy with BBE cache: direct SC read
                comp = f'{sc_l}{rows[yk]}'
            else:
                # non-yoy no BBE: use Revenue SC (scenario-independent)
                comp = f'{sc_l}{rows["rev_r"]}'
            rp.append(comp)
            gp.append(f'({comp}*{sc_l}{rows["gm_r"]})')
        if not rp:
            continue
        rf = '=' + '+'.join(rp)
        gf = '=' + '+'.join(gp)
        rev_total_sc = f'({rf[1:]})' if rf.startswith('=') else rf
        # NI/EBITDA metric — mirrors main P&L logic exactly:
        #   EBITDA depth: metric = Σ(line_rev × EBITDA_margin) = gf (already EBITDA)
        #   GP/OP depth: OI = GP − Rev×Opex/Rev, NI = OI − Rev×(OI-NI)/Rev
        #     → NI = GP − Rev×(Opex/Rev_cell) − Rev×((OI-NI)/Rev_cell)
        if ev_only:
            nf_s = gf  # EBITDA = sum of line EBITDA (gf already computes this)
        else:
            _opex_cell = f'{syc}{opex_r}'
            _gap_cell = f'{syc}{tax_r}'
            nf_s = f'={gf[1:]}-({rev_total_sc}*{_opex_cell})-({rev_total_sc}*{_gap_cell})'
        pf = f'=IFERROR({mref}/({nf_s[1:]}),"")'
        C(ws, R, 2, label, font=bf)
        C(ws, R, DS, rf, fmt=NUM)
        C(ws, R, DS + 1, gf, fmt=NUM)
        C(ws, R, DS + 2, nf_s, fmt=NUM)
        C(ws, R, DS + 3, pf, font=bf, fmt='0.0x')
        R += 1

    # ═══════════════ Post-format ═══════════════
    for row in range(1, R):
        # Clear placeholder zeros (only cells without explicit format)
        for c in range(DS, ALL_END + 1):
            cl = ws.cell(row=row, column=c)
            if cl.value == 0 and cl.number_format == 'General':
                cl.value = None
        # Clear D/E formula-only cells (Section 2 only, skip protected rows)
        if s1_end < row <= s2_end and row not in protected_rows:
            for c in (DS, DS + 1):
                cl = ws.cell(row=row, column=c)
                if cl.value and isinstance(cl.value, str) and cl.value.startswith('='):
                    cl.value = None
        # Bold key C-column labels
        cv = ws.cell(row=row, column=3).value
        if cv and isinstance(cv, str) and cv.strip() in {
            'Revenue', 'Cost', 'GP', 'GM', 'OP', 'OPM', 'OI YoY',
            'Total Revenue', 'Rev YoY', 'Total GP', 'Blended GM',
            'Opex', 'Operating Profit', 'EBITDA', 'EBIT',
            'Tax', 'Net Income', 'NPM', 'NI YoY',
        }:
            cell = ws.cell(row=row, column=3)
            if not (cell.fill and cell.fill.start_color and cell.fill.start_color.rgb == '00963634'):
                cell.font = bf
    # Clear gap columns between annual and Q (value + fill)
    if has_q:
        for row in range(1, R):
            for gc in (LC_ANNUAL + 1, LC_ANNUAL + 2):
                c = ws.cell(row=row, column=gc)
                c.value = None; c.fill = PatternFill()
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 36
    for ci in range(DS, ALL_END + 1):
        cl = get_column_letter(ci)
        ws.column_dimensions[cl].width = 5 if has_q and ci in (LC_ANNUAL + 1, LC_ANNUAL + 2) else 13

    ws.freeze_panes = 'D2'

    # Output: auto-derive to company root as YYYY-MM-DD-driver-model.xlsx
    if output_path:
        out_path = output_path
    else:
        # json_path: industry/<ind>/companies/<ticker>/.cache/scripts/research-model.json
        # → derive company dir
        json_dir = os.path.dirname(os.path.abspath(json_path))
        parts = json_dir.replace('\\', '/').split('/')
        try:
            ci = parts.index('companies')
            ticker_dir = parts[ci + 1]  # e.g. 'hwm', 'dpc'
            company_root = '/'.join(parts[:ci + 2])  # up to companies/<ticker>
            today = datetime.date.today().isoformat()
            out_path = os.path.join(company_root, f'{today}-{ticker_dir}-driver-model.xlsx')
        except (ValueError, IndexError):
            out_path = json_path.replace('.json', '.xlsx')
    wb.save(out_path)
    print(f'OK: {out_path}')

    # ── COM: compute + read Checks, write checks.json with flags ──
    THRESHOLDS = {
        'Check Rev': 0.02, 'Check GP': 0.05, 'Check OI': 0.05,
        'Check EBITDA': 0.05, 'Check D&A': 0.20, 'Check Tax': 0.10, 'Check NI': 0.15,
    }
    if is_ebitda_depth:
        THRESHOLDS['Check GP'] = 0.15; THRESHOLDS['Check OI'] = 0.15
    try:
        xl = win32com.client.Dispatch('Excel.Application')
        time.sleep(0.5)
        wb2 = xl.Workbooks.Open(os.path.abspath(out_path))
        ws2 = wb2.Sheets(1)
        check_data = {}
        q_check_data = {}
        flags_pnl = []; flags_q = []
        year_labels = [cfg['meta'].get('base_fy', 2025) - i for i in [2, 1, 0]]

        # Scan P&L checks (C column labels starting with '  Check ')
        for r in range(1, 3000):
            lab = str(ws2.Cells(r, 3).Value or '')
            if lab.startswith('  Check '):
                clean = lab.strip()
                vals = {}
                for ci, yr in zip([4, 5, 6], year_labels):
                    v = ws2.Cells(r, ci).Value
                    if isinstance(v, float):
                        pct = abs(v)
                        vals[str(yr)] = f'{v:.2%}' if pct < 100 else str(round(v, 1))
                        th = THRESHOLDS.get(clean, 0.10)
                        if pct > th:
                            flags_pnl.append(f'{clean} {yr}={v:.2%} (> {th:.0%})')
                    elif v is None:
                        vals[str(yr)] = '—'
                    else:
                        vals[str(yr)] = str(v)[:20]
                check_data[clean] = vals

        # Scan Q checks (U columns — FY{yr} Check headers, P&L F/F rows only)
        if has_q:
            if is_ebitda_depth:
                PNL_LABELS = {'Revenue', 'Cost', 'GP', 'Opex', 'OI', 'OPM', 'D&A', 'EBITDA', 'Tax', 'Net Income', 'NPM'}
            elif is_op_depth:
                PNL_LABELS = {'Revenue', 'Cost', 'GP', 'Opex', 'OI', 'OPM'}
            else:
                PNL_LABELS = {'Revenue', 'Cost', 'GP'}
            for chk_col in range(Q_END + 3, Q_END + 10):
                hdr = str(ws2.Cells(1, chk_col).Value or '')
                if not hdr.startswith('FY'): break
                fy_label = hdr.replace(' Check', '')
                q_check_data[fy_label] = {}
                for row in range(trev, _ni_end + 1):
                    cv = str(ws2.Cells(row, 3).Value or '').strip()
                    if cv not in PNL_LABELS: continue
                    v = ws2.Cells(row, chk_col).Value
                    if isinstance(v, float):
                        q_check_data[fy_label][cv] = f'{v:.2%}'
                        if abs(v) > 0.0001:
                            flags_q.append(f'{fy_label} {cv}={v:.2%} (ΣQ≠Annual)')
                    elif v is not None:
                        q_check_data[fy_label][cv] = str(v)[:20]

        # qq_checks: COM read P&L Q cells vs bridge Q actuals
        qq_check_data = {}
        if q_act_cells and q_actual_n > 0:
            xl.Calculation = -4105
            xl.CalculateFull()
            time.sleep(2)
            MAP = {'Revenue': trev, 'GP': tgp, 'OI': op, 'EBITDA': ebitda_r, 'NI': ni_r, 'Tax': tv}
            for qi in range(min(q_actual_n, 4)):
                qk = f'q{qi+1}'; qc = Q_START + qi
                qq_check_data[qk] = {}
                for lab, pnl_r in MAP.items():
                    if not pnl_r: continue
                    try:
                        cl = get_column_letter(qc)
                        v = ws2.Range(f'{cl}{pnl_r}').Value
                    except:
                        v = ws2.Cells(pnl_r, qc).Value
                    ar = q_act_cells.get(lab, {}).get(qk, 0)
                    try:
                        av = ws2.Range(f'{get_column_letter(DS)}{ar}').Value if ar else None
                    except:
                        av = ws2.Cells(ar, DS).Value if ar else None
                    if isinstance(v, float) and isinstance(av, float) and abs(av) > 0.01:
                        qq_check_data[qk][lab] = f'{(v - av) / abs(av):.2%}'

        # qq_checks: COM read P&L Q vs bridge actuals
        qq_check_data = {}
        if q_act_cells and q_actual_n > 0:
            xl.CalculateFull()
            time.sleep(2)
            MAP = {'Revenue':trev, 'GP':tgp, 'OI':op, 'EBITDA':ebitda_r, 'NI':ni_r, 'Tax':tv}
            for qi in range(min(q_actual_n, 4)):
                qk = f'q{qi+1}'; qc = Q_START + qi
                qq_check_data[qk] = {}
                for lab, pnl_r in MAP.items():
                    if not pnl_r: continue
                    v = ws2.Cells(pnl_r, qc).Value
                    ar = q_act_cells.get(lab, {}).get(qk, 0)
                    av = ws2.Cells(ar, DS).Value if ar else None
                    if isinstance(v, float) and isinstance(av, float) and abs(av) > 0.01:
                        qq_check_data[qk][lab] = f'{(v - av) / abs(av):.2%}'

        wb2.Close(False)
        xl.Quit()

        cj_path = json_path.replace('.json', '_checks.json')
        result = {
            'generated_at': datetime.datetime.now().isoformat(),
            'model': os.path.basename(out_path),
            'depth': cfg['meta'].get('p&l_depth', '?'),
            'checks': check_data,
            'flags': {'P&L': flags_pnl, 'Q': flags_q},
        }
        if has_q:
            result['q_checks'] = q_check_data
        if qq_check_data:
            result['qq_checks'] = qq_check_data
        with open(cj_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, indent=2, ensure_ascii=False)
        print(f'  Checks: {cj_path}')
        pl_checks = {k: v for k, v in check_data.items() if '%' not in k}
        for lab, vals in pl_checks.items():
            items = list(vals.items())[:3]
            print(f'    {lab:25s} {" | ".join(f"{k}={v}" for k,v in items)}')
        if flags_pnl:
            print(f'  [!] P&L flags ({len(flags_pnl)}):')
            for f in flags_pnl[:5]: print(f'    {f}')
        if flags_q:
            print(f'  [!] Q flags ({len(flags_q)}):')
            for f in flags_q[:3]: print(f'    {f}')
        if not flags_pnl and not flags_q:
            print(f'  [OK] All checks within threshold')
    except Exception as e:
        print(f'  [COM skip] {e}')


# ═══════════════════════════════════════════════════════════════
# CLI entry
# ═══════════════════════════════════════════════════════════════

if __name__ == '__main__':
    p = argparse.ArgumentParser()
    p.add_argument('json_path')
    p.add_argument('-o', '--output')
    a = p.parse_args()
    build(a.json_path, a.output)
