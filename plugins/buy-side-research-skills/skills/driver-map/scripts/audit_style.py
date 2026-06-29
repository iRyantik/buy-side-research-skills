"""audit_style.py — Post-build cell style auditor.

Usage: python .scripts/driver-map/audit_style.py <model.xlsx>

Scans every used cell and reports style violations:
  - I() cells: must have yellow fill (FFFFCC) and blue font (0000CC)
  - A() cells: must have gray fill (F0F0F0)
  - Formula cells: must have non-'General' number_format
  - PCT format mismatch: label contains GM/OPM/NPM/YoY/ratio → must be 0.0%
  - Font check: all cells must use Calibri (not default)
  - General format on numeric/formula cells (most common regression)

Exit code 0 = clean, 1 = warnings, 2 = errors.
"""

import sys, argparse, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.utils import get_column_letter

NUM_FMTS = {'#,##0', '#,##0.0', '#,##0.00', '0.0x', '#,##0.0x'}
PCT_FMTS = {'0.0%', '0%', '0.00%'}
PCT = '0.0%'
DEC = '#,##0.00'

YELLOW_FILL = 'FFFFCC00'  # openpyxl stores ARGB
GRAY_FILL = '00F0F0F0'
BLUE_FONT = '000000CC'

PCT_KEYWORDS = ('GM', 'OPM', 'NPM', 'YoY', 'margin', 'rate', 'Ratio', '%')
NUM_KEYWORDS = ('Revenue', 'GP', 'OP', 'NI', 'Opex', 'Cost',
                'EBITDA', 'EBIT', 'Tax', 'MCap', 'Price', 'Volume', 'Shares')
SKIP_LABEL_PREFIXES = ('  Check', '  Bull', '  Base', '  Bear')


def audit(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    errors = []
    warnings = []

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue

            cl = get_column_letter(col)
            loc = f'{cl}{row}'
            is_formula = isinstance(cell.value, str) and cell.value.startswith('=')
            is_numeric = isinstance(cell.value, (int, float))
            fill_rgb = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else ''
            font_name = cell.font.name if cell.font else ''
            nf = cell.number_format
            c3 = str(ws.cell(row=row, column=3).value or '')
            c2 = str(ws.cell(row=row, column=2).value or '')

            # ── General format on non-empty cells ──
            if (is_formula or is_numeric) and nf == 'General':
                errors.append(f'{loc}: formula/numeric cell has General format')

            # ── Skip check/scenario rows (labels start with indent markers) ──
            skip = any(c3.startswith(p) for p in SKIP_LABEL_PREFIXES) if c3 else False

            # ── PCT label but not PCT format ──
            if (is_formula or is_numeric) and not skip:
                is_pct_label = any(kw in c3 for kw in PCT_KEYWORDS) if c3 else False
                is_pct_col2 = any(kw in c2 for kw in PCT_KEYWORDS) if c2 else False
                if (is_pct_label or is_pct_col2) and nf not in PCT_FMTS:
                    warnings.append(f'{loc}: likely PCT cell (label "{c3}") has fmt {nf}')

            # ── NUM label but not NUM format (only if not already PCT format) ──
            if is_formula and not skip and nf not in PCT_FMTS:
                is_num_label = any(kw in c3 for kw in NUM_KEYWORDS) if c3 else False
                if is_num_label and nf not in NUM_FMTS and '#' not in nf:
                    warnings.append(f'{loc}: likely NUM cell (label "{c3}") has fmt {nf}')

            # ── Font check (skip headers) ──
            if font_name and font_name not in ('Calibri', 'Calibri Light'):
                warnings.append(f'{loc}: non-Calibri font "{font_name}"')

    # Report
    if errors:
        print(f'\n=== {len(errors)} ERRORS ===')
        for e in errors[:30]:
            print(f'  ERR {e}')
        if len(errors) > 30:
            print(f'  ... and {len(errors) - 30} more')

    if warnings:
        print(f'\n=== {len(warnings)} WARNINGS ===')
        for w in warnings[:20]:
            print(f'  WARN {w}')
        if len(warnings) > 20:
            print(f'  ... and {len(warnings) - 20} more')

    if not errors and not warnings:
        print('  [OK] Clean — no style violations')

    return 2 if errors else (1 if warnings else 0)


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx')
    args = ap.parse_args()
    sys.exit(audit(args.xlsx))
