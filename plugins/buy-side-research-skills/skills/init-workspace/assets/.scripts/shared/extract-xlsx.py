#!/usr/bin/env python3
"""extract-xlsx — Excel to structured JSON. Lightweight, no pandas.

Usage:
  python extract-xlsx.py <file>              # all sheets, JSON
  python extract-xlsx.py <file> --sheet 0     # single sheet by index
  python extract-xlsx.py <file> --markdown     # markdown table output
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path


def extract(filepath: str, sheet: int | str | None = None) -> list[dict]:
    """Extract all sheets. Returns [{name, headers, rows}]."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install --user openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    sheets = []

    names = [sheet] if isinstance(sheet, str) else wb.sheetnames
    if isinstance(sheet, int):
        names = [wb.sheetnames[sheet]] if sheet < len(wb.sheetnames) else []

    for sname in names:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(c) if c is not None else "" for c in rows[0]]
        data = []
        for row in rows[1:]:
            data.append([str(c) if c is not None else "" for c in row])
        sheets.append({"name": sname, "headers": headers, "rows": data})
    wb.close()
    return sheets


def to_markdown(sheets: list[dict]) -> str:
    """Convert extracted sheets to markdown tables."""
    lines = []
    for s in sheets:
        lines.append(f"### {s['name']}")
        lines.append("")
        header = "| " + " | ".join(s["headers"]) + " |"
        sep = "|---" * len(s["headers"]) + "|"
        lines.extend([header, sep])
        for row in s["rows"]:
            lines.append("| " + " | ".join(row) + " |")
        lines.append("")
    return "\n".join(lines)


def main():
    p = argparse.ArgumentParser(description="Excel to JSON/markdown")
    p.add_argument("file", help="Excel file path")
    p.add_argument("--sheet", help="Sheet name or index (default: all)")
    p.add_argument("--markdown", action="store_true", help="Output as markdown tables")
    p.add_argument("--json", action="store_true", help="Output as JSON (default)")
    args = p.parse_args()

    sheet = None
    if args.sheet:
        try:
            sheet = int(args.sheet)
        except ValueError:
            sheet = args.sheet

    sheets = extract(args.file, sheet)

    if args.markdown:
        print(to_markdown(sheets))
    else:
        print(json.dumps(sheets, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
