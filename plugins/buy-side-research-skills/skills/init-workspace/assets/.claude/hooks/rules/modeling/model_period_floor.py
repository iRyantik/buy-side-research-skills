"""Check: model must have at least 2 historical + 3 forecast years (3sm), or 5 forecast years (DCF)."""
import re, sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from common import block, warn
from rules.modeling._common import load_payload, get_xlsx_targets, get_sheet_names_from_payload

payload = load_payload()
for t in get_xlsx_targets(payload):
    import openpyxl
    wb = openpyxl.load_workbook(t["path"], data_only=True)
    if "IS" not in wb.sheetnames:
        wb.close()
        continue

    ws = wb["IS"]
    header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    # Count columns that look like years
    year_cols = 0
    for c in header:
        s = str(c).lower() if c else ""
        if re.search(r'(20\d{2}|fy\s*20\d{2})', s):
            year_cols += 1

    meta = {}
    if "_meta" in wb.sheetnames:
        for r in wb["_meta"].iter_rows(min_row=1, max_col=2, values_only=True):
            if r[0] and r[1]:
                meta[str(r[0]).strip().lower().replace(" ", "_")] = str(r[1]).strip()
    artifact = meta.get("artifact", "")

    d = t.get('display', 'xlsx')
    if "dcf" in artifact.lower():
        if year_cols < 5:
            warn(f"model_period_floor: {d} DCF should have >=5 forecast years (found {year_cols}).")
    elif "comps" in artifact.lower():
        pass  # Comps doesn't need multi-year
    else:
        # 3-statement
        hist = sum(1 for c in header if re.search(r'(?i)(202[0-5]|fy\s*202[0-5]|actual|historical)', str(c)))
        fcast = sum(1 for c in header if re.search(r'(?i)(202[6-9]|fy\s*202[6-9]|estimate|forecast)', str(c)))
        if hist < 2:
            warn(f"model_period_floor: {d} should have >=2 historical years (found {hist}).")
        if fcast < 3:
            warn(f"model_period_floor: {d} should have >=3 forecast years (found {fcast}).")

    wb.close()
sys.exit(0)
