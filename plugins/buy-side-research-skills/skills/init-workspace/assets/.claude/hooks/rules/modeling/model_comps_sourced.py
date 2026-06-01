"""Check: comps valuation multiples can be derived from actuals + market data."""
import re, sys, os, json, glob
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from common import block, warn
from rules.modeling._common import load_payload, get_xlsx_targets, get_sheet_names_from_payload

VALUATION_KEYWORDS = ["p/e", "pe ratio", "ev/ebitda", "ev/sales", "p/b", "price to book", "fwd pe", "ntm pe"]

payload = load_payload()
for t in get_xlsx_targets(payload):
    import openpyxl
    wb = openpyxl.load_workbook(t["path"], data_only=True)
    sheet_names = wb.sheetnames

    meta = {}
    if "_meta" in wb.sheetnames:
        for r in wb["_meta"].iter_rows(min_row=1, max_col=2, values_only=True):
            if r[0] and r[1]:
                meta[str(r[0]).strip().lower().replace(" ", "_")] = str(r[1]).strip()
    if "comps" not in meta.get("artifact", "").lower():
        wb.close()
        continue

    # Check if any sheet has valuation multiples with source annotation
    has_source = False
    has_multiples = False
    for sn in sheet_names:
        ws = wb[sn]
        text = " ".join(str(c).lower() for row in ws.iter_rows(values_only=True) for c in row if c)
        if any(kw in text for kw in VALUATION_KEYWORDS):
            has_multiples = True
            if re.search(r'(?i)(source|as.of|market data|actuals|yahoo|bloomberg|google finance|bridge)', text):
                has_source = True

    d = t.get('display', 'xlsx')
    if has_multiples and not has_source:
        block(f"comps_sourced: {d} has valuation multiples without source/as-of annotations. Each multiple must be traceable to actuals or market data.")

    wb.close()
sys.exit(0)
