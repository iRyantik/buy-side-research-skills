"""Check: model driver block must match driver-map JSON assumptions (within 2pp)."""
import re, sys, os, json, glob
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from common import block, warn
from rules.modeling._common import load_payload, get_xlsx_targets

DRIVER_FIELDS = {
    "revenue_growth": ["revenue_growth", "revenue growth", "rev growth", "growth rate"],
    "gross_margin": ["gross_margin", "gross margin"],
    "operating_margin": ["operating_margin", "operating margin", "op margin"],
    "capex_to_revenue": ["capex/revenue", "capex_revenue", "capex to revenue"],
}
TOLERANCE = 0.02  # 2 percentage points

def _find_driver_json(ticker):
    candidates = glob.glob(f"industry/*/companies/{ticker}/_cache/driver-map/*.json", recursive=True)
    if candidates:
        for p in candidates:
            with open(p, encoding="utf-8") as f:
                return json.load(f), p
    return None, None

payload = load_payload()
for t in get_xlsx_targets(payload):
    import openpyxl
    wb = openpyxl.load_workbook(t["path"], data_only=True)
    sheet_names = wb.sheetnames

    meta = {}
    if "_meta" in sheet_names:
        for r in wb["_meta"].iter_rows(min_row=1, max_col=2, values_only=True):
            if r[0] and r[1]:
                meta[str(r[0]).strip().lower().replace(" ", "_")] = str(r[1]).strip()

    ticker = meta.get("ticker", "")
    if not ticker:
        wb.close()
        continue

    driver, src_path = _find_driver_json(ticker)
    if not driver:
        wb.close()
        continue

    # Try to find driver block in xlsx — look for Assumptions/Drivers sheet
    driver_sheet = None
    for sn in sheet_names:
        if re.search(r'(?i)(assumptions|drivers?|inputs)', sn):
            driver_sheet = sn
            break
    if not driver_sheet:
        wb.close()
        continue

    ws = wb[driver_sheet]
    rows = list(ws.iter_rows(values_only=True))
    text = "\n".join(" ".join(str(c).lower() for c in r if c) for r in rows)

    failures = []
    assumptions = driver.get("assumptions", driver.get("drivers", {}))
    for field, keywords in DRIVER_FIELDS.items():
        expected = assumptions.get(field)
        if expected is None:
            continue
        expected = float(expected)
        for row in rows:
            row_str = " ".join(str(c).lower() for c in row if c)
            if any(kw in row_str for kw in keywords):
                for c in row:
                    if isinstance(c, (int, float)) and abs(c) > 0:
                        model_val = float(c)
                        if expected > 1:
                            diff = abs(model_val - expected) / expected
                        else:
                            diff = abs(model_val - expected)
                        if diff > TOLERANCE:
                            failures.append(f"{field}: driver={expected:.2%}, model={model_val:.2%})")
                        break
                break

    wb.close()
    if failures:
        block(f"driver_cross_check: {t.get('display','xlsx')} driver assumptions diverge from driver-map. {failures}")
sys.exit(0)
