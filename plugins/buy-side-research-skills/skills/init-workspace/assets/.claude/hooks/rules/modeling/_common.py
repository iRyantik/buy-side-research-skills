"""Shared utilities for modeling hooks."""
import sys, os, json, re, zipfile
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from common import block, warn


def load_payload():
    try:
        return json.load(sys.stdin)
    except Exception:
        return {}


def get_xlsx_targets(payload: dict) -> list[dict]:
    """Return xlsx targets from payload."""
    targets = []
    for t in payload.get("targets", []):
        if t.get("kind") == "file" and (t.get("path", "") or "").endswith(".xlsx"):
            targets.append(t)
    return targets


def get_sheet_names_from_payload(target: dict) -> list[str]:
    """Get sheet names from pre-extracted payload data."""
    return target.get("sheetNames", [])


def get_shared_strings(target: dict) -> str:
    """Get shared strings XML text from payload or extract from xlsx file."""
    text = target.get("sharedStringsText", "")
    if text:
        return text
    # Fallback: extract from xlsx
    path = target.get("path", "")
    if path and os.path.exists(path):
        try:
            with zipfile.ZipFile(path, 'r') as z:
                if 'xl/sharedStrings.xml' in z.namelist():
                    return z.read('xl/sharedStrings.xml').decode('utf-8', errors='replace')
        except Exception:
            pass
    return ""


def get_all_cell_text(target: dict) -> str:
    """Get all cell text from pre-extracted sheet data or xlsx file."""
    sheets = target.get("sheets", [])
    if sheets:
        return "\n".join(str(s.get("Text", "")) for s in sheets)
    # Fallback: read from xlsx using openpyxl
    path = target.get("path", "")
    if path and os.path.exists(path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            texts = []
            for sn in wb.sheetnames:
                ws = wb[sn]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            texts.append(str(cell.value))
            wb.close()
            return "\n".join(texts)
        except Exception:
            pass
    return ""


def get_formula_count(target: dict) -> int:
    """Get total formula count for relevant sheets."""
    sheets = target.get("sheets", [])
    total = sum(s.get("FormulaCount", 0) or 0 for s in sheets)
    if total > 0:
        return total
    # Fallback: count from xlsx
    path = target.get("path", "")
    if path and os.path.exists(path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            count = 0
            for sn in wb.sheetnames:
                ws = wb[sn]
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            count += 1
            wb.close()
            return count
        except Exception:
            pass
    return 0


def search_all_text(target: dict, pattern: str) -> bool:
    """Search pattern in sheet names, shared strings, and cell text."""
    # Sheet names
    for sn in get_sheet_names_from_payload(target):
        if re.search(pattern, sn):
            return True
    # Shared strings
    ss = get_shared_strings(target)
    if re.search(pattern, ss):
        return True
    # Cell text
    ct = get_all_cell_text(target)
    if re.search(pattern, ct):
        return True
    return False
