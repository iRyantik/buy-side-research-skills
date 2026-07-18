"""Check: DCF WACC components must have source + as-of annotations."""
import re, sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from common import block, warn
from rules.modeling._common import load_payload, get_xlsx_targets, get_sheet_names_from_payload

WACC_ITEMS = {
    "risk-free rate": ["risk.free", "risk_free", "10.?year", "government bond"],
    "beta": ["beta", "levered beta", "unlevered beta"],
    "equity risk premium": ["erp", "equity risk premium", "market risk premium"],
    "cost of debt": ["cost of debt", "pre-tax cost", "pre.tax cost"],
    "tax rate": ["tax rate", "marginal tax", "effective tax"],
}
SOURCE_PATTERN = re.compile(r'(?i)(source|from|as.of|as at|collected|bloomberg|yahoo|damodaran|fed|central bank|\d{8})')

payload = load_payload()
for t in get_xlsx_targets(payload):
    import openpyxl
    wb = openpyxl.load_workbook(t["path"], data_only=True)
    sheet_names = wb.sheetnames

    meta = {}
    if "_meta" in wb.sheetnames:
        for r in wb["_meta"].iter_rows(min_row=1, max_col=2, values_only=True):
            if r[0] and r[1]:
                meta[str(r[0]).strip().lower().replace(" ", "_")] = str(r[1]).strip()
    if "dcf" not in meta.get("artifact", "").lower():
        wb.close()
        continue

    # Find WACC sheet or section
    wacc_sheet = None
    for sn in sheet_names:
        if re.search(r'(?i)(wacc|valuation|inputs?|assumptions?)', sn):
            wacc_sheet = sn
            break
    if not wacc_sheet:
        wb.close()
        continue

    ws = wb[wacc_sheet]
    full_text = " ".join(str(c).lower() for row in ws.iter_rows(values_only=True) for c in row if c)

    missing = []
    for label, patterns in WACC_ITEMS.items():
        found = any(re.search(p, full_text) for p in patterns)
        if not found:
            missing.append(label)

    if missing:
        warn(f"dcf_input_sourcing: {t.get('display','xlsx')} WACC section missing components: {missing}.")

    # Check source annotations
    if not SOURCE_PATTERN.search(full_text):
        block(f"dcf_input_sourcing: {t.get('display','xlsx')} WACC inputs must have source and as-of annotations.")

    wb.close()
sys.exit(0)
