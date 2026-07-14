"""Cross-platform Excel control: open, calculate, read cells and formulas.

Windows: win32com via pywin32
macOS: osascript + openpyxl (read after AppleScript calc)
"""
import sys, os, json, tempfile
from pathlib import Path


def get_workbook_data(xlsx_path: str) -> dict:
    """Open xlsx, force calculate, return {sheet_name: {cells: [[value, formula],...], sheet_xml: str}}."""
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"xlsx not found: {xlsx_path}")

    if sys.platform == "win32":
        return _win32_read(xlsx_path)
    else:
        return _macos_read(xlsx_path)


def _win32_read(path: str) -> dict:
    """Windows: win32com open, calculate, read all cells."""
    import win32com.client

    abs_path = str(Path(path).resolve())
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    try:
        wb = excel.Workbooks.Open(abs_path)
        wb.Application.CalculateFull()  # Force full recalc cascade
        wb.Save()

        data = {"sheets": {}}
        for sheet in wb.Worksheets:
            sheet_name = sheet.Name
            used = sheet.UsedRange
            if used is None:
                data["sheets"][sheet_name] = {"cells": [], "formula_count": 0}
                continue

            rows = used.Rows.Count
            cols = used.Columns.Count
            cells = []
            formula_count = 0

            for r in range(1, rows + 1):
                row_cells = []
                for c in range(1, cols + 1):
                    cell = sheet.Cells(r, c)
                    val = cell.Value
                    formula = cell.Formula if str(cell.Formula).startswith("=") else None
                    if formula:
                        formula_count += 1
                    row_cells.append([val, formula])
                cells.append(row_cells)

            data["sheets"][sheet_name] = {
                "cells": cells,
                "formula_count": formula_count,
                "used_rows": rows,
                "used_cols": cols,
            }

        wb.Close(SaveChanges=False)
        return data
    finally:
        excel.Quit()


def _macos_read(path: str) -> dict:
    """macOS: AppleScript calc, then openpyxl read cached values."""
    import subprocess

    abs_path = str(Path(path).resolve())
    # AppleScript: open, calculate, save, close
    script = f'''
        tell application "Microsoft Excel"
            open POSIX file "{abs_path}"
            tell active workbook
                calculate
                save
                close
            end tell
            quit
        end tell
    '''
    subprocess.run(["osascript", "-e", script], capture_output=True, timeout=60)

    # Now read with openpyxl (cached values are present after calc+save)
    return _openpyxl_read(abs_path)


def _openpyxl_read(path: str) -> dict:
    """Fallback: openpyxl read (no calc — use after _macos_calc or for inspection)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=False)
    data = {"sheets": {}}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cells = []
        formula_count = 0

        for row in ws.iter_rows(max_row=ws.max_row, max_col=ws.max_column):
            row_cells = []
            for cell in row:
                val = cell.value
                formula = None
                if isinstance(cell, openpyxl.cell.cell.Cell):
                    # Check if formula
                    pass
                row_cells.append([val, None])
            cells.append(row_cells)

        data["sheets"][sheet_name] = {
            "cells": cells,
            "formula_count": formula_count,
            "used_rows": ws.max_row or 0,
            "used_cols": ws.max_column or 0,
        }

    wb.close()
    return data


def get_sheet_names(xlsx_path: str) -> list[str]:
    """Quick read: just sheet names, no calc."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def get_shared_strings_text(xlsx_path: str) -> str:
    """Extract shared strings XML from xlsx for text pattern matching."""
    import zipfile
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            if 'xl/sharedStrings.xml' in z.namelist():
                return z.read('xl/sharedStrings.xml').decode('utf-8', errors='replace')
    except Exception:
        pass
    return ""
