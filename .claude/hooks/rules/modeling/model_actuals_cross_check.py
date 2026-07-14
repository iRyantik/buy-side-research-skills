"""Check: model historical columns must match actuals-resolved.json (within 1%)."""
import re, sys, os, json, glob
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from common import block, warn
from rules.modeling._common import load_payload, get_xlsx_targets, get_sheet_names_from_payload

FIELDS = {
    "revenue": ("IS", ["revenue", "total_revenue", "sales"]),
    "net_income": ("IS", ["net_income", "net_income_attributable"]),
    "total_assets": ("BS", ["total_assets"]),
    "total_equity": ("BS", ["total_equity", "shareholders_equity"]),
    "operating_cf": ("CF", ["operating_cf", "cash_from_operations"]),
    "capex": ("CF", ["capex", "capital_expenditure"]),
}
TOLERANCE = 0.01  # 1%

def _find_actuals(actuals_path: str, xlsx_path: str):
    """Load actuals-resolved.json, return {field: value_or_None}."""
    candidates = [actuals_path]
    # Resolve relative to xlsx directory, then walk up to find workspace root
    xlsx_dir = os.path.dirname(os.path.abspath(xlsx_path))
    d = xlsx_dir
    for _ in range(6):
        g = os.path.join(d, actuals_path)
        if os.path.exists(g) and g not in candidates:
            candidates.append(g)
        parent = os.path.dirname(d)
        if parent == d:
            break
        d = parent
    for path in candidates:
        if os.path.exists(path):
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
            fy = data.get("income_statement", {}).get("latest_fy", {})
            bs = data.get("balance_sheet", {}).get("latest_fy", {})
            cf = data.get("cash_flow", {}).get("latest_fy", {})
            result = {}
            for field_name, (section, keys) in FIELDS.items():
                src = {"IS": fy, "BS": bs, "CF": cf}[section]
                for k in keys:
                    v = src.get(k)
                    if isinstance(v, dict):
                        v = v.get("value")
                    if v is not None and v != 0:
                        result[field_name] = float(v)
                        break
            return result, path
    return None, None

def _find_value_in_row(row, keywords):
    """Check if row labels contain any keyword; return first numeric cell in that row."""
    row_text = " ".join(str(c).lower() for c in row if c)
    for kw in keywords:
        if kw in row_text:
            for c in row:
                if isinstance(c, (int, float)) and abs(c) > 0:
                    return float(c)
    return None

payload = load_payload()
for t in get_xlsx_targets(payload):
    import openpyxl
    wb = openpyxl.load_workbook(t["path"], data_only=True)

    # Read _meta
    if "_meta" not in wb.sheetnames:
        wb.close()
        continue
    meta = {}
    for r in wb["_meta"].iter_rows(min_row=1, max_col=2, values_only=True):
        if r[0] and r[1]:
            meta[str(r[0]).strip().lower().replace(" ", "_")] = str(r[1]).strip()

    actuals_ref = meta.get("actuals_run_id", "") or meta.get("actuals_path", "")
    if not actuals_ref:
        wb.close()
        continue

    actuals, src_path = _find_actuals(actuals_ref, t["path"])
    if not actuals:
        warn(f"actuals_cross_check: cannot find actuals-resolved.json for {meta.get('ticker','?')}")
        wb.close()
        continue

    failures = []
    for sheet_name in ["IS", "BS", "CF"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # Find the first numeric column that looks like a historical year
        header = [str(c).lower() if c else "" for c in rows[0]]
        hist_col = None
        for ci, h in enumerate(header):
            if any(y in h for y in ["fy2025", "2025a", "2025", "fy2024", "2024a", "2024", "historical", "actual"]):
                hist_col = ci
                break
        if hist_col is None:
            continue

        for field, (section, keywords) in FIELDS.items():
            if section != sheet_name:
                continue
            expected = actuals.get(field)
            if expected is None:
                continue
            # Find matching row
            for row in rows:
                model_val = _find_value_in_row(row, keywords)
                if model_val is not None:
                    if hist_col < len(row) and isinstance(row[hist_col], (int, float)):
                        cell_val = float(row[hist_col])
                        if abs(cell_val - expected) / abs(expected) > TOLERANCE:
                            failures.append(f"{sheet_name}.{field}: model={cell_val:,.0f}, actuals={expected:,.0f} (diff={abs(cell_val-expected)/abs(expected)*100:.1f}%)")
                    break

    wb.close()
    if failures:
        block(f"actuals_cross_check: {t.get('display','xlsx')} has {len(failures)} mismatches vs actuals. {failures[:5]}...")
sys.exit(0)
