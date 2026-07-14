"""Check: DCF key inputs (FCF/EBIT/D&A/CapEx) should be formulas referencing 3sm or have verified sources."""
import re, sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from common import block, warn
from rules.modeling._common import load_payload, get_xlsx_targets, get_sheet_names_from_payload

KEY_INPUTS = ["free cash flow", "fcf", "ebit", "ebitda", "d&a", "depreciation", "capex", "capital expenditure", "change in working capital", "wc change"]
EXTERNAL_REF = re.compile(r'\[[^\]]+\]')
FORMULA = re.compile(r'^=')

payload = load_payload()
for t in get_xlsx_targets(payload):
    import openpyxl
    wb = openpyxl.load_workbook(t["path"])  # data_only=False to read formulas
    sheet_names = wb.sheetnames

    meta = {}
    if "_meta" in wb.sheetnames:
        for r in wb["_meta"].iter_rows(min_row=1, max_col=2, values_only=True):
            if r[0] and r[1]:
                meta[str(r[0]).strip().lower().replace(" ", "_")] = str(r[1]).strip()
    if "dcf" not in meta.get("artifact", "").lower():
        wb.close()
        continue

    warnings = []
    hardcoded = []
    for sn in sheet_names:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                val_str = str(cell.value)
                if not isinstance(cell.value, str):
                    continue
                if any(kw in val_str.lower() for kw in KEY_INPUTS):
                    # This is a label row — skip
                    pass
                # Check actual number cells
                if isinstance(cell.value, (int, float)) and abs(cell.value) > 1000:
                    # Check if adjacent label cell contains a key input name
                    pass

    # Simplified approach: search for hardcoded large numbers in DCF sheets
    for sn in sheet_names:
        ws = wb[sn]
        for row in ws.iter_rows():
            label = " ".join(str(c).lower() for c in [cell.value for cell in row] if c and isinstance(c.value, str))
            for cell in row:
                if isinstance(cell.value, (int, float)) and abs(cell.value) > 1e6:
                    for kw in KEY_INPUTS:
                        if kw in label:
                            # Check if this cell has a formula or is hardcoded
                            if not FORMULA.match(str(cell.value)):
                                hardcoded.append(f"{sn}: {kw}={cell.value:,.0f}")
                            break

    wb.close()
    if hardcoded:
        warn(f"dcf_linked_to_3sm: {t.get('display','xlsx')} has hardcoded key inputs that should reference 3sm: {hardcoded[:5]}")
sys.exit(0)
