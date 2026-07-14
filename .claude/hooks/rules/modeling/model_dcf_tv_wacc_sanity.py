"""Warn: TV growth unrealistic or WACC out of typical range."""
import re, sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from common import block, warn
from rules.modeling._common import load_payload, get_xlsx_targets, get_sheet_names_from_payload

def _extract_pct(text, patterns):
    for p in patterns:
        m = re.search(p + r'\s*[=:]\s*([\d.]+)\s*%?', text)
        if m:
            return float(m.group(1))
        m = re.search(r'([\d.]+)\s*%.{0,30}' + p, text)
        if m:
            return float(m.group(1))
    return None

payload = load_payload()
for t in get_xlsx_targets(payload):
    import openpyxl
    wb = openpyxl.load_workbook(t["path"], data_only=True)
    sheet_names = wb.sheetnames

    meta = {}
    if "_meta" in wb.sheetnames:
        for r in wb["_meta"].iter_rows(min_row=1, max_col=2, values_only=True):
            if r[0] and r[1]:
                meta[str(r[0]).strip().lower().replace(" ", "_")] = str(r[1]).strip()
    if "dcf" not in meta.get("artifact", "").lower():
        wb.close()
        continue

    full_text = ""
    for sn in sheet_names:
        if re.search(r'(?i)(wacc|valuation|terminal|dcf)', sn):
            ws = wb[sn]
            full_text += " " + " ".join(str(c).lower() for row in ws.iter_rows(values_only=True) for c in row if c)

    wacc = None
    for cell_text in full_text.split():
        m = re.search(r'wacc.*?([\d.]+)\s*%', full_text)
        if m:
            wacc = float(m.group(1))
            break

    tv_growth = None
    tg_pat = r'(?i)(terminal growth|perpetuity growth|long.?term growth)'
    m = re.search(tg_pat + r'.{0,50}([\d.]+)\s*%', full_text)
    if m:
        tv_growth = float(m.group(2))

    d = t.get('display', 'xlsx')
    if wacc is not None:
        if wacc < 6 or wacc > 15:
            warn(f"dcf_tv_wacc_sanity: {d} WACC={wacc}% outside typical range 6-15%.")
    if tv_growth is not None:
        if tv_growth > 5:
            warn(f"dcf_tv_wacc_sanity: {d} terminal growth={tv_growth}% seems high (most economies <5%).")
        if wacc is not None and tv_growth >= wacc - 1:
            warn(f"dcf_tv_wacc_sanity: {d} terminal growth ({tv_growth}%) close to or above WACC ({wacc}%).")

    wb.close()
sys.exit(0)
