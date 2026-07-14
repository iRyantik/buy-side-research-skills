"""Warn: comps analysis should indicate GAAP vs Non-GAAP consistency."""
import re, sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from common import block, warn
from rules.modeling._common import load_payload, get_xlsx_targets, get_sheet_names_from_payload

GAAP_TERMS = ["gaap", "reported", "ifrs", "as reported"]
NONGAAP_TERMS = ["non-gaap", "adjusted", "non gaap", "underlying"]

payload = load_payload()
for t in get_xlsx_targets(payload):
    import openpyxl
    wb = openpyxl.load_workbook(t["path"], data_only=True)
    sheet_names = wb.sheetnames

    meta = {}
    if "_meta" in wb.sheetnames:
        for r in wb["_meta"].iter_rows(min_row=1, max_col=2, values_only=True):
            if r[0] and r[1]:
                meta[str(r[0]).strip().lower().replace(" ", "_")] = str(r[1]).strip()
    if "comps" not in meta.get("artifact", "").lower():
        wb.close()
        continue

    full_text = ""
    for sn in sheet_names:
        ws = wb[sn]
        full_text += " " + " ".join(str(c).lower() for row in ws.iter_rows(values_only=True) for c in row if c)

    has_gaap = any(t in full_text for t in GAAP_TERMS)
    has_nongaap = any(t in full_text for t in NONGAAP_TERMS)

    if not has_gaap and not has_nongaap:
        warn(f"comps_denominator_parity: {t.get('display','xlsx')} does not specify GAAP vs Non-GAAP basis for earnings/EBITDA. Peer denominators may not be comparable.")
    elif has_nongaap and not has_gaap:
        # Only Non-GAAP mentioned — warn that GAAP isn't referenced
        pass  # This is acceptable if intentional

    wb.close()
sys.exit(0)
