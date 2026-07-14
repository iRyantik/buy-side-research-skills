"""Check: IS net income flows to BS retained earnings delta, CF cash ties to BS cash, debt feeds IS interest."""
import re, sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from common import block, warn
from rules.modeling._common import load_payload, get_xlsx_targets, get_sheet_names_from_payload

def _find_value(ws, row_keywords, col_offset=1):
    """Find first row matching keywords, return the numeric value in the col_offset-th numeric column."""
    for row in ws.iter_rows(values_only=True):
        text = " ".join(str(c).lower() for c in row if c)
        if all(kw.lower() in text for kw in row_keywords):
            count = 0
            for c in row:
                if isinstance(c, (int, float)):
                    if count == col_offset:
                        return float(c)
                    count += 1
    return None

payload = load_payload()
for t in get_xlsx_targets(payload):
    import openpyxl
    wb = openpyxl.load_workbook(t["path"], data_only=True)
    sheet_names = wb.sheetnames
    sheets = {}
    for sn in ["IS", "BS", "CF"]:
        if sn in sheet_names:
            sheets[sn] = wb[sn]

    issues = []

    # 1. IS Net Income → check if BS Retained Earnings references it
    if "IS" in sheets and "BS" in sheets:
        ni = _find_value(sheets["IS"], ["net", "income"], 0)
        if ni is not None:
            re_val = _find_value(sheets["BS"], ["retained", "earnings"], 0)
            if re_val is None:
                re_val = _find_value(sheets["BS"], ["total", "equity"], 0)
            if re_val is not None and abs(re_val) < 1:
                issues.append("BS retained earnings/equity appears zero — check IS NI → BS RE linkage")

    # 2. CF ending cash ↔ BS cash (last year)
    if "CF" in sheets and "BS" in sheets:
        cf_rows = list(sheets["CF"].iter_rows(values_only=True))
        bs_rows = list(sheets["BS"].iter_rows(values_only=True))
        # Find last column with data
        bs_cash = _find_value(sheets["BS"], ["cash"], -1)  # last column
        if bs_cash is None:
            bs_cash = _find_value(sheets["BS"], ["cash"], 1)
        # Find CF ending cash
        cf_end = None
        for row in cf_rows:
            text = " ".join(str(c).lower() for c in row if c)
            if "ending" in text and "cash" in text:
                for c in reversed(row):
                    if isinstance(c, (int, float)):
                        cf_end = float(c)
                        break
                break
        if cf_end is not None and bs_cash is not None:
            if cf_end == 0 and bs_cash == 0:
                issues.append("CF ending cash and BS cash both zero — model may not be integrated")
            elif abs(cf_end - bs_cash) / max(abs(bs_cash), 1) > 0.02:
                issues.append(f"CF ending cash ({cf_end:,.0f}) != BS cash ({bs_cash:,.0f}) — cash tie-out broken")

    # 3. BS debt → IS interest expense (if debt exists)
    if "BS" in sheets and "IS" in sheets:
        debt = None
        for kw in ["long_term_debt", "long term debt", "total debt"]:
            d = _find_value(sheets["BS"], [kw.replace("_", " ")], 1)
            if d and d > 0:
                debt = d
                break
        if debt and debt > 1000:
            interest = _find_value(sheets["IS"], ["interest", "expense"], 1)
            if interest is None or abs(interest) < 1:
                issues.append(f"BS has debt ({debt:,.0f}) but IS interest expense is missing/zero")

    wb.close()
    if issues:
        block(f"model_internal_consistency: {t.get('display','xlsx')} has cross-sheet linkage issues. {issues}")
sys.exit(0)
