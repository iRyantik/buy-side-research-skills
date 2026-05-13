#!/usr/bin/env python3
"""Build IONQ DCF Model xlsx from financial evidence pack + driver map.

Usage:
    python scripts/build-ionq-dcf.py

Output:
    examples/financial-data-pull/us/ionq/ionq-dcf-model.xlsx
"""

from __future__ import annotations

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent.parent / "examples" / "financial-data-pull" / "us" / "ionq"
DATA_FILE = BASE / "financials.normalized.json"
OUTPUT = BASE / "ionq-dcf-model.xlsx"

# ── Color palette ──────────────────────────────────────────────────────────
DARK_BLUE = "1F3864"
MED_BLUE = "4472C4"
LIGHT_BLUE = "D6E4F0"
DARK_GRAY = "404040"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
RED_FILL = "FCE4EC"
AMBER_FILL = "FFF3E0"
GREEN_FILL = "E8F5E9"

hdr_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
hdr_fill = PatternFill("solid", fgColor=DARK_BLUE)
sub_hdr_font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=11)
sub_hdr_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
num_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", bold=True, size=10)
label_font = Font(name="Calibri", bold=True, size=10, color=DARK_GRAY)
thin_border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

NUM_FMT = '#,##0'
NUM_FMT_DEC = '#,##0.0'
PCT_FMT = '0.0%'


def load_data() -> dict:
    return json.loads(DATA_FILE.read_text(encoding="utf-8"))


def get_concept(data: dict, statement: str, concept: str) -> float | None:
    items = data.get(statement, [])
    for item in items:
        if item.get("concept") == concept:
            return item
    return None


def get_value(item: dict | None, period: str) -> float | None:
    if item is None:
        return None
    return item.get("values", {}).get(period)


def safe(v, default=0.0):
    return v if v is not None else default


def style_header_row(ws, row, max_col, fill=None):
    fill = fill or hdr_fill
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = hdr_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border


def style_data_cell(cell, is_bold=False, is_pct=False):
    cell.font = bold_font if is_bold else num_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="right" if cell.column > 1 else "left")
    if is_pct:
        cell.number_format = PCT_FMT
    elif cell.column > 1:
        cell.number_format = NUM_FMT


def write_section_header(ws, row, label, cols_max):
    ws.cell(row=row, column=1, value=label).font = sub_hdr_font
    for c in range(1, cols_max + 1):
        ws.cell(row=row, column=c).fill = sub_hdr_fill
        ws.cell(row=row, column=c).border = thin_border
    return row + 1


def write_data_row(ws, row, label, values, is_bold=False, is_pct=False):
    ws.cell(row=row, column=1, value=label).font = bold_font if is_bold else label_font
    ws.cell(row=row, column=1).border = thin_border
    for i, v in enumerate(values, start=2):
        cell = ws.cell(row=row, column=i, value=v)
        style_data_cell(cell, is_bold=is_bold, is_pct=is_pct)
    return row + 1


# ────────────────────────────────────────────────────────────────────────────
def build():
    data = load_data()
    d = data  # shorthand
    periods_hist = ["FY 2022", "FY 2023", "FY 2024", "FY 2025"]
    periods_proj = ["FY 2026E", "FY 2027E", "FY 2028E", "FY 2029E", "FY 2030E"]
    all_periods = periods_hist + periods_proj
    n_hist = len(periods_hist)
    n_proj = len(periods_proj)
    n_all = len(all_periods)

    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════════════
    # Helper: extract historical values from normalized JSON
    # ═══════════════════════════════════════════════════════════════════════
    def hist_val(statement: str, concept: str) -> list[float]:
        item = get_concept(d, statement, concept)
        return [safe(get_value(item, p)) for p in periods_hist]

    def all_val(statement: str, concept: str, proj: list[float]) -> list[float]:
        """Hist (from data) + proj (passed in)."""
        h = hist_val(statement, concept)
        return h + list(proj)

    # ── Actuals from normalized JSON ──────────────────────────────────────
    rev_hist = hist_val("income_statement", "RevenueFromContractWithCustomerExcludingAssessedTax")
    cogs_hist = hist_val("income_statement", "CostOfGoodsAndServiceExcludingDepreciationDepletionAndAmortization")
    rnd_hist = hist_val("income_statement", "ResearchAndDevelopmentExpense")
    sga_hist = hist_val("income_statement", "SellingAndMarketingExpense")
    ga_hist = hist_val("income_statement", "GeneralAndAdministrativeExpense")
    da_hist = hist_val("income_statement", "DepreciationDepletionAndAmortizationNonproduction")
    oi_hist = hist_val("income_statement", "OperatingIncomeLoss")
    interest_hist = hist_val("income_statement", "InterestIncomeExpenseNet")
    tax_hist = hist_val("income_statement", "IncomeTaxExpenseBenefit")
    net_income_hist = hist_val("income_statement", "NetIncomeLossAttributableToParent")

    # BS actuals
    cash_hist = hist_val("balance_sheet", "CashAndCashEquivalentsAtCarryingValue")
    ar_hist = hist_val("balance_sheet", "AccountsReceivableAfterAllowanceForCreditLossCurrent")
    inv_hist = hist_val("balance_sheet", "InventoryNet")
    other_ca_hist_raw = hist_val("balance_sheet", "PrepaidExpenseAndOtherAssetsCurrent")
    st_inv_hist = hist_val("balance_sheet", "ShortTermInvestments")
    total_ca_hist = hist_val("balance_sheet", "AssetsCurrent")
    ppe_hist = hist_val("balance_sheet", "PropertyPlantAndEquipmentNet")
    goodwill_hist = hist_val("balance_sheet", "Goodwill")
    intang_hist = hist_val("balance_sheet", "IntangibleAssetsNetExcludingGoodwill")
    other_na_hist = hist_val("balance_sheet", "OtherAssetsNoncurrent")
    total_assets_hist = hist_val("balance_sheet", "Assets")
    ap_hist = hist_val("balance_sheet", "AccountsPayableCurrent")
    accrued_hist = hist_val("balance_sheet", "AccruedLiabilitiesCurrent")
    defrev_c_hist = hist_val("balance_sheet", "DeferredRevenueCurrent")
    ocl_hist_raw = hist_val("balance_sheet", "OperatingLeaseLiabilityCurrent")
    total_cl_hist = hist_val("balance_sheet", "LiabilitiesCurrent")
    defrev_nc_hist = hist_val("balance_sheet", "DeferredRevenueNoncurrent")
    other_ncl_hist = hist_val("balance_sheet", "OtherLiabilitiesNoncurrent")
    op_lease_nc_hist = hist_val("balance_sheet", "OperatingLeaseLiabilityNoncurrent")
    deferred_tax_hist = hist_val("balance_sheet", "DeferredTaxLiabilitiesNet")
    total_liab_hist = hist_val("balance_sheet", "Liabilities")
    apic_hist = hist_val("balance_sheet", "AdditionalPaidInCapital")
    retained_hist = hist_val("balance_sheet", "RetainedEarningsAccumulatedDeficit")
    total_equity_hist = hist_val("balance_sheet", "StockholdersEquityAttributableToParent")

    # CF actuals
    sbc_hist = hist_val("cash_flow", "ShareBasedCompensation")
    dep_hist = hist_val("cash_flow", "Depreciation")
    amort_intang_hist = hist_val("cash_flow", "AmortizationOfIntangibleAssets")
    # Combined D&A
    da_cf_hist = hist_val("cash_flow", "DepreciationAndAmortization")
    # Use D&A from CF if available, else from IS
    da_combined_hist = da_cf_hist if any(da_cf_hist) else da_hist

    # ═══════════════════════════════════════════════════════════════════════
    # Projections (Base Scenario)
    # ═══════════════════════════════════════════════════════════════════════
    # RPO model: FY2025 closing RPO = ~$370M; 40% converts in 12 months
    # Revenue build: organic QCaaS growth + RPO conversion + acquisitions step-in
    base_rev = [200_000_000, 300_000_000, 420_000_000, 550_000_000, 700_000_000]  # 40% CAGR
    bull_rev = [280_000_000, 480_000_000, 720_000_000, 950_000_000, 1_200_000_000]
    bear_rev = [140_000_000, 180_000_000, 220_000_000, 260_000_000, 300_000_000]

    # Cost projections
    gross_margin_pct = [0.35, 0.38, 0.42, 0.45, 0.48]  # improving with scale
    base_cogs = [base_rev[i] * (1 - gross_margin_pct[i]) for i in range(n_proj)]
    sbc_proj = [350_000_000, 380_000_000, 400_000_000, 420_000_000, 440_000_000]
    # Cash opex (ex SBC) as % of rev
    cash_opex_pct = [1.2, 0.9, 0.7, 0.55, 0.45]
    da_proj = [95_000_000, 110_000_000, 125_000_000, 140_000_000, 155_000_000]
    capex_proj = [20_000_000, 25_000_000, 30_000_000, 35_000_000, 40_000_000]

    # Balance sheet projections (simplified)
    # NWC as % of revenue
    def nwc_projections(revenue_proj):
        ar_pct = 0.05
        ap_pct = 0.02
        defrev_pct = 0.03  # deferred revenue grows with rev
        accrued_pct = 0.06
        inv_pct = 0.01
        ar_p = [r * ar_pct for r in revenue_proj]
        inv_p = [r * inv_pct for r in revenue_proj]
        ap_p = [r * ap_pct for r in revenue_proj]
        accrued_p = [r * accrued_pct for r in revenue_proj]
        defrev_p = [r * defrev_pct for r in revenue_proj]
        return ar_p, inv_p, ap_p, accrued_p, defrev_p

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 1: Summary
    # ═══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = DARK_BLUE

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 60

    r = 1
    ws.cell(row=r, column=1, value="IONQ DCF Model — Summary").font = Font(name="Calibri", bold=True, size=16, color=DARK_BLUE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r = 2
    ws.cell(row=r, column=1, value="Built: 3-statement-model / dcf-model / comps-analysis / model-update skill | Source: 10-K FY2025 | Data as-of: Dec 31, 2025").font = Font(name="Calibri", size=9, italic=True, color="808080")

    r = 4
    ws.cell(row=r, column=1, value="⚠️  No Price Target Guard").font = Font(name="Calibri", bold=True, size=12, color="C62828")
    r = 5
    ws.cell(row=r, column=1, value="This model is scenario-based. All revenue drivers are Low confidence (see driver-map.md).").font = num_font
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

    r = 7
    ws.cell(row=r, column=1, value="Key Market Data").font = sub_hdr_font
    ws.cell(row=r, column=1).fill = sub_hdr_fill
    ws.cell(row=r, column=2).fill = sub_hdr_fill
    r += 1
    for label, val in [
        ("Stock Price", "$56.89"),
        ("Market Cap", "$21.24B"),
        ("Shares Outstanding", "373M"),
        ("Cash + ST Investments (FY2025)", "$2.4B"),
        ("Enterprise Value (approx)", "~$18.8B"),
        ("TTM Revenue (FY2025)", "$130M"),
        ("Implied EV/Sales (TTM)", "~145x"),
        ("WACC (central)", "13.0%"),
        ("Terminal EV/Sales (central)", "8x"),
        ("Forecast Horizon", "FY2026E-FY2030E"),
    ]:
        ws.cell(row=r, column=1, value=label).font = label_font
        ws.cell(row=r, column=2, value=val).font = num_font
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2).border = thin_border
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Revenue Scenarios (Base | Bull | Bear)").font = sub_hdr_font
    ws.cell(row=r, column=1).fill = sub_hdr_fill
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1
    for label, vals in [
        ("Base (FY2026→2030)", base_rev),
        ("Bull (FY2026→2030)", bull_rev),
        ("Bear (FY2026→2030)", bear_rev),
    ]:
        ws.cell(row=r, column=1, value=label).font = label_font
        ws.cell(row=r, column=2, value=f"${vals[0]/1e6:.0f}M → ${vals[-1]/1e6:.0f}M  (CAGR {((vals[-1]/vals[0])**(1/4)-1)*100:.0f}%)").font = num_font
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2).border = thin_border
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Key Assumptions").font = sub_hdr_font
    ws.cell(row=r, column=1).fill = sub_hdr_fill
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1
    for label, val in [
        ("RPO Opening FY2026", "$370M (10-K)"),
        ("RPO 12-month conversion", "~40% (10-K)"),
        ("Terminal FCF margin (2030)", "~20% of revenue"),
        ("Terminal EV/Sales range", "6x – 10x"),
        ("WACC range", "12% – 15%"),
        ("Annual share dilution", "~3-5%"),
        ("Cash runway (at base burn)", "~6-8 years (no near-term dilution)"),
    ]:
        ws.cell(row=r, column=1, value=label).font = label_font
        ws.cell(row=r, column=2, value=val).font = num_font
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2).border = thin_border
        r += 1

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 2: Income Statement
    # ═══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Income Statement")
    ws2.sheet_properties.tabColor = MED_BLUE
    ws2.column_dimensions["A"].width = 42
    for i in range(2, n_all + 2):
        ws2.column_dimensions[get_column_letter(i)].width = 16

    headers = ["$ thousands"] + all_periods
    max_c = len(headers)
    r = 1
    ws2.cell(row=r, column=1, value="IONQ — Income Statement (Actuals + Base Projections)").font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_c)
    r = 2
    for ci, h in enumerate(headers, start=1):
        ws2.cell(row=r, column=ci, value=h)
    style_header_row(ws2, r, max_c)

    # Compute projected lines
    base_gross = [base_rev[i] - base_cogs[i] for i in range(n_proj)]
    base_rnd = [base_rev[i] * 0.7 for i in range(n_proj)]  # declining toward 40%
    base_rnd = [min(base_rnd[i], base_rnd[0] * (0.85 ** i)) for i in range(n_proj)]
    base_sga = [base_rev[i] * 0.15 for i in range(n_proj)]
    base_ga = [base_rev[i] * 0.25 for i in range(n_proj)]
    # SBC allocated across R&D/S&M/G&A — total SBC in separate row
    base_total_opex = [base_cogs[i] + base_rnd[i] + base_sga[i] + base_ga[i] + da_proj[i] for i in range(n_proj)]
    base_ebit = [base_rev[i] - base_total_opex[i] for i in range(n_proj)]
    # Interest declines as cash burns down
    base_interest = [55_000_000, 45_000_000, 35_000_000, 25_000_000, 20_000_000]
    base_pretax = [base_ebit[i] + base_interest[i] for i in range(n_proj)]
    base_tax = [0, 0, 0, 0, 5_000_000]  # assume NOL shields through FY2029
    base_ni = [base_pretax[i] - base_tax[i] for i in range(n_proj)]

    rows_is = [
        ("", "", True),
        ("Revenue", rev_hist + base_rev, False),
        ("Cost of Revenue (ex-D&A)", cogs_hist + base_cogs, False),
        ("Gross Profit", [rev_hist[i] - cogs_hist[i] for i in range(n_hist)] + base_gross, True),
        ("", "", True),
        ("Operating Expenses", "", True),
        ("  Research & Development", rnd_hist + base_rnd, False),
        ("  Sales & Marketing", sga_hist + base_sga, False),
        ("  General & Administrative", ga_hist + base_ga, False),
        ("  Depreciation & Amortization", da_combined_hist + da_proj, False),
        ("Total Operating Expenses", [rnd_hist[i] + sga_hist[i] + ga_hist[i] + da_combined_hist[i] for i in range(n_hist)] + base_total_opex, False),
        ("", "", True),
        ("Operating Income (EBIT)", oi_hist + base_ebit, True),
        ("Interest Income (Expense), Net", interest_hist + base_interest, False),
        ("Pre-Tax Income", [oi_hist[i] + interest_hist[i] for i in range(n_hist)] + base_pretax, False),
        ("Income Tax (Benefit)", tax_hist + base_tax, False),
        ("Net Income (to Parent)", net_income_hist + base_ni, True),
        ("", "", True),
        ("Memo: Stock-Based Compensation", sbc_hist + sbc_proj, False),
        ("Memo: EBITDA (EBIT + D&A)", [oi_hist[i] + da_combined_hist[i] for i in range(n_hist)] + [base_ebit[i] + da_proj[i] for i in range(n_proj)], False),
    ]

    r = 3
    for label, vals, is_bold in rows_is:
        if label == "":
            r += 1
            continue
        r = write_data_row(ws2, r, label, vals, is_bold=is_bold)
        # Color actuals vs projected boundary
        if is_bold and label != "":
            for c in range(n_hist + 1, n_all + 1):
                ws2.cell(row=r - 1, column=c + 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    # Auto-format actuals/projected boundary with a light blue column
    r_start = 3
    r_end = r
    for c in range(n_hist + 2, n_all + 2):
        for rr in range(r_start, r_end):
            cell = ws2.cell(row=rr, column=c)
            if cell.value and isinstance(cell.value, (int, float)) and cell.value > 0:
                pass  # keep formatting

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 3: Balance Sheet
    # ═══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Balance Sheet")
    ws3.sheet_properties.tabColor = MED_BLUE
    ws3.column_dimensions["A"].width = 42
    for i in range(2, n_all + 2):
        ws3.column_dimensions[get_column_letter(i)].width = 16

    r = 1
    ws3.cell(row=r, column=1, value="IONQ — Balance Sheet (Actuals + Base Projections)").font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_c)
    r = 2
    for ci, h in enumerate(headers, start=1):
        ws3.cell(row=r, column=ci, value=h)
    style_header_row(ws3, r, max_c)

    # BS projections
    cash_proj = [0.0] * n_proj  # will calc from CF
    ar_p, inv_p, ap_p, accrued_p, defrev_p = nwc_projections(base_rev)

    # Non-cash adjustments for BS
    ppe_proj = [ppe_hist[-1] + capex_proj[i] - da_proj[i] * 0.3 for i in range(n_proj)]
    # Goodwill & intang amortize slowly
    gw_proj = [goodwill_hist[-1] * (1 - 0.02 * i) for i in range(1, n_proj + 1)]
    intang_proj = [intang_hist[-1] - da_proj[i] * 0.7 for i in range(n_proj)]
    other_ca_proj = [other_ca_hist_raw[-1] * 1.1 for _ in range(n_proj)]  # placeholder
    st_inv_proj = [st_inv_hist[-1] * 0.8 for _ in range(n_proj)]  # deplete slowly
    ocl_proj = [ap_p[i] * 0.3 for i in range(n_proj)]
    defrev_nc_proj = [defrev_nc_hist[-1] * 1.1 for _ in range(n_proj)]
    other_ncl_proj = [other_ncl_hist[-1] * 1.1 for _ in range(n_proj)]
    op_lease_nc_proj = [op_lease_nc_hist[-1] * 0.9 for _ in range(n_proj)]
    deferred_tax_proj = [deferred_tax_hist[-1] * 1.1 for _ in range(n_proj)]

    # Equity (accumulated)
    apic_proj = [apic_hist[-1] + sbc_proj[i] * 0.8 for i in range(n_proj)]  # SBC new equity
    retained_proj_base = [retained_hist[-1]]
    for i in range(n_proj):
        retained_proj_base.append(retained_proj_base[-1] + base_ni[i])

    rows_bs = [
        ("", "", True),
        ("Assets", "", True),
        ("  Cash & Cash Equivalents", cash_hist + cash_proj, False),
        ("  Short-Term Investments", st_inv_hist + st_inv_proj, False),
        ("  Accounts Receivable", ar_hist + ar_p, False),
        ("  Inventory", inv_hist + inv_p, False),
        ("  Prepaid & Other CA", other_ca_hist_raw + other_ca_proj, False),
        ("Total Current Assets", [total_ca_hist[i] if i < n_hist else 0 for i in range(n_all)], True),
        ("", "", True),
        ("  PP&E, Net", ppe_hist + ppe_proj, False),
        ("  Goodwill", goodwill_hist + gw_proj, False),
        ("  Intangibles, Net", intang_hist + intang_proj, False),
        ("  Other Noncurrent Assets", other_na_hist + [other_na_hist[-1] * 1.05 for _ in range(n_proj)], False),
        ("Total Assets", total_assets_hist + [0.0] * n_proj, True),
        ("", "", True),
        ("Liabilities & Equity", "", True),
        ("  Accounts Payable", ap_hist + ap_p, False),
        ("  Accrued Liabilities", accrued_hist + accrued_p, False),
        ("  Deferred Revenue (Current)", defrev_c_hist + defrev_p, False),
        ("  Other Current Liabilities", ocl_hist_raw + ocl_proj, False),
        ("Total Current Liabilities", total_cl_hist + [ap_p[i] + accrued_p[i] + defrev_p[i] + ocl_proj[i] for i in range(n_proj)], True),
        ("", "", True),
        ("  Deferred Revenue (Noncurrent)", defrev_nc_hist + defrev_nc_proj, False),
        ("  Deferred Tax Liabilities", deferred_tax_hist + deferred_tax_proj, False),
        ("  Operating Lease (Noncurrent)", op_lease_nc_hist + op_lease_nc_proj, False),
        ("  Other Noncurrent Liabilities", other_ncl_hist + other_ncl_proj, False),
        ("Total Liabilities", total_liab_hist + [0.0] * n_proj, True),
        ("", "", True),
        ("  Additional Paid-in Capital", apic_hist + apic_proj, False),
        ("  Retained Earnings (Deficit)", retained_hist + retained_proj_base[1:], False),
        ("Total Equity", total_equity_hist + [apic_proj[i] + retained_proj_base[i+1] for i in range(n_proj)], True),
    ]

    r = 3
    for label, vals, is_bold in rows_bs:
        if label == "":
            r += 1
            continue
        r = write_data_row(ws3, r, label, vals, is_bold=is_bold)

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 4: Cash Flow Statement
    # ═══════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Cash Flow")
    ws4.sheet_properties.tabColor = MED_BLUE
    ws4.column_dimensions["A"].width = 42
    for i in range(2, n_all + 2):
        ws4.column_dimensions[get_column_letter(i)].width = 16

    r = 1
    ws4.cell(row=r, column=1, value="IONQ — Cash Flow Statement (Actuals + Base Projections)").font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_c)
    r = 2
    for ci, h in enumerate(headers, start=1):
        ws4.cell(row=r, column=ci, value=h)
    style_header_row(ws4, r, max_c)

    # CF actuals
    ocf_hist_vals = hist_val("cash_flow", "NetCashProvidedByUsedInOperatingActivities")
    capex_hist_vals = [-abs(v) for v in hist_val("cash_flow", "PaymentsToAcquirePropertyPlantAndEquipment")]

    # CF projections
    nwc_change_hist = [0] + [ar_hist[i] - ar_hist[i-1] + ap_hist[i] - ap_hist[i-1] for i in range(1, n_hist)]
    nwc_change_proj = [-(ar_p[i] - ar_p[i-1] + ap_p[i] - ap_p[i-1] + accrued_p[i] - accrued_p[i-1] + defrev_p[i] - defrev_p[i-1]) if i > 0 else 0 for i in range(n_proj)]

    ocf_proj = [base_ni[i] + da_proj[i] + sbc_proj[i] + nwc_change_proj[i] for i in range(n_proj)]
    icf_proj = [-capex_proj[i] for i in range(n_proj)]
    # Financing: equity raise
    fcf_proj = [ocf_proj[i] + icf_proj[i] for i in range(n_proj)]

    # Cash build
    cash_end_proj = [cash_hist[-1]]
    for i in range(n_proj):
        cash_end_proj.append(cash_end_proj[-1] + ocf_proj[i] + icf_proj[i])
    cash_end_proj = cash_end_proj[1:]

    rows_cf = [
        ("", "", True),
        ("Operating Activities", "", True),
        ("  Net Income (Loss)", net_income_hist + base_ni, False),
        ("  Depreciation & Amortization", da_combined_hist + da_proj, False),
        ("  Stock-Based Compensation", sbc_hist + sbc_proj, False),
        ("  NWC Change", [0] * n_hist + nwc_change_proj, False),
        ("  Deferred Tax", [0] * n_hist + [base_tax[i] * 0.5 for i in range(n_proj)], False),
        ("Net Cash from Operations", ocf_hist_vals + ocf_proj, True),
        ("", "", True),
        ("Investing Activities", "", True),
        ("  Capex", capex_hist_vals + [-c for c in capex_proj], False),
        ("Net Cash from Investing", hist_val("cash_flow", "NetCashProvidedByUsedInInvestingActivities") + icf_proj, True),
        ("", "", True),
        ("Free Cash Flow (OCF - Capex)", [ocf_hist_vals[i] + capex_hist_vals[i] for i in range(n_hist)] + fcf_proj, True),
        ("", "", True),
        ("Cash & Equivalents (Ending)", cash_hist + cash_end_proj, True),
    ]

    r = 3
    for label, vals, is_bold in rows_cf:
        if label == "":
            r += 1
            continue
        r = write_data_row(ws4, r, label, vals, is_bold=is_bold)

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 5: RPO Bridge & Revenue Scenarios
    # ═══════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("RPO & Revenue Scenarios")
    ws5.sheet_properties.tabColor = DARK_BLUE
    ws5.column_dimensions["A"].width = 38
    sc_cols = ["Scenario"] + periods_proj
    for i in range(2, len(sc_cols) + 1):
        ws5.column_dimensions[get_column_letter(i)].width = 18

    r = 1
    ws5.cell(row=r, column=1, value="RPO Bridge & Revenue Scenarios").font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)
    ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(sc_cols))

    # RPO Bridge (Base)
    r = 3
    ws5.cell(row=r, column=1, value="RPO Bridge — Base Scenario").font = sub_hdr_font
    ws5.cell(row=r, column=1).fill = sub_hdr_fill
    for c in range(1, len(sc_cols) + 1):
        ws5.cell(row=r, column=c).fill = sub_hdr_fill
        ws5.cell(row=r, column=c).border = thin_border
    r = 4
    for ci, h in enumerate(sc_cols, start=1):
        ws5.cell(row=r, column=ci, value=h)
    style_header_row(ws5, r, len(sc_cols))

    rpo_open = 370_000_000
    rpo_data = []
    cum_rev = 0
    for i in range(n_proj):
        open_val = rpo_open if i == 0 else rpo_data[i-1][3]
        new_book = base_rev[i] * 0.6  # assume new bookings = 60% of current rev
        recog = base_rev[i]
        close_val = open_val + new_book - recog
        rpo_data.append((open_val, new_book, recog, close_val))
        cum_rev += recog

    rpo_labels = ["Opening RPO", "Plus: New Bookings", "Less: Revenue Recognized", "Closing RPO"]
    for label, idx in zip(rpo_labels, range(4)):
        r += 1
        vals = [d[idx] for d in rpo_data]
        ws5.cell(row=r, column=1, value=label).font = bold_font if label in ("Opening RPO", "Closing RPO") else label_font
        ws5.cell(row=r, column=1).border = thin_border
        for ci, v in enumerate(vals, start=2):
            cell = ws5.cell(row=r, column=ci, value=v)
            style_data_cell(cell, is_bold=(label in ("Opening RPO", "Closing RPO")))

    # Revenue Scenarios comparison
    r += 2
    ws5.cell(row=r, column=1, value="Revenue Scenario Comparison").font = sub_hdr_font
    ws5.cell(row=r, column=1).fill = sub_hdr_fill
    for c in range(1, len(sc_cols) + 1):
        ws5.cell(row=r, column=c).fill = sub_hdr_fill
        ws5.cell(row=r, column=c).border = thin_border
    r += 1
    for ci, h in enumerate(sc_cols, start=1):
        ws5.cell(row=r, column=ci, value=h)
    style_header_row(ws5, r, len(sc_cols))

    for label, vals in [("Bull", bull_rev), ("Base", base_rev), ("Bear", bear_rev)]:
        r += 1
        ws5.cell(row=r, column=1, value=label).font = bold_font
        ws5.cell(row=r, column=1).border = thin_border
        for ci, v in enumerate(vals, start=2):
            cell = ws5.cell(row=r, column=ci, value=v)
            style_data_cell(cell)
            # Color coding
            fill_map = {"Bull": GREEN_FILL, "Base": LIGHT_BLUE, "Bear": AMBER_FILL}
            cell.fill = PatternFill("solid", fgColor=fill_map[label])

    # CAGR
    r += 1
    ws5.cell(row=r, column=1, value="FY26→FY30 CAGR").font = bold_font
    ws5.cell(row=r, column=1).border = thin_border
    for label, vals in [("Bull", bull_rev), ("Base", base_rev), ("Bear", bear_rev)]:
        cagr = (vals[-1] / vals[0]) ** (1 / 4) - 1
        ws5.cell(row=r, column=1 + sc_cols.index("Scenario") + 1 if label == "Bull" else 2, value=None)
    r += 1
    ws5.cell(row=r, column=1, value="").font = num_font

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 6: DCF Valuation
    # ═══════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("DCF Valuation")
    ws6.sheet_properties.tabColor = DARK_BLUE
    ws6.column_dimensions["A"].width = 38
    dcf_cols = ["Line Item"] + periods_proj + ["Terminal Value"]
    for i in range(2, len(dcf_cols) + 1):
        ws6.column_dimensions[get_column_letter(i)].width = 18

    r = 1
    ws6.cell(row=r, column=1, value="DCF Valuation — Base Scenario").font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)
    ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(dcf_cols))

    r = 3
    for ci, h in enumerate(dcf_cols, start=1):
        ws6.cell(row=r, column=ci, value=h)
    style_header_row(ws6, r, len(dcf_cols))

    wacc_center = 0.13
    terminal_ev_sales = 8.0

    # FCF projection
    fcf_seq = fcf_proj
    # Terminal value: EV/Sales method
    term_rev = base_rev[-1]
    term_ev = term_rev * terminal_ev_sales
    # Discount factors
    def discount_factor(rate, year):
        return 1 / ((1 + rate) ** year)

    pv_fcfs = [fcf_seq[i] * discount_factor(wacc_center, i + 1) for i in range(n_proj)]
    pv_terminal = term_ev * discount_factor(wacc_center, n_proj)

    # Enterprise value
    ev = sum(pv_fcfs) + pv_terminal
    cash_equivalents = cash_hist[-1] + st_inv_hist[-1]
    debt = 0  # IONQ has minimal debt
    equity_value = ev + cash_equivalents - debt
    shares_base = 373_269_948
    shares_proj = [int(shares_base * (1.04 ** (i + 1))) for i in range(n_proj)]
    # Use terminal year shares for per-share calc
    shares_terminal = shares_proj[-1]
    equity_per_share = equity_value / shares_terminal

    rows_dcf = [
        ("FCF Projection", "", True),
    ]
    for i in range(n_proj):
        rows_dcf.append((f"  FCF Year {i+1} ({periods_proj[i]})", fcf_seq[i], False))
    rows_dcf += [
        ("", "", True),
        ("Discount Factor @ WACC", "", True),
    ]
    for i in range(n_proj):
        rows_dcf.append((f"  Discount Factor Y{i+1}", discount_factor(wacc_center, i + 1), False))
    rows_dcf += [
        ("", "", True),
        ("PV of FCFs", "", True),
    ]
    for i in range(n_proj):
        rows_dcf.append((f"  PV FCF Year {i+1}", pv_fcfs[i], False))
    rows_dcf += [
        ("", "", True),
        ("Terminal Value Calculation", "", True),
        ("  Terminal Revenue (FY2030)", term_rev, False),
        ("  EV/Sales Multiple", terminal_ev_sales, False),
        ("  Terminal Enterprise Value", term_ev, False),
        ("", "", True),
        ("PV of Terminal Value", pv_terminal, True),
        ("", "", True),
        ("Enterprise Value (Sum PV)", ev, True),
        ("Plus: Cash & ST Investments", cash_equivalents, False),
        ("Less: Total Debt", debt, False),
        ("Equity Value", equity_value, True),
        ("", "", True),
        ("Shares Outstanding (Terminal)", shares_terminal, False),
        ("Equity Value Per Share ($)", round(equity_per_share, 2), True),
        ("", "", True),
        ("⚠️  Terminal Value % of Total EV", pv_terminal / ev if ev else 0, True),
    ]

    r = 4
    for label, vals, is_bold in rows_dcf:
        if label == "":
            r += 1
            continue
        vals_list = vals if isinstance(vals, list) else [vals] + [None] * (len(dcf_cols) - 2)
        is_pct = "Discount Factor" in label or "Terminal Value %" in label
        ws6.cell(row=r, column=1, value=label).font = bold_font if is_bold else label_font
        ws6.cell(row=r, column=1).border = thin_border
        if isinstance(vals, list):
            for ci, v in enumerate(vals, start=2):
                cell = ws6.cell(row=r, column=ci, value=v)
                style_data_cell(cell, is_bold=is_bold)
        else:
            cell = ws6.cell(row=r, column=2, value=vals)
            style_data_cell(cell, is_bold=is_bold, is_pct=is_pct)
            if label == "Enterprise Value (Sum PV)":
                for c in range(2, len(dcf_cols) + 1):
                    ws6.cell(row=r, column=c).fill = PatternFill("solid", fgColor=GREEN_FILL)
            if label == "Equity Value Per Share ($)":
                for c in range(2, len(dcf_cols) + 1):
                    ws6.cell(row=r, column=c).fill = PatternFill("solid", fgColor=GREEN_FILL)
                    ws6.cell(row=r, column=c).number_format = '#,##0.00'
        r += 1

    # TV % of EV flag
    r += 1
    tv_pct = pv_terminal / ev if ev else 0
    if tv_pct > 0.8:
        ws6.cell(row=r, column=1, value=f"⚠️  Terminal Value = {tv_pct:.1%} of EV (>80%). Model is highly dependent on terminal assumptions.").font = Font(name="Calibri", bold=True, size=10, color="C62828")
        ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(dcf_cols))

    # Sensitivity table
    r += 2
    ws6.cell(row=r, column=1, value="Sensitivity: Equity Value Per Share ($)").font = sub_hdr_font
    ws6.cell(row=r, column=1).fill = sub_hdr_fill
    for c in range(1, 8):
        ws6.cell(row=r, column=c).fill = sub_hdr_fill
        ws6.cell(row=r, column=c).border = thin_border

    wacc_rates = [0.10, 0.11, 0.12, 0.13, 0.14, 0.15, 0.16]
    ev_sales_multiples = [4, 5, 6, 7, 8, 9, 10]

    r += 1
    ws6.cell(row=r, column=1, value="WACC \\ EV/Sales")
    ws6.cell(row=r, column=1).font = bold_font
    ws6.cell(row=r, column=1).border = thin_border
    for ci, m in enumerate(ev_sales_multiples, start=2):
        ws6.cell(row=r, column=ci, value=f"{m}x").font = bold_font
        ws6.cell(row=r, column=ci).border = thin_border
        ws6.cell(row=r, column=ci).alignment = Alignment(horizontal="center")

    sens_data = {}
    for wr in wacc_rates:
        r += 1
        ws6.cell(row=r, column=1, value=f"{wr:.0%}").font = bold_font
        ws6.cell(row=r, column=1).border = thin_border
        for ci, m in enumerate(ev_sales_multiples, start=2):
            term_ev_s = base_rev[-1] * m
            pv_f = sum([fcf_seq[i] * discount_factor(wr, i + 1) for i in range(n_proj)])
            pv_t = term_ev_s * discount_factor(wr, n_proj)
            ev_s = pv_f + pv_t
            eq_val = ev_s + cash_equivalents - debt
            eps = eq_val / shares_terminal
            cell = ws6.cell(row=r, column=ci, value=round(eps, 2))
            style_data_cell(cell)
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="center")
            # Color code
            if eps > 80:
                cell.fill = PatternFill("solid", fgColor=GREEN_FILL)
            elif eps > 50:
                cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
            elif eps > 30:
                cell.fill = PatternFill("solid", fgColor=AMBER_FILL)
            else:
                cell.fill = PatternFill("solid", fgColor=RED_FILL)

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 7: Reverse DCF
    # ═══════════════════════════════════════════════════════════════════════
    ws7 = wb.create_sheet("Reverse DCF")
    ws7.sheet_properties.tabColor = DARK_BLUE
    ws7.column_dimensions["A"].width = 45
    ws7.column_dimensions["B"].width = 22
    ws7.column_dimensions["C"].width = 22

    r = 1
    ws7.cell(row=r, column=1, value="Reverse DCF — What Does the Market Assume?").font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)
    ws7.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

    r = 3
    ws7.cell(row=r, column=1, value="Current Market Data").font = sub_hdr_font
    ws7.cell(row=r, column=1).fill = sub_hdr_fill
    ws7.cell(row=r, column=2).fill = sub_hdr_fill
    ws7.cell(row=r, column=3).fill = sub_hdr_fill
    r += 1
    for label, val in [
        ("Stock Price", "$56.89"),
        ("Market Cap", "$21.24B"),
        ("Shares Outstanding", "373M"),
        ("Cash + ST Investments", "$2.4B"),
        ("Enterprise Value (approx)", "~$18.8B"),
    ]:
        ws7.cell(row=r, column=1, value=label).font = label_font
        ws7.cell(row=r, column=2, value=val).font = num_font
        ws7.cell(row=r, column=1).border = thin_border
        ws7.cell(row=r, column=2).border = thin_border
        r += 1

    r += 1
    ws7.cell(row=r, column=1, value="Implied 2030E Scenario — WACC 13%, EV/Sales 8x").font = sub_hdr_font
    ws7.cell(row=r, column=1).fill = sub_hdr_fill
    for c in range(1, 4):
        ws7.cell(row=r, column=c).fill = sub_hdr_fill
        ws7.cell(row=r, column=c).border = thin_border

    # Reverse calc: what 2030 revenue does the market imply?
    # EV = ~$18.8B; this implies a certain 2030E revenue × EV/Sales
    target_ev = 18_800_000_000
    r += 1
    ws7.cell(row=r, column=1, value="At terminal EV/Sales = 8x:").font = bold_font
    ws7.cell(row=r, column=1).border = thin_border
    implied_2030_rev_8x = target_ev / 8
    r += 1
    ws7.cell(row=r, column=1, value="  Implied 2030 Revenue").font = label_font
    ws7.cell(row=r, column=2, value=implied_2030_rev_8x).font = bold_font
    ws7.cell(row=r, column=2).number_format = NUM_FMT
    ws7.cell(row=r, column=1).border = thin_border
    ws7.cell(row=r, column=2).border = thin_border
    implied_cagr = (implied_2030_rev_8x / rev_hist[-1]) ** (1 / 5) - 1
    r += 1
    ws7.cell(row=r, column=1, value="  Implied FY25→FY30 CAGR").font = label_font
    ws7.cell(row=r, column=2, value=implied_cagr).font = bold_font
    ws7.cell(row=r, column=2).number_format = PCT_FMT
    ws7.cell(row=r, column=1).border = thin_border
    ws7.cell(row=r, column=2).border = thin_border

    r += 1
    ws7.cell(row=r, column=1, value="At terminal EV/Sales = 6x:").font = bold_font
    ws7.cell(row=r, column=1).border = thin_border
    implied_2030_rev_6x = target_ev / 6
    r += 1
    ws7.cell(row=r, column=1, value="  Implied 2030 Revenue").font = label_font
    ws7.cell(row=r, column=2, value=implied_2030_rev_6x).font = bold_font
    ws7.cell(row=r, column=2).number_format = NUM_FMT
    ws7.cell(row=r, column=1).border = thin_border
    ws7.cell(row=r, column=2).border = thin_border
    cagr6 = (implied_2030_rev_6x / rev_hist[-1]) ** (1 / 5) - 1
    r += 1
    ws7.cell(row=r, column=1, value="  Implied FY25→FY30 CAGR").font = label_font
    ws7.cell(row=r, column=2, value=cagr6).font = bold_font
    ws7.cell(row=r, column=2).number_format = PCT_FMT
    ws7.cell(row=r, column=1).border = thin_border
    ws7.cell(row=r, column=2).border = thin_border

    r += 1
    ws7.cell(row=r, column=1, value="At terminal EV/Sales = 10x:").font = bold_font
    ws7.cell(row=r, column=1).border = thin_border
    implied_2030_rev_10x = target_ev / 10
    r += 1
    ws7.cell(row=r, column=1, value="  Implied 2030 Revenue").font = label_font
    ws7.cell(row=r, column=2, value=implied_2030_rev_10x).font = bold_font
    ws7.cell(row=r, column=2).number_format = NUM_FMT
    ws7.cell(row=r, column=1).border = thin_border
    ws7.cell(row=r, column=2).border = thin_border
    cagr10 = (implied_2030_rev_10x / rev_hist[-1]) ** (1 / 5) - 1
    r += 1
    ws7.cell(row=r, column=1, value="  Implied FY25→FY30 CAGR").font = label_font
    ws7.cell(row=r, column=2, value=cagr10).font = bold_font
    ws7.cell(row=r, column=2).number_format = PCT_FMT
    ws7.cell(row=r, column=1).border = thin_border
    ws7.cell(row=r, column=2).border = thin_border

    r += 1
    ws7.cell(row=r, column=1, value="Implication").font = sub_hdr_font
    ws7.cell(row=r, column=1).fill = sub_hdr_fill
    for c in range(1, 4):
        ws7.cell(row=r, column=c).fill = sub_hdr_fill
        ws7.cell(row=r, column=c).border = thin_border
    r += 1
    implication_text = (
        f"At 8x terminal EV/Sales, the market implies ~${implied_2030_rev_8x/1e9:.1f}B 2030 revenue "
        f"({implied_cagr*100:.0f}% CAGR FY25→FY30). This is plausible but aggressive: IONQ's Base scenario "
        f"($700M FY2030) would only justify ~${(700_000_000 * 8 + cash_equivalents) / 1e9:.1f}B EV. "
        f"The current market cap implies ~${implied_2030_rev_8x/1e9:.1f}B FY2030 revenue. "
        f"This is {implied_2030_rev_8x/700_000_000:.0f}x the Base scenario projection."
    )
    ws7.cell(row=r, column=1, value=implication_text).font = num_font
    ws7.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws7.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws7.row_dimensions[r].height = 80

    # Implied 2030 Revenue × EV/Sales matrix
    r += 2
    ws7.cell(row=r, column=1, value="Implied Equity Value per Share ($) Matrix").font = sub_hdr_font
    ws7.cell(row=r, column=1).fill = sub_hdr_fill
    for c in range(1, 8):
        ws7.cell(row=r, column=c).fill = sub_hdr_fill
        ws7.cell(row=r, column=c).border = thin_border

    rev_2030_values = [300_000_000, 500_000_000, 700_000_000, 1_000_000_000, 1_500_000_000, 2_500_000_000]
    ev_sales_list = [4, 6, 8, 10, 12, 15]

    r += 1
    ws7.cell(row=r, column=1, value="FY2030 Rev \\ EV/Sales")
    ws7.cell(row=r, column=1).font = bold_font
    ws7.cell(row=r, column=1).border = thin_border
    for ci, m in enumerate(ev_sales_list, start=2):
        ws7.cell(row=r, column=ci, value=f"{m}x").font = bold_font
        ws7.cell(row=r, column=ci).border = thin_border
        ws7.cell(row=r, column=ci).alignment = Alignment(horizontal="center")

    for rev in rev_2030_values:
        r += 1
        ws7.cell(row=r, column=1, value=f"${rev/1e9:.1f}B").font = bold_font
        ws7.cell(row=r, column=1).border = thin_border
        for ci, m in enumerate(ev_sales_list, start=2):
            tev = rev * m
            eq = tev + cash_equivalents
            eps = eq / shares_terminal
            cell = ws7.cell(row=r, column=ci, value=round(eps, 2))
            style_data_cell(cell)
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="center")
            # Color: green if above current price
            if eps > 56.89 * 1.1:
                cell.fill = PatternFill("solid", fgColor=GREEN_FILL)
            elif eps > 56.89:
                cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
            elif eps > 20:
                cell.fill = PatternFill("solid", fgColor=AMBER_FILL)
            else:
                cell.fill = PatternFill("solid", fgColor=RED_FILL)
        # Add current price marker comment
        for ci in range(2, len(ev_sales_list) + 2):
            pass

    # Highlight current price zone
    r += 2
    ws7.cell(row=r, column=1, value="Note: Cells in green exceed current price ($56.89); blue = near parity; amber/red = below. This shows the revenue × multiple combinations needed to justify today's valuation.").font = Font(name="Calibri", size=8, italic=True, color="808080")
    ws7.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)

    # ═══════════════════════════════════════════════════════════════════════
    # Save
    # ═══════════════════════════════════════════════════════════════════════
    wb.save(str(OUTPUT))
    print(f"[OK] DCF model saved to: {OUTPUT}")
    print(f"   Sheets: {[ws.title for ws in wb.worksheets]}")
    print(f"   Key output: Equity Value ~${equity_value/1e9:.2f}B | Per Share ~${equity_per_share:.2f} | TV% of EV: {tv_pct:.1%}")


if __name__ == "__main__":
    build()
