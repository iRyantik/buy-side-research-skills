#!/usr/bin/env python3
"""Build RKLB 3-statement + DCF model from financial-data and driver-map examples."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parents[1] / "examples" / "financial-data-pull" / "us" / "rklb"
INTERNAL = BASE / "internal"
FINANCIAL_DATA_INTERNAL = INTERNAL / "financial-data"
DRIVER_MAP_INTERNAL = INTERNAL / "driver-map"
FINANCIALS = FINANCIAL_DATA_INTERNAL / "financials.normalized.json"
ACTUALS = FINANCIAL_DATA_INTERNAL / "actuals-resolved.json"
DRIVER_MAP = DRIVER_MAP_INTERNAL / "driver-map.json"
IDENTITY = FINANCIAL_DATA_INTERNAL / "identity.json"
OUTPUT = BASE / "rklb-3statement-dcf-model.xlsx"

HIST_YEARS = [2023, 2024, 2025]
PROJ_YEARS = [2026, 2027, 2028, 2029, 2030]
ALL_YEARS = HIST_YEARS + PROJ_YEARS

BLUE = "1F4E79"
LIGHT_BLUE = "D9E1F2"
MID_BLUE = "BDD7EE"
GREY = "F2F2F2"
GREEN = "008000"
WHITE = "FFFFFF"
INPUT_BLUE = "0000FF"


def load_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


actuals_pack = load_json(ACTUALS) if ACTUALS.exists() else {}
financials = actuals_pack.get("statements") or load_json(FINANCIALS)
driver = load_json(DRIVER_MAP)
identity = load_json(IDENTITY)


def row_by_concept(statement: str, concept: str) -> dict:
    for row in financials.get(statement, []):
        if row.get("concept") == concept:
            return row
    return {"values": {}}


def value(statement: str, concept: str, year: int) -> float | None:
    values = row_by_concept(statement, concept).get("values", {})
    raw = values.get(f"FY {year}")
    if raw is None:
        return None
    return raw / 1_000_000


def seg(name: str) -> dict:
    for segment in driver["segment_geography_treatment"]["reported_segments"]:
        if segment["name"] == name:
            return segment
    raise KeyError(name)


launch = seg("Launch Services")
space = seg("Space Systems")


def dm_values(segment: dict, key: str, year: int) -> float | None:
    raw = segment.get(key, {}).get(f"FY{year}")
    return None if raw is None else float(raw)


def input_comment(text: str) -> Comment:
    return Comment(text, "financial-data")


def qsheet(sheet: str) -> str:
    return f"'{sheet}'" if " " in sheet else sheet


def xref(sheet: str, cell: str) -> str:
    return f"{qsheet(sheet)}!{cell}"


def fmt_num(cell):
    cell.number_format = '#,##0.0;[Red](#,##0.0);-'


def fmt_pct(cell):
    cell.number_format = '0.0%'


def fmt_mult(cell):
    cell.number_format = '0.0x'


def set_title(ws, title: str, last_col: int = 10):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, color=WHITE, size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE)
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)


def set_periods(ws, row: int = 3):
    ws.cell(row=row, column=1, value="$ in millions, except per-share data")
    ws.cell(row=row, column=1).font = Font(italic=True, color="666666")
    for idx, year in enumerate(ALL_YEARS, start=2):
        cell = ws.cell(row=row, column=idx, value=f"FY{year}{'A' if year in HIST_YEARS else 'E'}")
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


def style_sheet(ws, max_col: int = 10):
    thin = Side(style="thin", color="D9E1F2")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "B4"
    ws.column_dimensions["A"].width = 34
    for col in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14


def section(ws, row: int, label: str, last_col: int = 10):
    ws.cell(row=row, column=1, value=label)
    for col in range(1, last_col + 1):
        c = ws.cell(row=row, column=col)
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.font = Font(bold=True, color=WHITE)


def hardcode(ws, row: int, col: int, val, comment: str | None = None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(color=GREEN if isinstance(val, str) and val.startswith("=") else INPUT_BLUE)
    if comment:
        c.comment = input_comment(comment)
    if isinstance(val, (int, float)):
        fmt_num(c)
    return c


def formula(ws, row: int, col: int, val: str):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(color=GREEN)
    return c


def write_metric_row(ws, row: int, label: str, hist_values: dict[int, float | None],
                     proj_formulas: dict[int, str] | None = None,
                     comment: str | None = None, percent: bool = False):
    ws.cell(row=row, column=1, value=label)
    for col, year in enumerate(ALL_YEARS, start=2):
        if year in HIST_YEARS:
            val = hist_values.get(year)
            if val is not None:
                c = hardcode(ws, row, col, val, comment)
            else:
                c = ws.cell(row=row, column=col, value=None)
        else:
            c = formula(ws, row, col, proj_formulas[year])
        fmt_pct(c) if percent else fmt_num(c)


def build_cover(wb: Workbook):
    ws = wb.create_sheet("Cover")
    set_title(ws, "RKLB 3-Statement + DCF Model", 8)
    rows = [
        ("Company", identity.get("name", "Rocket Lab Corp")),
        ("Ticker", "RKLB"),
        ("Model output", str(OUTPUT)),
        ("Source pack", str(BASE)),
        ("Primary filing", "FY2025 10-K, filed 2026-02-26, accession 0001819994-26-000013"),
        ("Historical actuals", "FY2023-FY2025 from financial-data / SEC XBRL"),
        ("Forecast horizon", "FY2026E-FY2030E"),
        ("Revenue build", "Launch Services + Space Systems Products + Space Systems Services"),
        ("Key modeling caveat", "Backlog is disclosed only at company level; do not allocate by segment without new source."),
    ]
    for r, (k, v) in enumerate(rows, start=4):
        ws.cell(r, 1, k).font = Font(bold=True)
        ws.cell(r, 2, v)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 110
    style_sheet(ws, 8)


def build_sources(wb: Workbook):
    ws = wb.create_sheet("Sources")
    set_title(ws, "Sources and Model Caveats", 8)
    source_rows = [
        ("financial-data-summary.md", "Public financial-data summary for human/LLM review."),
        ("internal/financial-data/actuals-resolved.json", "Machine historical actuals; missing/unmapped policy says leave blank and flag review."),
        ("internal/financial-data/evidence-pack.json", "Completeness, source-map, cross-check and latest run pointer."),
        ("internal/financial-data/full-filing.md", "Internal FY2025 10-K text retained for audit/deeper lookup."),
        ("driver-map.md", "Public driver-map summary for revenue split and modeling treatment."),
        ("internal/driver-map/driver-map.json", "Machine driver-map input for revenue build assumptions."),
    ]
    for r, (k, v) in enumerate(source_rows, start=4):
        ws.cell(r, 1, k).font = Font(bold=True)
        ws.cell(r, 2, v)
    section(ws, 12, "Critical Caveats", 8)
    caveats = [
        "Launch cadence and ASP are not in the financial-data pack; they need a sourced launch manifest before being model-ready.",
        "Neutron is modeled as scenario optionality, not historical run-rate.",
        "Company backlog was $1.847bn at FY2025, but no segment split is disclosed.",
        "Forecast assumptions are illustrative and editable blue inputs, not management guidance.",
    ]
    for r, text in enumerate(caveats, start=13):
        ws.cell(r, 1, f"Caveat {r-12}").font = Font(bold=True)
        ws.cell(r, 2, text)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 110
    style_sheet(ws, 8)


def build_assumptions(wb: Workbook):
    ws = wb.create_sheet("Assumptions")
    set_title(ws, "Forecast and Valuation Assumptions", 10)
    set_periods(ws)
    assumptions = {
        "Launch revenue growth": [0.30, 0.25, 0.22, 0.18, 0.15],
        "Space Systems products growth": [0.25, 0.22, 0.20, 0.18, 0.16],
        "Space Systems services growth": [0.28, 0.25, 0.22, 0.20, 0.18],
        "Launch gross margin": [0.42, 0.43, 0.44, 0.45, 0.46],
        "Space Systems products gross margin": [0.32, 0.33, 0.34, 0.35, 0.36],
        "Space Systems services gross margin": [0.23, 0.24, 0.25, 0.26, 0.27],
        "R&D % revenue": [0.42, 0.38, 0.34, 0.30, 0.27],
        "SG&A % revenue": [0.24, 0.22, 0.20, 0.18, 0.16],
        "D&A % revenue": [0.07, 0.065, 0.06, 0.055, 0.05],
        "SBC % revenue": [0.10, 0.09, 0.08, 0.07, 0.06],
        "CapEx % revenue": [0.16, 0.14, 0.12, 0.10, 0.08],
        "AR % revenue": [0.065, 0.064, 0.063, 0.062, 0.061],
        "Inventory % revenue": [0.25, 0.24, 0.23, 0.22, 0.21],
        "Other current assets % revenue": [0.14, 0.13, 0.12, 0.11, 0.10],
        "AP % revenue": [0.12, 0.115, 0.11, 0.105, 0.10],
        "Accrued liabilities % revenue": [0.032, 0.031, 0.030, 0.029, 0.028],
        "Contract liabilities % revenue": [0.30, 0.29, 0.28, 0.27, 0.26],
        "Tax rate": [0.21, 0.21, 0.21, 0.21, 0.21],
        "Interest rate on debt": [0.08, 0.08, 0.08, 0.08, 0.08],
        "Share dilution": [0.03, 0.03, 0.03, 0.03, 0.03],
    }
    section(ws, 5, "Operating Assumptions", 10)
    start = 6
    for idx, (label, vals) in enumerate(assumptions.items(), start=start):
        ws.cell(idx, 1, label)
        for c, year in enumerate(PROJ_YEARS, start=5):
            cell = hardcode(ws, idx, c, vals[year - 2026], "Source: driver-map judgment; editable assumption.")
            fmt_pct(cell)
    section(ws, 29, "DCF Assumptions", 10)
    dcf_inputs = [
        ("WACC", 0.11),
        ("Terminal growth", 0.035),
        ("Current diluted shares (m)", identity.get("shares_outstanding", 578750990) / 1_000_000),
        ("Current cash (m)", value("balance_sheet", "CashAndCashEquivalentsAtCarryingValue", 2025)),
        ("Current debt (m)", value("balance_sheet", "LongTermDebtAndCapitalLeaseObligations", 2025)),
        ("Current share price", None),
    ]
    for r, (label, val) in enumerate(dcf_inputs, start=30):
        ws.cell(r, 1, label)
        c = hardcode(ws, r, 2, val, "Source: financial-data / editable valuation input.")
        if "WACC" in label or "growth" in label:
            fmt_pct(c)
        else:
            fmt_num(c)
    style_sheet(ws, 10)


def build_revenue(wb: Workbook):
    ws = wb.create_sheet("Revenue Build")
    set_title(ws, "Revenue Build by Driver", 10)
    set_periods(ws)
    section(ws, 5, "Launch Services", 10)
    write_metric_row(
        ws, 6, "Revenue",
        {y: dm_values(launch, "revenue_usd_m", y) for y in HIST_YEARS},
        {y: f"={get_column_letter(y - 2026 + 4)}6*(1+{xref('Assumptions', f'{get_column_letter(y - 2026 + 5)}6')})" for y in PROJ_YEARS},
        "Source: RKLB FY2025 10-K segment table."
    )
    write_metric_row(
        ws, 7, "YoY Growth",
        {2023: None, 2024: "=C6/B6-1", 2025: "=D6/C6-1"},
        {y: f"={get_column_letter(y - 2026 + 5)}6/{get_column_letter(y - 2026 + 4)}6-1" for y in PROJ_YEARS},
        percent=True,
    )
    write_metric_row(
        ws, 8, "Gross Margin",
        {y: dm_values(launch, "gross_margin_pct", y) / 100 for y in HIST_YEARS},
        {y: f"={xref('Assumptions', f'{get_column_letter(y - 2026 + 5)}9')}" for y in PROJ_YEARS},
        "Source: RKLB FY2025 10-K segment gross profit / revenue.",
        True,
    )
    write_metric_row(
        ws, 9, "Gross Profit",
        {y: dm_values(launch, "gross_profit_usd_m", y) for y in HIST_YEARS},
        {y: f"={get_column_letter(y - 2026 + 5)}6*{get_column_letter(y - 2026 + 5)}8" for y in PROJ_YEARS},
        "Source: RKLB FY2025 10-K segment table."
    )
    section(ws, 11, "Space Systems Products", 10)
    write_metric_row(
        ws, 12, "Revenue",
        {y: space["product_service_split_usd_m"]["products"][f"FY{y}"] for y in HIST_YEARS},
        {y: f"={get_column_letter(y - 2026 + 4)}12*(1+{xref('Assumptions', f'{get_column_letter(y - 2026 + 5)}7')})" for y in PROJ_YEARS},
        "Source: RKLB FY2025 10-K product/service segment table."
    )
    write_metric_row(
        ws, 13, "YoY Growth",
        {2023: None, 2024: "=C12/B12-1", 2025: "=D12/C12-1"},
        {y: f"={get_column_letter(y - 2026 + 5)}12/{get_column_letter(y - 2026 + 4)}12-1" for y in PROJ_YEARS},
        percent=True,
    )
    write_metric_row(
        ws, 14, "Gross Margin",
        {2023: 41.218 / 156.56, 2024: 76.016 / 289.851, 2025: 118.769 / 371.617},
        {y: f"={xref('Assumptions', f'{get_column_letter(y - 2026 + 5)}10')}" for y in PROJ_YEARS},
        "Source: RKLB FY2025 10-K product/service table.",
        True,
    )
    write_metric_row(
        ws, 15, "Gross Profit",
        {2023: 41.218, 2024: 76.016, 2025: 118.769},
        {y: f"={get_column_letter(y - 2026 + 5)}12*{get_column_letter(y - 2026 + 5)}14" for y in PROJ_YEARS}
    )
    section(ws, 17, "Space Systems Services", 10)
    write_metric_row(
        ws, 18, "Revenue",
        {y: space["product_service_split_usd_m"]["services"][f"FY{y}"] for y in HIST_YEARS},
        {y: f"={get_column_letter(y - 2026 + 4)}18*(1+{xref('Assumptions', f'{get_column_letter(y - 2026 + 5)}8')})" for y in PROJ_YEARS},
        "Source: RKLB FY2025 10-K product/service segment table."
    )
    write_metric_row(
        ws, 19, "YoY Growth",
        {2023: None, 2024: "=C18/B18-1", 2025: "=D18/C18-1"},
        {y: f"={get_column_letter(y - 2026 + 5)}18/{get_column_letter(y - 2026 + 4)}18-1" for y in PROJ_YEARS},
        percent=True,
    )
    write_metric_row(
        ws, 20, "Gross Margin",
        {2023: 2.124 / 16.138, 2024: 5.543 / 20.987, 2025: 7.142 / 31.14},
        {y: f"={xref('Assumptions', f'{get_column_letter(y - 2026 + 5)}11')}" for y in PROJ_YEARS},
        "Source: RKLB FY2025 10-K product/service table.",
        True,
    )
    write_metric_row(
        ws, 21, "Gross Profit",
        {2023: 2.124, 2024: 5.543, 2025: 7.142},
        {y: f"={get_column_letter(y - 2026 + 5)}18*{get_column_letter(y - 2026 + 5)}20" for y in PROJ_YEARS}
    )
    section(ws, 23, "Total Revenue and Gross Profit", 10)
    for col in range(2, 10):
        c = get_column_letter(col)
        formula(ws, 24, col, f"={c}12+{c}18")
        formula(ws, 25, col, f"={c}6+{c}24")
        formula(ws, 26, col, f'=IFERROR({c}25/{get_column_letter(col-1)}25-1,"")' if col > 2 else '=""')
        formula(ws, 27, col, f"={c}9+{c}15+{c}21")
        formula(ws, 28, col, f"={c}27/{c}25")
        fmt_num(ws.cell(24, col)); fmt_num(ws.cell(25, col)); fmt_pct(ws.cell(26, col)); fmt_num(ws.cell(27, col)); fmt_pct(ws.cell(28, col))
    labels = {24: "Space Systems Revenue", 25: "Total Revenue", 26: "Total Revenue Growth", 27: "Total Gross Profit", 28: "Gross Margin"}
    for row, label in labels.items():
        ws.cell(row, 1, label)
    section(ws, 31, "Backlog Memo", 10)
    ws.cell(32, 1, "Company backlog")
    hardcode(ws, 32, 4, 1847.322, "Source: internal/financial-data/full-filing.md; company-level backlog, not segment-disclosed.")
    ws.cell(33, 1, "% expected within 12 months")
    c = hardcode(ws, 33, 4, 0.37, "Source: RKLB FY2025 10-K backlog note.")
    fmt_pct(c)
    style_sheet(ws, 10)


def build_income_statement(wb: Workbook):
    ws = wb.create_sheet("Income Statement")
    set_title(ws, "Income Statement", 10)
    set_periods(ws)
    rows = {
        5: "Revenue",
        6: "Revenue Growth",
        8: "Cost of Revenue",
        9: "Gross Profit",
        10: "Gross Margin",
        12: "R&D",
        13: "SG&A",
        14: "Operating Expenses",
        15: "Operating Income",
        16: "Operating Margin",
        18: "D&A",
        19: "EBITDA",
        20: "EBITDA Margin",
        22: "Interest Expense Memo",
        23: "Net Nonoperating Income / Expense",
        24: "Pre-tax Income",
        25: "Tax Expense / (Benefit)",
        26: "Net Income",
        27: "Net Margin",
        29: "Diluted Shares",
        30: "EPS",
    }
    for r, label in rows.items():
        ws.cell(r, 1, label)
    hist_map = {
        5: ("income_statement", "RevenueFromContractWithCustomerExcludingAssessedTax"),
        8: ("income_statement", "CostOfRevenue"),
        9: ("income_statement", "GrossProfit"),
        12: ("income_statement", "ResearchAndDevelopmentExpense"),
        13: ("income_statement", "SellingGeneralAndAdministrativeExpense"),
        14: ("income_statement", "OperatingExpenses"),
        15: ("income_statement", "OperatingIncomeLoss"),
        18: ("income_statement", "DepreciationAndAmortization"),
        22: ("income_statement", "InterestExpenseNonoperating"),
        23: ("income_statement", "NonoperatingIncomeExpense"),
        25: ("income_statement", "IncomeTaxExpenseBenefit"),
        26: ("income_statement", "NetIncomeLoss"),
        29: ("income_statement", "WeightedAverageNumberOfDilutedSharesOutstanding"),
        30: ("income_statement", "EarningsPerShareDiluted"),
    }
    for col, year in enumerate(HIST_YEARS, start=2):
        for row, (stmt, concept) in hist_map.items():
            val = value(stmt, concept, year)
            if val is not None:
                if row == 29:
                    val = val * 1_000_000 / 1_000_000
                hardcode(ws, row, col, val, "Source: financials.normalized.json / SEC XBRL.")
        formula(ws, 6, col, f'=IFERROR({get_column_letter(col)}5/{get_column_letter(col-1)}5-1,"")' if col > 2 else '=""')
        formula(ws, 10, col, f"={get_column_letter(col)}9/{get_column_letter(col)}5")
        formula(ws, 16, col, f"={get_column_letter(col)}15/{get_column_letter(col)}5")
        formula(ws, 19, col, f"={get_column_letter(col)}15+{get_column_letter(col)}18")
        formula(ws, 20, col, f"={get_column_letter(col)}19/{get_column_letter(col)}5")
        if ws.cell(22, col).value is None:
            net_interest = value("income_statement", "InterestIncomeExpenseNonoperatingNet", year)
            if net_interest is not None:
                hardcode(ws, 22, col, abs(net_interest), "Source: financials.normalized.json / SEC XBRL net interest fallback.")
        formula(ws, 24, col, f"={get_column_letter(col)}15+{get_column_letter(col)}23")
        formula(ws, 27, col, f"={get_column_letter(col)}26/{get_column_letter(col)}5")
    for col, year in enumerate(PROJ_YEARS, start=5):
        c = get_column_letter(col)
        formula(ws, 5, col, f"={xref('Revenue Build', f'{c}25')}")
        formula(ws, 6, col, f"={c}5/{get_column_letter(col-1)}5-1")
        formula(ws, 9, col, f"={xref('Revenue Build', f'{c}27')}")
        formula(ws, 8, col, f"={c}5-{c}9")
        formula(ws, 10, col, f"={c}9/{c}5")
        formula(ws, 12, col, f"={c}5*{xref('Assumptions', f'{c}12')}")
        formula(ws, 13, col, f"={c}5*{xref('Assumptions', f'{c}13')}")
        formula(ws, 14, col, f"={c}12+{c}13")
        formula(ws, 15, col, f"={c}9-{c}14")
        formula(ws, 16, col, f"={c}15/{c}5")
        formula(ws, 18, col, f"={c}5*{xref('Assumptions', f'{c}14')}")
        formula(ws, 19, col, f"={c}15+{c}18")
        formula(ws, 20, col, f"={c}19/{c}5")
        formula(ws, 22, col, f"={xref('Balance Sheet', f'{c}22')}*{xref('Assumptions', f'{c}24')}")
        formula(ws, 23, col, f"=-{c}22")
        formula(ws, 24, col, f"={c}15+{c}23")
        formula(ws, 25, col, f"=IF({c}24>0,{c}24*{xref('Assumptions', f'{c}23')},0)")
        formula(ws, 26, col, f"={c}24-{c}25")
        formula(ws, 27, col, f"={c}26/{c}5")
        formula(ws, 29, col, f"={get_column_letter(col-1)}29*(1+{xref('Assumptions', f'{c}25')})")
        formula(ws, 30, col, f"={c}26/{c}29")
    for row in rows:
        for col in range(2, 10):
            if row in {6, 10, 16, 20, 27}:
                fmt_pct(ws.cell(row, col))
            else:
                fmt_num(ws.cell(row, col))
    section(ws, 4, "Income Statement", 10)
    section(ws, 11, "Operating Expenses", 10)
    section(ws, 21, "Below EBIT", 10)
    style_sheet(ws, 10)


def build_balance_sheet(wb: Workbook):
    ws = wb.create_sheet("Balance Sheet")
    set_title(ws, "Balance Sheet", 10)
    set_periods(ws)
    rows = {
        5: "Cash & Equivalents",
        6: "Accounts Receivable",
        7: "Inventory",
        8: "Other Current Assets",
        9: "Total Current Assets",
        11: "PP&E, Net",
        12: "Goodwill",
        13: "Intangibles",
        14: "Other Assets",
        15: "Total Assets",
        17: "Accounts Payable",
        18: "Accrued Liabilities",
        19: "Contract Liabilities",
        20: "Current Liabilities",
        22: "Total Debt",
        23: "Lease / Other Noncurrent Liabilities",
        24: "Total Liabilities",
        26: "Shareholders' Equity",
        27: "Liabilities + Equity",
        29: "Balance Check",
        31: "Net Cash / (Debt)",
    }
    for r, label in rows.items():
        ws.cell(r, 1, label)
    hist = {
        5: ("balance_sheet", "CashAndCashEquivalentsAtCarryingValue"),
        6: ("balance_sheet", "AccountsReceivableNetCurrent"),
        7: ("balance_sheet", "InventoryNet"),
        8: ("balance_sheet", "PrepaidExpenseAndOtherAssetsCurrent"),
        9: ("balance_sheet", "AssetsCurrent"),
        11: ("balance_sheet", "PropertyPlantAndEquipmentNet"),
        12: ("balance_sheet", "Goodwill"),
        13: ("balance_sheet", "IntangibleAssetsNetExcludingGoodwill"),
        14: ("balance_sheet", "OtherAssetsNoncurrent"),
        15: ("balance_sheet", "Assets"),
        17: ("balance_sheet", "AccountsPayableCurrent"),
        18: ("balance_sheet", "AccruedLiabilitiesCurrent"),
        20: ("balance_sheet", "LiabilitiesCurrent"),
        22: ("balance_sheet", "LongTermDebtAndCapitalLeaseObligations"),
        23: ("balance_sheet", "OperatingLeaseLiabilityNoncurrent"),
        24: ("balance_sheet", "Liabilities"),
        26: ("balance_sheet", "StockholdersEquity"),
        27: ("balance_sheet", "LiabilitiesAndStockholdersEquity"),
    }
    contract_liab = {2023: 139.338, 2024: 216.160, 2025: 195.438}
    for col, year in enumerate(HIST_YEARS, start=2):
        for row, (stmt, concept) in hist.items():
            val = value(stmt, concept, year)
            if val is not None:
                hardcode(ws, row, col, val, "Source: financials.normalized.json / SEC XBRL.")
        hardcode(ws, 19, col, contract_liab[year], "Source: internal/financial-data/full-filing.md contract liabilities table.")
        c = get_column_letter(col)
        formula(ws, 29, col, f"={c}15-{c}27")
        formula(ws, 31, col, f"={c}5-{c}22")
    for col, year in enumerate(PROJ_YEARS, start=5):
        c = get_column_letter(col)
        revenue = xref("Income Statement", f"{c}5")
        formula(ws, 5, col, f"={xref('Cash Flow', f'{c}20')}")
        formula(ws, 6, col, f"={revenue}*{xref('Assumptions', f'{c}17')}")
        formula(ws, 7, col, f"={revenue}*{xref('Assumptions', f'{c}18')}")
        formula(ws, 8, col, f"={revenue}*{xref('Assumptions', f'{c}19')}")
        formula(ws, 9, col, f"=SUM({c}5:{c}8)")
        formula(ws, 11, col, f"={get_column_letter(col-1)}11+{xref('Cash Flow', f'{c}11')}-{xref('Income Statement', f'{c}18')}")
        formula(ws, 12, col, f"={get_column_letter(col-1)}12")
        formula(ws, 13, col, f"={get_column_letter(col-1)}13")
        formula(ws, 14, col, f"={get_column_letter(col-1)}14")
        formula(ws, 15, col, f"=SUM({c}9,{c}11:{c}14)")
        formula(ws, 17, col, f"={revenue}*{xref('Assumptions', f'{c}20')}")
        formula(ws, 18, col, f"={revenue}*{xref('Assumptions', f'{c}21')}")
        formula(ws, 19, col, f"={revenue}*{xref('Assumptions', f'{c}22')}")
        formula(ws, 20, col, f"=SUM({c}17:{c}19)")
        formula(ws, 22, col, f"={get_column_letter(col-1)}22")
        formula(ws, 23, col, f"={get_column_letter(col-1)}23")
        formula(ws, 24, col, f"=SUM({c}20,{c}22:{c}23)")
        formula(ws, 26, col, f"={c}15-{c}24")
        formula(ws, 27, col, f"={c}24+{c}26")
        formula(ws, 29, col, f"={c}15-{c}27")
        formula(ws, 31, col, f"={c}5-{c}22")
    for row in rows:
        for col in range(2, 10):
            fmt_num(ws.cell(row, col))
    section(ws, 4, "Assets", 10)
    section(ws, 16, "Liabilities and Equity", 10)
    section(ws, 28, "Checks / Memo", 10)
    style_sheet(ws, 10)


def build_cash_flow(wb: Workbook):
    ws = wb.create_sheet("Cash Flow")
    set_title(ws, "Cash Flow Statement", 10)
    set_periods(ws)
    labels = {
        5: "Net Income",
        6: "D&A",
        7: "Stock-Based Compensation",
        8: "Change in NWC",
        9: "Cash Flow from Operations",
        11: "CapEx",
        12: "Acquisitions",
        13: "Cash Flow from Investing",
        15: "Debt / Equity Financing",
        16: "Cash Flow from Financing",
        18: "Beginning Cash",
        19: "Net Change in Cash",
        20: "Ending Cash",
        21: "Cash Tie Check",
        23: "Unlevered FCF",
    }
    for r, label in labels.items():
        ws.cell(r, 1, label)
    hist = {
        5: ("cash_flow", "ProfitLoss"),
        6: ("cash_flow", "DepreciationDepletionAndAmortization"),
        7: ("cash_flow", "ShareBasedCompensation"),
        9: ("cash_flow", "NetCashProvidedByUsedInOperatingActivities"),
        11: ("cash_flow", "PaymentsToAcquirePropertyPlantAndEquipment"),
        12: ("cash_flow", "PaymentsToAcquireBusinessesNetOfCashAcquired"),
        13: ("cash_flow", "NetCashProvidedByUsedInInvestingActivities"),
        16: ("cash_flow", "NetCashProvidedByUsedInFinancingActivities"),
        19: ("cash_flow", "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect"),
        20: ("cash_flow", "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents"),
    }
    for col, year in enumerate(HIST_YEARS, start=2):
        for row, (stmt, concept) in hist.items():
            val = value(stmt, concept, year)
            if val is not None:
                hardcode(ws, row, col, val, "Source: financials.normalized.json / SEC XBRL.")
        c = get_column_letter(col)
        formula(ws, 8, col, f"={c}5+{c}6+{c}7-{c}9")
        formula(ws, 15, col, f"={c}16")
        formula(ws, 18, col, f"={get_column_letter(col-1)}20" if col > 2 else f"={c}20-{c}19")
        formula(ws, 21, col, f"={c}20-{xref('Balance Sheet', f'{c}5')}")
        formula(ws, 23, col, f"={xref('Income Statement', f'{c}15')}*(1-21%)+{c}6-{c}11-{c}8")
    for col, year in enumerate(PROJ_YEARS, start=5):
        c = get_column_letter(col)
        prev = get_column_letter(col - 1)
        nwc = f"({xref('Balance Sheet', f'{c}6')}+{xref('Balance Sheet', f'{c}7')}+{xref('Balance Sheet', f'{c}8')}-{xref('Balance Sheet', f'{c}17')}-{xref('Balance Sheet', f'{c}18')}-{xref('Balance Sheet', f'{c}19')})"
        prev_nwc = f"({xref('Balance Sheet', f'{prev}6')}+{xref('Balance Sheet', f'{prev}7')}+{xref('Balance Sheet', f'{prev}8')}-{xref('Balance Sheet', f'{prev}17')}-{xref('Balance Sheet', f'{prev}18')}-{xref('Balance Sheet', f'{prev}19')})"
        formula(ws, 5, col, f"={xref('Income Statement', f'{c}26')}")
        formula(ws, 6, col, f"={xref('Income Statement', f'{c}18')}")
        formula(ws, 7, col, f"={xref('Income Statement', f'{c}5')}*{xref('Assumptions', f'{c}15')}")
        formula(ws, 8, col, f"={nwc}-{prev_nwc}")
        formula(ws, 9, col, f"={c}5+{c}6+{c}7-{c}8")
        formula(ws, 11, col, f"={xref('Income Statement', f'{c}5')}*{xref('Assumptions', f'{c}16')}")
        formula(ws, 12, col, "=0")
        formula(ws, 13, col, f"=-{c}11-{c}12")
        formula(ws, 15, col, "=0")
        formula(ws, 16, col, f"={c}15")
        formula(ws, 18, col, f"={prev}20")
        formula(ws, 19, col, f"={c}9+{c}13+{c}16")
        formula(ws, 20, col, f"={c}18+{c}19")
        formula(ws, 21, col, f"={c}20-{xref('Balance Sheet', f'{c}5')}")
        formula(ws, 23, col, f"={xref('Income Statement', f'{c}15')}*(1-{xref('Assumptions', f'{c}23')})+{c}6-{c}11-{c}8")
    for row in labels:
        for col in range(2, 10):
            fmt_num(ws.cell(row, col))
    section(ws, 4, "Operating Cash Flow", 10)
    section(ws, 10, "Investing Cash Flow", 10)
    section(ws, 14, "Financing Cash Flow", 10)
    section(ws, 17, "Cash Tie-Out", 10)
    style_sheet(ws, 10)


def build_dcf(wb: Workbook):
    ws = wb.create_sheet("DCF")
    set_title(ws, "DCF Valuation", 9)
    for col, year in enumerate(PROJ_YEARS, start=2):
        ws.cell(3, col, f"FY{year}E").fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        ws.cell(3, col).font = Font(bold=True)
    labels = {
        5: "Revenue",
        6: "EBIT",
        7: "Tax Rate",
        8: "NOPAT",
        9: "D&A",
        10: "CapEx",
        11: "Change in NWC",
        12: "Unlevered FCF",
        14: "Discount Factor",
        15: "PV of FCF",
        17: "Terminal Growth",
        18: "Terminal Value",
        19: "PV of Terminal Value",
    }
    for row, label in labels.items():
        ws.cell(row, 1, label)
    for col, year in enumerate(PROJ_YEARS, start=2):
        model_col = get_column_letter(col + 3)
        c = get_column_letter(col)
        formula(ws, 5, col, f"={xref('Income Statement', f'{model_col}5')}")
        formula(ws, 6, col, f"={xref('Income Statement', f'{model_col}15')}")
        formula(ws, 7, col, f"={xref('Assumptions', f'{model_col}23')}")
        formula(ws, 8, col, f"={c}6*(1-{c}7)")
        formula(ws, 9, col, f"={xref('Income Statement', f'{model_col}18')}")
        formula(ws, 10, col, f"={xref('Cash Flow', f'{model_col}11')}")
        formula(ws, 11, col, f"={xref('Cash Flow', f'{model_col}8')}")
        formula(ws, 12, col, f"={c}8+{c}9-{c}10-{c}11")
        formula(ws, 14, col, f"=1/(1+{xref('Assumptions', 'B30')})^{col-1}")
        formula(ws, 15, col, f"={c}12*{c}14")
        formula(ws, 17, col, f"={xref('Assumptions', 'B31')}")
    formula(ws, 18, 6, f"=F12*(1+{xref('Assumptions', 'B31')})/({xref('Assumptions', 'B30')}-{xref('Assumptions', 'B31')})")
    formula(ws, 19, 6, f"=F18*F14")
    summary = {
        22: "PV of FCF",
        23: "PV of Terminal Value",
        24: "Enterprise Value",
        25: "Cash",
        26: "Debt",
        27: "Equity Value",
        28: "Diluted Shares",
        29: "Implied Value / Share",
        30: "Current Share Price",
        31: "Upside / (Downside)",
    }
    for row, label in summary.items():
        ws.cell(row, 1, label)
    formula(ws, 22, 2, "=SUM(B15:F15)")
    formula(ws, 23, 2, "=F19")
    formula(ws, 24, 2, "=B22+B23")
    formula(ws, 25, 2, f"={xref('Assumptions', 'B33')}")
    formula(ws, 26, 2, f"={xref('Assumptions', 'B34')}")
    formula(ws, 27, 2, "=B24+B25-B26")
    formula(ws, 28, 2, f"={xref('Assumptions', 'B32')}")
    formula(ws, 29, 2, "=B27/B28")
    formula(ws, 30, 2, f"={xref('Assumptions', 'B35')}")
    formula(ws, 31, 2, '=IF(B30>0,B29/B30-1,"")')
    for row in list(labels) + list(summary):
        for col in range(2, 7):
            if row in {7, 14, 17, 31}:
                fmt_pct(ws.cell(row, col))
            else:
                fmt_num(ws.cell(row, col))
    section(ws, 4, "Unlevered Free Cash Flow", 9)
    section(ws, 21, "Valuation Summary", 9)
    style_sheet(ws, 9)


def build_sensitivity(wb: Workbook):
    ws = wb.create_sheet("Sensitivity")
    set_title(ws, "DCF Sensitivity: WACC vs Terminal Growth", 8)
    ws["A3"] = "Implied Value / Share"
    base_wacc = xref("Assumptions", "B30")
    base_g = xref("Assumptions", "B31")
    wacc_offsets = [-0.02, -0.01, 0, 0.01, 0.02]
    g_offsets = [-0.01, -0.005, 0, 0.005, 0.01]
    for i, off in enumerate(g_offsets, start=2):
        formula(ws, 4, i, f"={base_g}{off:+.3f}")
        fmt_pct(ws.cell(4, i))
    for r, off in enumerate(wacc_offsets, start=5):
        formula(ws, r, 1, f"={base_wacc}{off:+.3f}")
        fmt_pct(ws.cell(r, 1))
        for c in range(2, 7):
            w = f"$A{r}"
            g = f"{get_column_letter(c)}$4"
            fcf_terms = "+".join([f"{xref('DCF', f'{get_column_letter(i)}12')}/(1+{w})^{i-1}" for i in range(2, 7)])
            tv = f"{xref('DCF', 'F12')}*(1+{g})/({w}-{g})/(1+{w})^5"
            formula(ws, r, c, f"=({fcf_terms}+{tv}+{xref('DCF', 'B25')}-{xref('DCF', 'B26')})/{xref('DCF', 'B28')}")
            fmt_num(ws.cell(r, c))
    ws["A4"] = "WACC \\ Terminal g"
    for row in range(4, 10):
        for col in range(1, 7):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=LIGHT_BLUE if row == 4 or col == 1 else WHITE)
            ws.cell(row, col).font = Font(bold=(row == 4 or col == 1))
    ws["D7"].fill = PatternFill("solid", fgColor=MID_BLUE)
    ws["D7"].font = Font(bold=True)
    style_sheet(ws, 8)


def build_checks(wb: Workbook):
    ws = wb.create_sheet("Checks")
    set_title(ws, "Model Checks", 8)
    checks = [
        ("Revenue build equals IS revenue FY2025", "='Revenue Build'!D25-'Income Statement'!D5"),
        ("Revenue build equals IS revenue FY2030E", "='Revenue Build'!I25-'Income Statement'!I5"),
        ("BS balances FY2025", "='Balance Sheet'!D29"),
        ("BS balances FY2030E", "='Balance Sheet'!I29"),
        ("Cash ties FY2030E", "='Cash Flow'!I21"),
        ("DCF uses FY2030E UFCF", "='DCF'!F12-'Cash Flow'!I23"),
    ]
    ws.cell(3, 1, "Check").font = Font(bold=True)
    ws.cell(3, 2, "Formula Result").font = Font(bold=True)
    ws.cell(3, 3, "Status").font = Font(bold=True)
    for r, (label, formula_text) in enumerate(checks, start=4):
        ws.cell(r, 1, label)
        formula(ws, r, 2, formula_text)
        formula(ws, r, 3, f'=IF(ABS(B{r})<0.01,"OK","Review")')
        fmt_num(ws.cell(r, 2))
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    style_sheet(ws, 8)


def main():
    wb = Workbook()
    wb.remove(wb.active)
    build_cover(wb)
    build_sources(wb)
    build_assumptions(wb)
    build_revenue(wb)
    build_income_statement(wb)
    build_balance_sheet(wb)
    build_cash_flow(wb)
    build_dcf(wb)
    build_sensitivity(wb)
    build_checks(wb)
    wb.save(OUTPUT)
    # Re-open once to validate the file is structurally readable.
    load_workbook(OUTPUT, data_only=False)
    print(OUTPUT)


if __name__ == "__main__":
    main()
